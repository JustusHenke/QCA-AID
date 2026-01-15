"""
Ergebnis-Exporter für QCA-AID
==============================
Exportiert Analyseergebnisse in verschiedene Formate (Excel, PDF, etc.).
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from collections import defaultdict, Counter

# Fix für Unicode-Encoding auf Windows-Konsolen
if sys.platform == 'win32':
    try:
        # Setze stdout und stderr auf UTF-8
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        # Fallback für ältere Python-Versionen
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..core.config import CONFIG
from ..core.data_models import CategoryDefinition, CodingResult, CategoryChange
from ..analysis.inductive_coding import InductiveCoder
from ..quality.reliability import ReliabilityCalculator
from ..utils.export.helpers import (
    sanitize_text_for_excel, generate_pastel_colors, format_confidence
)
from ..utils.impact_analysis import export_multiple_coding_report

try:
    from ..utils.export.pdf_annotator import PDFAnnotator
    from ..utils.export.converters import DocumentToPDFConverter
    pdf_annotation_available = True
except ImportError:
    pdf_annotation_available = False


class ResultsExporter:
    """
    Exports the results of the analysis in various formats (JSON, CSV, Excel, Markdown).
    Supports the documentation of coded chunks and the category system.
    """
    
    def __init__(self, output_dir: str, attribute_labels: Dict[str, str], inductive_coder: InductiveCoder = None,
                 analysis_manager: 'IntegratedAnalysisManager' = None):
        self.output_dir = output_dir
        self.attribute_labels = attribute_labels
        self.inductive_coder = inductive_coder
        self.analysis_manager = analysis_manager
        self.relevance_checker = analysis_manager.relevance_checker if analysis_manager else None
        self.category_colors = {}
        self.current_categories = {}  # Wird von main() gesetzt
        # FIX: Cache für Segment-Texte zur besseren Text-Extraktion
        self._segment_text_cache = {}

        os.makedirs(output_dir, exist_ok=True)

    def cache_segment_texts(self, codings: List[Dict], chunks: Dict[str, List[str]] = None) -> None:
        """
        FIX: Erstelle Cache für Segment-Texte zur besseren Text-Extraktion bei Mehrfachkodierung.
        Nutzt sowohl codings als auch chunks für vollständige Text-Abdeckung.
        """
        # FIX: Erste Pass - sammle alle verfügbaren Texte aus codings
        available_texts = {}
        
        for coding in codings:
            segment_id = coding.get('segment_id', '')
            
            # FIX: Prüfe ob Display-IDs als segment_id verwendet werden - das sollte NICHT passieren!
            if segment_id and not '_chunk_' in segment_id and '-' in segment_id:
                print(f"   ❌ CRITICAL ERROR: Display-ID als segment_id im Cache: {segment_id}")
                print(f"   ❌ Das ist ein Bug - Display-IDs sollen niemals intern verwendet werden!")
                print(f"   ❌ Quelle des Problems muss gefunden und behoben werden!")
                # Überspringe diese Kodierung um den Fehler sichtbar zu machen
                continue
            
            # FIX: Robuste Text-Extraktion mit derselben Logik wie im Export
            text = coding.get('text', '')
            if not text and 'result' in coding:
                text = coding['result'].get('text', '')
            
            if text and segment_id:
                available_texts[segment_id] = text
                
                # Auch original_segment_id sammeln falls verfügbar
                original_id = coding.get('original_segment_id', '')
                if original_id and original_id != segment_id:
                    available_texts[original_id] = text
                
                # Auch aus result-Feld sammeln
                if 'result' in coding:
                    result_original_id = coding['result'].get('original_segment_id', '')
                    if result_original_id and result_original_id != segment_id:
                        available_texts[result_original_id] = text
        
        # FIX: NEUE LOGIK - Nutze chunks für vollständige Text-Abdeckung
        if chunks:
            print(f"   📝 Ergänze Cache mit {len(chunks)} Dokumenten aus chunks...")
            from ..core.segment_id_manager import SegmentIDManager
            
            chunks_added = 0
            for doc_name, doc_chunks in chunks.items():
                for chunk_idx, chunk_text in enumerate(doc_chunks):
                    # Erstelle segment_id für diesen Chunk
                    segment_id = SegmentIDManager.create_segment_id(doc_name, chunk_idx)
                    
                    # Füge zum Cache hinzu (überschreibt nicht vorhandene Texte)
                    if segment_id not in available_texts and chunk_text:
                        available_texts[segment_id] = chunk_text
                        chunks_added += 1
            
            print(f"   📝 {chunks_added} zusätzliche Texte aus chunks hinzugefügt")
        else:
            print(f"   ⚠️ Keine chunks verfügbar - Cache nur aus codings erstellt")
        
        # FIX: Zweite Pass - erstelle umfassenden Cache mit Fallback-Strategien
        for coding in codings:
            segment_id = coding.get('segment_id', '')
            if not segment_id:
                continue
                
            # Versuche Text für dieses segment_id zu finden
            text = None
            
            # Strategie 1: Direkter Text aus diesem coding
            text = coding.get('text', '')
            if not text and 'result' in coding:
                text = coding['result'].get('text', '')
            
            # Strategie 2: Suche in available_texts
            if not text:
                text = available_texts.get(segment_id, '')
            
            # Strategie 3: Für Mehrfachkodierung - suche Base-ID
            if not text and '-' in segment_id:
                # Entferne Mehrfachkodierungs-Suffix: "document.ext_chunk_1-1" -> "document.ext_chunk_1"
                base_id = segment_id.rsplit('-', 1)[0]  # Entferne nur das letzte Suffix
                text = available_texts.get(base_id, '')
                if text:
                    # print(f"   📝 Found text for {segment_id} via base_id: {base_id}")
                    pass
            
            # Strategie 4: Suche über original_segment_id
            if not text:
                original_id = coding.get('original_segment_id', '')
                if original_id:
                    text = available_texts.get(original_id, '')
                    if text:
                        # print(f"   📝 Found text for {segment_id} via original_segment_id: {original_id}")
                        pass
                    
                    # Auch Base-ID von original_segment_id versuchen
                    if not text and '-' in original_id:
                        original_base_id = original_id.rsplit('-', 1)[0]
                        text = available_texts.get(original_base_id, '')
                        if text:
                            # print(f"   📝 Found text for {segment_id} via original_base_id: {original_base_id}")
                            pass
            
            # Cache den gefundenen Text unter verschiedenen IDs
            if text:
                # FIX: Use standardized segment IDs for consistent caching
                standardized_id = self._standardize_segment_id(segment_id)
                
                # Cache under standardized ID
                self._segment_text_cache[standardized_id] = text
                
                # Also cache under original segment_id if different
                if segment_id != standardized_id:
                    self._segment_text_cache[segment_id] = text
                
                # Also cache under original_segment_id if available
                original_id = coding.get('original_segment_id', '')
                if original_id and original_id != segment_id and original_id != standardized_id:
                    self._segment_text_cache[original_id] = text
                
                # For multiple coding, also cache under base ID
                if '-' in segment_id:
                    base_id = segment_id.rsplit('-', 1)[0]
                    if base_id not in self._segment_text_cache:
                        self._segment_text_cache[base_id] = text
                
                # Also cache under original segment_id for backward compatibility
                if segment_id != standardized_id:
                    self._segment_text_cache[segment_id] = text
                
                # Cache under base ID for multiple coding lookups
                if '-' in standardized_id:
                    base_id = standardized_id.rsplit('-', 1)[0]
                    self._segment_text_cache[base_id] = text
                
                # Cache under original_segment_id if different
                original_id = coding.get('original_segment_id', '')
                if original_id and original_id != segment_id:
                    standardized_original = self._standardize_segment_id(original_id)
                    self._segment_text_cache[standardized_original] = text
                    if original_id != standardized_original:
                        self._segment_text_cache[original_id] = text
                
                # print(f"   📝 Cached text for {standardized_id} (length: {len(text)})")
                pass
            else:
                print(f"   ⚠️ No text found for {segment_id} during cache population")

    def _get_consensus_coding(self, segment_codes: List[Dict]) -> Dict:
        """
        FIX: Einheitliche Consensus-Bildung mit robuster Subkategorien-Validierung
        FÜr IntegratedAnalysisManager Klasse
        """
        if not segment_codes:
            return {}

        # FIX: Debug-Info Über Eingabe-Kodierungen
        categories = [coding.get('category', 'UNKNOWN') for coding in segment_codes]
        unique_categories = list(set(categories))
        
        print(f"🕵️ DEBUG Consensus: {len(segment_codes)} Kodierungen, Kategorien: {unique_categories}")
        
        # FIX: Detaillierte Analyse der Subkategorien VOR Consensus
        for i, coding in enumerate(segment_codes):
            cat = coding.get('category', 'UNKNOWN')
            subcats = coding.get('subcategories', [])
            print(f"   Kodierung {i+1}: {cat} -> {subcats}")
        
        # Wenn alle dieselbe Hauptkategorie haben, normale Konsensbildung
        if len(unique_categories) == 1:
            return self._get_single_consensus_coding(segment_codes)
        
        # Mehrfachkodierung: Erstelle prÄzises Kategorie-Subkategorie-Mapping
        print(f"ðŸ”€ Mehrfachkodierung erkannt mit Kategorien: {unique_categories}")
        
        best_coding = None
        highest_confidence = 0
                
        for coding in segment_codes:
            category = coding.get('category', '')
            subcats = coding.get('subcategories', [])
            confidence = self._extract_confidence_value(coding)
            
            print(f"   Prüfe Kodierung: {category} (Subkat: {len(subcats)}, Konfidenz: {confidence:.2f})")
                                   
            # Globale beste Kodierung
            if confidence > highest_confidence:
                highest_confidence = confidence
                best_coding = coding

        if best_coding:
            consensus_coding = best_coding.copy()
            
            # FIX: Konsistente Subkategorien-Behandlung mit detailliertem Logging
            main_category = consensus_coding.get('category', '')
            original_subcats = best_coding.get('subcategories', [])
            
            print(f"🎯 Beste Kodierung gewÄhlt: {main_category}")
            print(f"   Original Subkategorien: {original_subcats}")
            
            # FIX: Verwende IMMER die robuste CategoryValidator-Methode
            try:
                validated_subcats = CategoryValidator.validate_subcategories_for_category(
                    original_subcats, main_category, self.current_categories, warn_only=False
                )
                
                print(f"   Nach Validierung: {validated_subcats}")
                
                # FIX: Dokumentiere Validierungsaktionen
                removed_subcats = set(original_subcats) - set(validated_subcats)
                if removed_subcats:
                    print(f"   🔧 ENTFERNT: {removed_subcats}")
                    # FIX: Füge Validierungs-Info zur Begründung hinzu
                    original_justification = consensus_coding.get('justification', '')
                    consensus_coding['justification'] = f"{original_justification} [FIX: Subkategorien-Validierung entfernte: {list(removed_subcats)}]"
                
                consensus_coding['subcategories'] = validated_subcats
                
            except Exception as e:
                print(f"❌ FEHLER bei Subkategorien-Validierung: {str(e)}")
                print(f"   Fallback: Verwende ursprüngliche Subkategorien ohne Validierung")
                consensus_coding['subcategories'] = original_subcats
            
            # FIX: Füge Validierungs-Metadaten hinzu
            consensus_coding['validation_applied'] = True
            consensus_coding['original_subcategory_count'] = len(original_subcats)
            consensus_coding['validated_subcategory_count'] = len(consensus_coding['subcategories'])
            
            return consensus_coding
        
        # Fallback: Erste Kodierung verwenden
        print("❌ FALLBACK: Verwende erste verfügbare Kodierung")
        fallback_coding = segment_codes[0] if segment_codes else {}
        
        # FIX: Auch Fallback-Kodierung validieren
        if fallback_coding:
            try:
                main_cat = fallback_coding.get('category', '')
                orig_subcats = fallback_coding.get('subcategories', [])
                
                validated_subcats = CategoryValidator.validate_subcategories_for_category(
                    orig_subcats, main_cat, self.current_categories, warn_only=False
                )
                
                fallback_coding['subcategories'] = validated_subcats
                fallback_coding['validation_applied'] = True
                fallback_coding['is_fallback_coding'] = True
                
            except Exception as e:
                print(f"❌ Fallback-Validierung fehlgeschlagen: {str(e)}")
        
        return fallback_coding

    def _get_majority_coding(self, segment_codes: List[Dict]) -> Optional[Dict]:
        """
        VEREINFACHT: Nutzt dieselbe Logik wie Schlüsselwörter - nimmt aus bester Kodierung
        """
        if not segment_codes:
            return None

        print(f"\nMehrheitsentscheidung fuer Segment mit {len(segment_codes)} Kodierungen...")

        # 1. ZÄhle Hauptkategorien
        category_counts = Counter(coding['category'] for coding in segment_codes)
        total_coders = len(segment_codes)
        
        # Finde hÄufigste Hauptkategorie(n)
        max_count = max(category_counts.values())
        majority_categories = [
            category for category, count in category_counts.items()
            if count == max_count
        ]
        
        print(f"  Kategorieverteilung: {dict(category_counts)}")
        print(f"  Häufigste Kategorie(n): {majority_categories} ({max_count}/{total_coders})")
        
        # 2. Bei eindeutiger Mehrheit
        if len(majority_categories) == 1:
            majority_category = majority_categories[0]
            print(f"  ✅ Eindeutige Mehrheit fuer: '{majority_category}'")
        else:
            # 3. Bei Gleichstand: WÄhle nach hÖchster Konfidenz
            print(f"  Gleichstand zwischen {len(majority_categories)} Kategorien")
            
            # Sammle Kodierungen fuer die gleichstehenden Kategorien
            tied_codings = [
                coding for coding in segment_codes
                if coding['category'] in majority_categories
            ]
            
            # Finde die Kodierung mit der hÖchsten Konfidenz
            highest_confidence = -1
            best_coding = None
            
            for coding in tied_codings:
                # Extrahiere Konfidenzwert
                confidence = 0.0
                if isinstance(coding.get('confidence'), dict):
                    confidence = float(coding['confidence'].get('total', 0))
                    if confidence == 0:  # Fallback auf category-Konfidenz
                        confidence = float(coding['confidence'].get('category', 0))
                elif isinstance(coding.get('confidence'), (int, float)):
                    confidence = float(coding['confidence'])
                
                if confidence > highest_confidence:
                    highest_confidence = confidence
                    best_coding = coding
                    majority_category = coding['category']
            
            print(f"  ✅ Tie-Breaking durch Konfidenz: '{majority_category}' (Konfidenz: {highest_confidence:.2f})")
        
        # 4. Sammle alle Kodierungen fuer die gewÄhlte Mehrheitskategorie
        matching_codings = [
            coding for coding in segment_codes
            if coding['category'] == majority_category
        ]
        
        # VEREINFACHT: WÄhle beste Kodierung und nutze ihre Subkategorien direkt
        base_coding = max(
            matching_codings,
            key=lambda x: self._extract_confidence_value(x)
        )
        
        # VEREINFACHT: Keine komplexe Subkategorien-Sammlung
        majority_coding = base_coding.copy()
        main_category = majority_coding.get('category', '')
        original_subcats = base_coding.get('subcategories', [])
        validated_subcats = CategoryValidator.validate_subcategories_for_category(
                original_subcats, main_category, self.current_categories, warn_only=False
            )
        majority_coding['subcategories'] = validated_subcats
        
        # Kombiniere Begründungen (bleibt gleich)
        all_justifications = []
        for coding in matching_codings:
            justification = coding.get('justification', '')
            if justification and justification not in all_justifications:
                all_justifications.append(justification)  
        
        if all_justifications:
            majority_coding['justification'] = f"[Mehrheit aus {len(matching_codings)} Kodierern] " + " | ".join(all_justifications[:3])
        
        # Rest der Dokumentation bleibt gleich
        majority_coding['consensus_info'] = {
            'total_coders': total_coders,
            'category_votes': max_count,
            'category_agreement': max_count / total_coders,
            'tied_categories': majority_categories if len(majority_categories) > 1 else [],
            'source_codings': len(matching_codings),
            'selection_type': 'majority',
            'tie_broken_by_confidence': len(majority_categories) > 1
        }
        
        print(f"  ✅ Mehrheits-Kodierung erstellt: '{majority_category}' mit {len(majority_coding['subcategories'])} Subkategorien direkt Übernommen: {', '.join(majority_coding['subcategories'])}")
        
        return majority_coding


    def _get_single_consensus_coding(self, segment_codes: List[Dict]) -> Optional[Dict]:
        """
        Ermittelt die Konsens-Kodierung fuer ein Segment basierend auf einem mehrstufigen Prozess.
        KORRIGIERT: PrÄzise Subkategorien-Zuordnung ohne Vermischung zwischen Hauptkategorien
        """
        if not segment_codes:
            return None

        # 1. Analyse der Hauptkategorien
        category_counts = Counter(coding['category'] for coding in segment_codes)
        total_coders = len(segment_codes)
        
        # Finde hÄufigste Hauptkategorie(n)
        max_count = max(category_counts.values())
        majority_categories = [
            category for category, count in category_counts.items()
            if count == max_count
        ]
        
        # Prüfe ob es eine klare Mehrheit gibt (>50%)
        if max_count <= total_coders / 2:
            print(f"Keine Mehrheit fuer Hauptkategorie gefunden: {dict(category_counts)}")
            
            # Suche nach Kodierung mit hÖchster Konfidenz
            highest_confidence = -1
            best_coding = None
            
            for coding in segment_codes:
                confidence = self._extract_confidence_value(coding)
                
                if confidence > highest_confidence:
                    highest_confidence = confidence
                    best_coding = coding
            
            # Minimalschwelle fuer Konfidenz (kann angepasst werden)
            confidence_threshold = 0.7
            
            if highest_confidence >= confidence_threshold:
                # Verwende die Kodierung mit der hÖchsten Konfidenz
                result_coding = best_coding.copy()
                
                # KORRIGIERT: Behalte nur Subkategorien der gewÄhlten Hauptkategorie
                result_coding['subcategories'] = best_coding.get('subcategories', [])
                
                # Füge Hinweis zur konfidenzbedingten Auswahl hinzu
                result_coding['justification'] = (f"[Konfidenzbasierte Auswahl: {highest_confidence:.2f}] " + 
                                                result_coding.get('justification', ''))
                
                # Dokumentiere den Konsensprozess
                result_coding['consensus_info'] = {
                    'total_coders': total_coders,
                    'category_distribution': dict(category_counts),
                    'max_agreement': max_count / total_coders,
                    'selection_type': 'confidence_based',
                    'confidence': highest_confidence
                }
                
                print(f"  Konfidenzbasierte Auswahl: '{result_coding['category']}' mit {len(result_coding.get('subcategories', []))} Subkategorien")
                return result_coding
            else:
                # Erstelle "Kein Kodierkonsens"-Eintrag
                base_coding = segment_codes[0].copy()
                base_coding['category'] = "Kein Kodierkonsens"
                base_coding['subcategories'] = []  # Keine Subkategorien bei fehlendem Konsens
                base_coding['justification'] = (f"Keine Mehrheit unter den Kodierern und keine Kodierung " +
                                            f"mit ausreichender Konfidenz (max: {highest_confidence:.2f}). " +
                                            f"Kategorien: {', '.join(category_counts.keys())}")
                
                # Dokumentiere den Konsensprozess
                base_coding['consensus_info'] = {
                    'total_coders': total_coders,
                    'category_distribution': dict(category_counts),
                    'max_agreement': max_count / total_coders,
                    'selection_type': 'no_consensus',
                    'max_confidence': highest_confidence
                }
                
                print(f"  Kein Konsens und unzureichende Konfidenz (max: {highest_confidence:.2f})")
                return base_coding

        # 2. Wenn es mehrere gleichhÄufige Hauptkategorien gibt, verwende Tie-Breaking
        if len(majority_categories) > 1:
            print(f"Gleichstand zwischen Kategorien: {majority_categories}")
            # Sammle alle Kodierungen fuer die Mehrheitskategorien
            candidate_codings = [
                coding for coding in segment_codes
                if coding['category'] in majority_categories
            ]
            
            # WÄhle basierend auf hÖchster durchschnittlicher Konfidenz
            highest_avg_confidence = -1
            selected_category = None
            
            for category in majority_categories:
                category_codings = [c for c in candidate_codings if c['category'] == category]
                total_confidence = 0.0
                
                for coding in category_codings:
                    confidence = self._extract_confidence_value(coding)
                    total_confidence += confidence
                    
                avg_confidence = total_confidence / len(category_codings) if category_codings else 0
                
                if avg_confidence > highest_avg_confidence:
                    highest_avg_confidence = avg_confidence
                    selected_category = category
                    
            majority_category = selected_category
            print(f"  Kategorie '{majority_category}' durch hÖchste Konfidenz ({highest_avg_confidence:.2f}) gewÄhlt")
        else:
            majority_category = majority_categories[0]

        # 3. KORRIGIERT: Sammle nur Kodierungen fuer die gewÄhlte Mehrheitskategorie
        matching_codings = [
            coding for coding in segment_codes
            if coding['category'] == majority_category
        ]
        
        # VEREINFACHT: WÄhle beste Kodierung und nutze ihre Subkategorien direkt
        base_coding = max(
            matching_codings,
            key=lambda x: self._extract_confidence_value(x)
        )
        
        # VEREINFACHT: Keine komplexe Subkategorien-Sammlung mehr
        consensus_coding = base_coding.copy()
        main_category = consensus_coding.get('category', '')
        original_subcats = base_coding.get('subcategories', [])
        # FIX: Prüfe ob wir Zugriff auf das vollständige Kategoriensystem haben
        categories_for_validation = getattr(self, 'current_categories', {})
        
        if categories_for_validation and main_category in categories_for_validation:
            validated_subcats = CategoryValidator.validate_subcategories_for_category(
                original_subcats, main_category, categories_for_validation, warn_only=False
            )
            
            # FIX: Debug-Ausgabe der Validierung
            if len(original_subcats) != len(validated_subcats):
                removed_subcats = set(original_subcats) - set(validated_subcats)
                print(f"🔧 FIX: Consensus-Validierung entfernte {len(removed_subcats)} Subkategorien: {removed_subcats}")
                
                # FIX: Dokumentiere Änderung in Begründung
                original_justification = consensus_coding.get('justification', '')
                consensus_coding['justification'] = f"{original_justification} [FIX: Consensus-Validierung entfernte ungültige Subkategorien: {list(removed_subcats)}]"
                
        else:
            # FIX: Fallback ohne Validierung
            print(f"❌ WARNUNG: Keine Kategorie-Validierung mÖglich fuer '{main_category}' - verwende ursprüngliche Subkategorien")
            validated_subcats = original_subcats
        
        consensus_coding['subcategories'] = validated_subcats
        
        # Kombiniere nur Begründungen der matching codings
        all_justifications = []
        for coding in matching_codings:
            justification = coding.get('justification', '')
            if justification and justification not in all_justifications:
                all_justifications.append(justification)
        
        if all_justifications:
            consensus_coding['justification'] = f"[Konsens aus {len(matching_codings)} Kodierern] " + " | ".join(all_justifications[:3])
        
        consensus_coding['consensus_info'] = {
            'total_coders': total_coders,
            'category_agreement': max_count / total_coders,
            'source_codings': len(matching_codings),
            'selection_type': 'consensus'
        }
        
        print(f"\nKonsens-Kodierung erstellt:")
        print(f"- Hauptkategorie: {consensus_coding['category']} ({max_count}/{total_coders} Kodierer)")
        print(f"- Subkategorien: {len(consensus_coding['subcategories'])} direkt Übernommen: {', '.join(consensus_coding['subcategories'])}")
        print(f"- Übereinstimmung: {(max_count/total_coders)*100:.1f}%")
        
        return consensus_coding

    def _create_category_specific_codings(self, segment_codes: List[Dict], segment_id: str) -> List[Dict]:
        """
        KORRIGIERT: PrÄzise Subkategorien-Zuordnung OHNE Mehrfachkodierung zu verhindern
        """
        # Gruppiere Kodierungen nach Hauptkategorien
        category_groups = {}
        
        for coding in segment_codes:
            main_cat = coding.get('category', '')
            if main_cat and main_cat not in ['Nicht kodiert', 'Kein Kodierkonsens']:
                if main_cat not in category_groups:
                    category_groups[main_cat] = []
                category_groups[main_cat].append(coding)
        
        result_codings = []
        
        for i, (main_cat, codings_for_cat) in enumerate(category_groups.items(), 1):
            # print(f"DEBUG: Verarbeite Hauptkategorie '{main_cat}' mit {len(codings_for_cat)} Kodierungen")
            
            # WÄhle die beste Kodierung fuer diese Kategorie als Basis
            best_coding = max(codings_for_cat, key=lambda x: self._extract_confidence_value(x))
            
            # KRITISCH: Sammle NUR Subkategorien, die fuer DIESE Hauptkategorie kodiert wurden
            relevant_subcats = []
            
            for coding in codings_for_cat:
                # Prüfe ob diese Kodierung wirklich fuer die aktuelle Hauptkategorie ist
                if coding.get('category') == main_cat:
                    # Diese Kodierung gehört zu dieser Hauptkategorie
                    subcats = coding.get('subcategories', [])
                    if isinstance(subcats, (list, tuple)):
                        relevant_subcats.extend(subcats)
                    elif isinstance(subcats, str) and subcats:
                        subcat_list = [s.strip() for s in subcats.split(',') if s.strip()]
                        relevant_subcats.extend(subcat_list)
            
            # Entferne Duplikate
            final_subcats = list(set(relevant_subcats))
            
            # OPTIONAL: Validiere gegen Kategoriensystem (aber nur als Warnung)
            if hasattr(self, 'current_categories') and main_cat in self.current_categories:
                valid_subcats_for_main = set(self.current_categories[main_cat].subcategories.keys())
                invalid_subcats = [sub for sub in final_subcats if sub not in valid_subcats_for_main]
                
                if invalid_subcats:
                    print(f"  WARNUNG: Ungültige Subkategorien fuer '{main_cat}' gefunden: {invalid_subcats}")
                    print(f"  GÜltige Subkategorien: {list(valid_subcats_for_main)}")
                    # NICHT entfernen, nur warnen!
            
            # print(f"  Finale Subkategorien fuer '{main_cat}': {final_subcats}")
            
            # Erstelle konsolidierte Kodierung
            consolidated_coding = best_coding.copy()
            consolidated_coding['category'] = main_cat
            consolidated_coding['subcategories'] = final_subcats  # Nur relevante Subkategorien
            consolidated_coding['multiple_coding_instance'] = i
            consolidated_coding['total_coding_instances'] = len(category_groups)
            
            # Erweiterte Begründung
            original_justification = consolidated_coding.get('justification', '')
            consolidated_coding['justification'] = f"[Mehrfachkodierung - Kategorie {i}/{len(category_groups)}] {original_justification}"
            
            result_codings.append(consolidated_coding)
        
        # print(f"DEBUG: Erstellt {len(result_codings)} kategorie-spezifische Kodierungen fuer {segment_id}")
        return result_codings
   
   
    # Zusätzliche Methode fuer ResultsExporter Klasse
    def debug_export_process(self, codings: List[Dict]) -> None:
        """
        Ae–ffentliche Debug-Methode fuer Export-Prozess
        Kann vor dem eigentlichen Export aufgerufen werden
        """
        print(f"\n🕵️ STARTE EXPORT-DEBUG fuer {len(codings)} Kodierungen")
        self._debug_export_preparation(codings)
        
        # Zusätzliche Checks
        segments_with_issues = []
        
        for coding in codings:
            segment_id = coding.get('segment_id', '')
            category = coding.get('category', '')
            subcats = coding.get('subcategories', [])
            
            # Prüfe auf leere Subkategorien bei kategorisierten Segmenten
            if category and category not in ['Nicht kodiert', 'Kein Kodierkonsens']:
                if not subcats or (isinstance(subcats, list) and len(subcats) == 0):
                    segments_with_issues.append({
                        'segment_id': segment_id,
                        'category': category,
                        'issue': 'Keine Subkategorien trotz Kategorisierung'
                    })
        
        if segments_with_issues:
            print(f"\nâš  GEFUNDENE PROBLEME: {len(segments_with_issues)} Segmente mit fehlenden Subkategorien")
            for issue in segments_with_issues[:3]:
                print(f"  - {issue['segment_id']}: {issue['category']} -> {issue['issue']}")
            if len(segments_with_issues) > 3:
                print(f"  ... und {len(segments_with_issues) - 3} weitere")
        else:
            print(f"\n✅ Keine offensichtlichen Subkategorien-Probleme gefunden")
        
        print(f"\n🕵️ EXPORT-DEBUG ABGESCHLOSSEN")

    def _extract_confidence_value(self, coding: Dict) -> float:
        """
        Hilfsmethode zum Extrahieren des Konfidenzwerts aus einer Kodierung.
        
        Args:
            coding: Kodierung mit Konfidenzinformation
            
        Returns:
            float: Konfidenzwert zwischen 0 und 1
        """
        try:
            if isinstance(coding.get('confidence'), dict):
                confidence = float(coding['confidence'].get('total', 0))
                if confidence == 0:  # Fallback auf category-Konfidenz
                    confidence = float(coding['confidence'].get('category', 0))
            elif isinstance(coding.get('confidence'), (int, float)):
                confidence = float(coding['confidence'])
            else:
                confidence = 0.0
            return confidence
        except (ValueError, TypeError):
            return 0.0

    def _get_manual_priority_coding(self, segment_codes: List[Dict]) -> Optional[Dict]:
        """
        Bevorzugt manuelle Kodierungen vor automatischen Kodierungen.
        KORRIGIERT: Subkategorien werden korrekt verarbeitet und zusammengefÜhrt.
        
        Args:
            segment_codes: Liste der Kodierungen fuer ein Segment von verschiedenen Kodierern
                
        Returns:
            Optional[Dict]: Priorisierte Kodierung mit korrekten Subkategorien
        """
        if not segment_codes:
            return None

        print(f"\nManuelle Priorisierung fuer Segment mit {len(segment_codes)} Kodierungen...")

        # 1. Trenne manuelle und automatische Kodierungen
        manual_codings = []
        auto_codings = []
        
        for coding in segment_codes:
            coder_id = coding.get('coder_id', '')
            # Erkenne manuelle Kodierungen anhand der coder_id
            if 'human' in coder_id.lower() or 'manual' in coder_id.lower() or coding.get('manual_coding', False):
                manual_codings.append(coding)
            else:
                auto_codings.append(coding)
        
        print(f"  Gefunden: {len(manual_codings)} manuelle, {len(auto_codings)} automatische Kodierungen")
        
        # 2. Wenn manuelle Kodierungen vorhanden sind, bevorzuge diese
        if manual_codings:
            print("  ✅ Verwende manuelle Kodierungen mit PrioritÄt")
            
            if len(manual_codings) == 1:
                # Nur eine manuelle Kodierung - verwende diese direkt
                selected_coding = manual_codings[0].copy()
                
                # WICHTIG: Subkategorien beibehalten!
                if 'subcategories' in manual_codings[0]:
                    selected_coding['subcategories'] = manual_codings[0]['subcategories']
                
                selected_coding['consensus_info'] = {
                    'total_coders': len(segment_codes),
                    'manual_coders': len(manual_codings),
                    'auto_coders': len(auto_codings),
                    'selection_type': 'single_manual',
                    'priority_reason': 'Einzige manuelle Kodierung verfügbar'
                }
                print(f"    Einzige manuelle Kodierung: '{selected_coding['category']}' mit {len(selected_coding.get('subcategories', []))} Subkategorien")
                
            else:
                # Mehrere manuelle Kodierungen - suche Konsens unter diesen
                print(f"    Suche Konsens unter {len(manual_codings)} manuellen Kodierungen")
                
                # Prüfe ob alle dieselbe Hauptkategorie haben
                manual_categories = [c['category'] for c in manual_codings]
                if len(set(manual_categories)) == 1:
                    # Alle haben dieselbe Hauptkategorie - konsolidiere Subkategorien
                    main_category = manual_categories[0]
                    
                    # Sammle alle Subkategorien von manuellen Kodierungen
                    all_manual_subcats = []
                    for coding in manual_codings:
                        subcats = coding.get('subcategories', [])
                        if isinstance(subcats, (list, tuple)):
                            all_manual_subcats.extend(subcats)
                        elif isinstance(subcats, str) and subcats:
                            subcats_list = [s.strip() for s in subcats.split(',') if s.strip()]
                            all_manual_subcats.extend(subcats_list)
                    
                    # Finde Konsens-Subkategorien (mindestens von der HÄlfte verwendet)
                    subcat_counts = Counter(all_manual_subcats)
                    min_votes = len(manual_codings) / 2
                    consensus_subcats = [
                        subcat for subcat, count in subcat_counts.items()
                        if count >= min_votes
                    ]
                    
                    # WÄhle beste manuelle Kodierung als Basis
                    selected_coding = max(
                        manual_codings,
                        key=lambda x: self._extract_confidence_value(x)
                    ).copy()
                    
                    # Setze konsolidierte Subkategorien
                    selected_coding['subcategories'] = consensus_subcats
                    
                    # Kombiniere Begründungen
                    manual_justifications = [c.get('justification', '')[:100] for c in manual_codings if c.get('justification', '')]
                    if manual_justifications:
                        selected_coding['justification'] = f"[Konsens aus {len(manual_codings)} manuellen Kodierungen] " + " | ".join(manual_justifications[:3])
                    
                    selected_coding['consensus_info'] = {
                        'total_coders': len(segment_codes),
                        'manual_coders': len(manual_codings),
                        'auto_coders': len(auto_codings),
                        'selection_type': 'manual_consensus',
                        'priority_reason': 'Konsens unter manuellen Kodierungen',
                        'subcategory_distribution': dict(subcat_counts)
                    }
                    print(f"    Konsens bei manuellen Kodierungen: '{selected_coding['category']}' mit {len(consensus_subcats)} Subkategorien: {', '.join(consensus_subcats)}")
                else:
                    # Verschiedene Hauptkategorien - wÄhle nach Konfidenz
                    selected_coding = max(
                        manual_codings,
                        key=lambda x: self._extract_confidence_value(x)
                    ).copy()

                    # VEREINFACHT: Direkte Übernahme
                    # selected_coding['subcategories'] = selected_coding.get('subcategories', [])  # DIREKT
                    main_category = selected_coding.get('category', '')
                    original_subcats = selected_coding.get('subcategories', [])
                    validated_subcats = CategoryValidator.validate_subcategories_for_category(
                        original_subcats, main_category, self.current_categories, warn_only=False
                    )
                    selected_coding['subcategories'] = validated_subcats
    
        else:
            # 3. Keine manuellen Kodierungen - verwende automatische mit Konsens
            print("  Keine manuellen Kodierungen - verwende automatische Kodierungen")
            
            # Verwende die bestehende Konsens-Logik fuer automatische Kodierungen
            consensus_coding = self._get_consensus_coding(auto_codings)
            
            
            if consensus_coding:
                selected_coding = consensus_coding.copy()
                # Aktualisiere consensus_info
                selected_coding['consensus_info'] = selected_coding.get('consensus_info', {})
                selected_coding['consensus_info'].update({
                    'total_coders': len(segment_codes),
                    'manual_coders': 0,
                    'auto_coders': len(auto_codings),
                    'selection_type': 'auto_consensus',
                    'priority_reason': 'Keine manuellen Kodierungen verfügbar - automatischer Konsens'
                })
                print(f"    Automatischer Konsens: '{selected_coding['category']}' mit {len(selected_coding.get('subcategories', []))} Subkategorien")
            else:
                # Fallback: WÄhle automatische Kodierung mit hÖchster Konfidenz
                selected_coding = max(
                    auto_codings,
                    key=lambda x: self._extract_confidence_value(x)
                ).copy()
                
                # VEREINFACHT: Direkte Übernahme
            selected_coding['subcategories'] = selected_coding.get('subcategories', [])  # DIREKT
    
        return selected_coding


    def _calculate_coding_quality(self, coding: Dict, consensus_subcats: List[str]) -> float:
        """
        Berechnet einen QualitÄtsscore fuer eine Kodierung.
        BerÜcksichtigt mehrere Faktoren:
        - Konfidenz der Kodierung
        - Übereinstimmung mit Konsens-Subkategorien
        - QualitÄt der Begründung

        Args:
            coding: Einzelne Kodierung
            consensus_subcats: Liste der Konsens-Subkategorien

        Returns:
            float: QualitÄtsscore zwischen 0 und 1
        """
        try:
            # Hole Konfidenzwert (gesamt oder Hauptkategorie)
            if isinstance(coding.get('confidence'), dict):
                confidence = float(coding['confidence'].get('total', 0))
            else:
                confidence = float(coding.get('confidence', 0))

            # Berechne Übereinstimmung mit Konsens-Subkategorien
            coding_subcats = set(coding.get('subcategories', []))
            consensus_subcats_set = set(consensus_subcats)
            if consensus_subcats_set:
                subcat_overlap = len(coding_subcats & consensus_subcats_set) / len(consensus_subcats_set)
            else:
                subcat_overlap = 1.0  # Volle Punktzahl wenn keine Konsens-Subkategorien

            # Bewerte QualitÄt der Begründung
            justification = coding.get('justification', '')
            if isinstance(justification, str):
                justification_score = min(len(justification.split()) / 20, 1.0)  # Max bei 20 WÖrtern
            else:
                justification_score = 0.0  # Keine Begründung vorhanden oder ungültiger Typ

            # Gewichtete Kombination der Faktoren
            quality_score = (
                confidence * 0.5 +          # 50% Konfidenz
                subcat_overlap * 0.3 +      # 30% Subkategorien-Übereinstimmung
                justification_score * 0.2   # 20% BegründungsqualitÄt
            )

            return quality_score

        except Exception as e:
            print(f"Fehler bei der Berechnung der CodierungsqualitÄt: {str(e)}")
            return 0.0  # RÜckgabe eines neutralen Scores im Fehlerfall
    
    def export_optimization_analysis(self, 
                                original_categories: Dict[str, CategoryDefinition],
                                optimized_categories: Dict[str, CategoryDefinition],
                                optimization_log: List[Dict]):
        """Exportiert eine detaillierte Analyse der Kategorienoptimierungen."""
        
        analysis_path = os.path.join(self.output_dir, 
                                    f'category_optimization_{datetime.now().strftime("%Y%m%d_%H%M%S")}.md')
        
        with open(analysis_path, 'w', encoding='utf-8') as f:
            f.write("# Analyse der Kategorienoptimierungen\n\n")
            
            f.write("## Übersicht\n")
            f.write(f"- UrsprÜngliche Kategorien: {len(original_categories)}\n")
            f.write(f"- Optimierte Kategorien: {len(optimized_categories)}\n")
            f.write(f"- Anzahl der Optimierungen: {len(optimization_log)}\n\n")
            
            f.write("## Detaillierte Optimierungen\n")
            for entry in optimization_log:
                if entry['type'] == 'merge':
                    f.write(f"\n### ZusammenfÜhrung zu: {entry['result_category']}\n")
                    f.write(f"- UrsprÜngliche Kategorien: {', '.join(entry['original_categories'])}\n")
                    f.write(f"- Zeitpunkt: {entry['timestamp']}\n\n")
                    
                    f.write("#### UrsprÜngliche Definitionen:\n")
                    for cat in entry['original_categories']:
                        if cat in original_categories:
                            f.write(f"- {cat}: {original_categories[cat].definition}\n")
                    f.write("\n")
                    
                    if entry['result_category'] in optimized_categories:
                        f.write("#### Neue Definition:\n")
                        f.write(f"{optimized_categories[entry['result_category']].definition}\n\n")
                        
                # Weitere Optimierungstypen hier...
            
            f.write("\n## Statistiken\n")
            f.write(f"- Kategorienreduktion: {(1 - len(optimized_categories) / len(original_categories)) * 100:.1f}%\n")
            
            # ZÄhle Optimierungstypen
            optimization_types = Counter(entry['type'] for entry in optimization_log)
            f.write("\nOptimierungstypen:\n")
            for opt_type, count in optimization_types.items():
                f.write(f"- {opt_type}: {count}\n")
        
        print(f"Optimierungsanalyse exportiert nach: {analysis_path}")
  
    
    def _validate_export_data(self, export_data: List[dict]) -> bool:
        """
        Validiert die zu exportierenden Daten.
        
        Args:
            export_data: Liste der aufbereiteten Export-Daten
            
        Returns:
            bool: True wenn Daten valide sind
        """
        required_columns = {
            'Dokument', 'Chunk_Nr', 'Text', 'Kodiert', 
            'Hauptkategorie', 'Kategorietyp', 'Subkategorien', 
            'Begründung', 'Konfidenz', 'Mehrfachkodierung'
        }
        
        try:
            if not export_data:
                print("Warnung: Keine Daten zum Exportieren vorhanden")
                return False
                
            # Prüfe ob alle erforderlichen Spalten vorhanden sind
            for entry in export_data:
                missing_columns = required_columns - set(entry.keys())
                if missing_columns:
                    print(f"Warnung: Fehlende Spalten in Eintrag: {missing_columns}")
                    return False
                    
                # Prüfe Kodiert-Status
                if entry['Kodiert'] not in {'Ja', 'Nein'}:
                    print(f"Warnung: Ungueltiger Kodiert-Status: {entry['Kodiert']}")
                    return False
                    
            print("Validierung der Export-Daten erfolgreich")
            return True
            
        except Exception as e:
            print(f"Fehler bei der Validierung der Export-Daten: {str(e)}")
            return False
            
    def _extract_metadata(self, filename: str) -> tuple:
        """Extrahiert Metadaten aus dem Dateinamen"""
        from pathlib import Path
        tokens = Path(filename).stem.split("_")
        attribut1 = tokens[0] if len(tokens) >= 1 else ""
        attribut2 = tokens[1] if len(tokens) >= 2 else ""
        attribut3 = tokens[2] if len(tokens) >= 3 else "" 
        return attribut1, attribut2, attribut3

    def _initialize_category_colors(self, df: pd.DataFrame) -> None:
        """
        Initialisiert die Farbzuordnung fuer alle Kategorien einmalig.
        
        Args:
            df: DataFrame mit einer 'Hauptkategorie' Spalte
        """
        if not self.category_colors:  # Nur initialisieren wenn noch nicht geschehen
            import unicodedata
            
            # Hole alle eindeutigen Hauptkategorien auẞer 'Nicht kodiert'
            raw_categories = df['Hauptkategorie'].unique()
            
            # FIX: Normalize category strings consistently
            normalized_categories = []
            for cat in raw_categories:
                if cat is not None and cat != 'Nicht kodiert':
                    normalized_cat = unicodedata.normalize('NFKC', str(cat).strip())
                    if normalized_cat and normalized_cat != 'Nicht kodiert':
                        normalized_categories.append(normalized_cat)
            
            # Sort for consistent color assignment
            categories = sorted(set(normalized_categories))
            
            # Generiere Pastellfarben
            colors = generate_pastel_colors(len(categories))
            
            # Erstelle Mapping in alphabetischer Reihenfolge
            self.category_colors = {
                category: color for category, color in zip(categories, colors)
            }
            
            # Füge 'Nicht kodiert' mit grauer Farbe hinzu
            if 'Nicht kodiert' in raw_categories:
                self.category_colors['Nicht kodiert'] = 'CCCCCC'
            
            print("\nFarbzuordnung initialisiert:")
            for cat, color in self.category_colors.items():
                print(f"- {cat}: {color}")


    async def export_results(self,
                            codings: List[Dict],
                            reliability: float,
                            categories: Dict[str, CategoryDefinition],
                            chunks: Dict[str, List[str]],
                            revision_manager: 'CategoryRevisionManager',
                            export_mode: str = "consensus",
                            original_categories: Dict[str, CategoryDefinition] = None,
                            original_codings: List[Dict] = None,  
                            inductive_coder: 'InductiveCoder' = None,
                            document_summaries: Dict[str, str] = None,
                            is_intermediate_export: bool = False) -> None: 
        """
        FIX: Exportiert mit korrekten ursprünglichen Kodierungen fuer Reliabilität
        FÜr ResultsExporter Klasse
        
        Args:
            codings: Finale/Review-Kodierungen fuer Export
            categories: Finale Kategorien
            original_categories: UrsprÜngliche Kategorien
            document_summaries: Document Summaries
            revision_manager: Revision Manager
            original_codings: FIX: UrsprÜngliche Kodierungen fuer Reliabilitätsberechnung
            export_mode: Export-Modus
            reliability: FIX: Bereits berechnete Reliabilität aus main()
        """
        try:
            # FIX: Cache Segment-Texte für bessere Text-Extraktion
            print("   📝 Erstelle Segment-Text-Cache für Export...")
            self.cache_segment_texts(codings, chunks)
            if original_codings:
                self.cache_segment_texts(original_codings, chunks)
            print(f"   📝 Cache erstellt mit {len(self._segment_text_cache)} Einträgen")
            
            # FIX: Debug - prüfe Relevance-Checker Verfügbarkeit
            relevance_checker = None
            if hasattr(self, 'relevance_checker') and self.relevance_checker:
                relevance_checker = self.relevance_checker
                print(f"   🔍 DEBUG: Relevance-Checker direkt verfügbar mit {len(relevance_checker.relevance_details)} Details")
            elif hasattr(self, 'analysis_manager') and self.analysis_manager and hasattr(self.analysis_manager, 'relevance_checker'):
                relevance_checker = self.analysis_manager.relevance_checker
                if relevance_checker:
                    print(f"   🔍 DEBUG: Relevance-Checker via analysis_manager verfügbar mit {len(relevance_checker.relevance_details)} Details")
                    # Zeige Beispiel-Keys
                    sample_keys = list(relevance_checker.relevance_details.keys())[:3]
                    print(f"   🔍 DEBUG: Beispiel Relevance-Keys: {sample_keys}")
                else:
                    print("   ❌ DEBUG: analysis_manager.relevance_checker ist None")
            else:
                print("   ❌ DEBUG: Kein Relevance-Checker verfügbar")
            
            # FIX: Debug - zeige Beispiel-Kodierungen
            multiple_codings = [c for c in codings if c.get('is_multiple_coding', False)]
            if multiple_codings:
                print(f"   🔍 DEBUG: {len(multiple_codings)} Mehrfachkodierungen gefunden")
                for i, mc in enumerate(multiple_codings[:2]):  # Zeige erste 2
                    segment_id = mc.get('segment_id', 'N/A')
                    has_text = bool(mc.get('text', ''))
                    has_result_text = bool(mc.get('result', {}).get('text', ''))
                    original_id = mc.get('original_segment_id', 'N/A')
                    print(f"     Beispiel {i+1}: {segment_id}")
                    print(f"       - text: {has_text}")
                    print(f"       - result.text: {has_result_text}")
                    print(f"       - original_segment_id: {original_id}")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            analysis_mode = CONFIG.get('ANALYSIS_MODE', 'deductive')
            # FIX: Unterschiedliche Dateinamen fuer normale und Zwischenexporte
            if is_intermediate_export:
                filename = f"QCA-AID_Analysis_INTERMEDIATE_{analysis_mode}_{timestamp}.xlsx"
                print(f"🧾 Exportiere Zwischenergebnisse bei Abbruch...")
            else:
                filename = f"QCA-AID_Analysis_{analysis_mode}_{timestamp}.xlsx"
                print(f"🧾 Exportiere finale Ergebnisse mit {export_mode}-Modus...")
            
            filepath = os.path.join(self.output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                print(f"🧾 Exportiere umfassende Ergebnisse mit {export_mode}-Modus...")
                
                # 1. HAUPT-SHEET: Kodierungsergebnisse (finale Kodierungen)
                print("🔀 Exportiere Hauptergebnisse...")
                self._export_main_results(writer, codings, original_categories)
                
                # 2. FIX: HÄUFIGKEITEN-SHEET (finale Kodierungen)
                print("🧾 Exportiere HÄufigkeiten...")
                df_coded = self._prepare_dataframe_for_frequency_analysis(codings)
                if not df_coded.empty:
                    attribut1_label = self.attribute_labels.get('attribut1', 'Attribut1')
                    attribut2_label = self.attribute_labels.get('attribut2', 'Attribut2')
                    self._export_frequency_analysis(writer, df_coded, attribut1_label, attribut2_label)
                
                # 3. FIX: INTERCODER-BERICHT mit ursprünglichen Kodierungen
                print("🧾 Exportiere IntercoderBericht...")
                if original_codings and reliability is not None:
                    # FIX: Verwende bereits berechnete Reliabilität und ursprüngliche Kodierungen
                    self._export_intercoder_bericht(writer, original_codings, reliability)
                    print(f"✅ IntercoderBericht mit Alpha={reliability:.3f} erstellt")
                else:
                    print("❌ Keine ursprünglichen Kodierungen oder Reliabilität verfügbar")
                    self._create_empty_intercoder_sheet(writer)
                
                # 4. KATEGORIEN-ÜBERSICHT
                if categories:
                    print("ℹ️ Exportiere Kategorien-Übersicht...")
                    self._export_categories_sheet_formatted(writer, categories, original_categories)
                
                # 5. PROGRESSIVE SUMMARIES (falls vorhanden)
                if document_summaries:
                    print("🔀 Exportiere Progressive Summaries...")
                    self._export_progressive_summaries(writer, document_summaries)
                
                # 6. REVIEW-STATISTIKEN
                print("🎯 Exportiere Review-Statistiken...")
                review_stats = self._calculate_review_statistics(codings, export_mode, original_codings)
                self._export_review_statistics(writer, review_stats, export_mode)
                
                # 7. REVISIONSHISTORIE (falls verfügbar)
                if revision_manager and hasattr(revision_manager, 'changes'):
                    print("🧾 Exportiere Revisionshistorie...")
                    revision_manager._export_revision_history(writer, revision_manager.changes)
                else:
                    print("ℹ️ Keine Revisionshistorie verfügbar")
                
                # 8. KONFIGURATION-SHEET 
                print("✅ Exportiere Konfiguration...")
                self._export_configuration(writer, export_mode)
                
                if is_intermediate_export:
                    print(f"✅ Zwischenergebnisse erfolgreich exportiert!")
                else:
                    print(f"✅ Export erfolgreich: {filename}")
                
            # FIX: Nur bei normalem Export Dateiinfo anzeigen
            if not is_intermediate_export:
                print(f"ℹ️ Dateien im Ordner: {self.output_dir}")
                print(f"ℹ️ Export-Datei: {filename}")
            
            # FIX: Return filename am Ende hinzugefÜgt
            return filename
            
        except Exception as e:
            print(f"⚠️ Fehler beim Export: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    

    def _prepare_dataframe_for_frequency_analysis(self, codings: List[Dict]) -> pd.DataFrame:
        """
        NEUE METHODE: Bereitet DataFrame fuer HÄufigkeitsanalyse vor
        """
        try:
            # Erstelle DataFrame aus Kodierungen
            data = []
            
            for coding in codings:
                doc_name = coding.get('document', '')
                
                # Extrahiere Attribute aus Dokumentname
                attribut1, attribut2, attribut3 = self._extract_attributes_from_document(doc_name)
                
                # Grunddaten
                row_data = {
                    'Dokument': doc_name,
                    self.attribute_labels.get('attribut1', 'Attribut1'): attribut1,
                    self.attribute_labels.get('attribut2', 'Attribut2'): attribut2,
                    'Chunk_Nr': coding.get('chunk_id', coding.get('segment_id', '')),
                    'Hauptkategorie': coding.get('category', ''),
                    'Kodiert': 'Ja' if coding.get('category') and coding.get('category') not in ['Nicht kodiert', 'Kein Kodierkonsens'] else 'Nein',
                    'Subkategorien': ', '.join(coding.get('subcategories', [])),
                    'Konfidenz': self._extract_confidence_from_coding(coding)
                }
                
                # FIX: Füge attribut3 hinzu, wenn es definiert ist
                if 'attribut3' in self.attribute_labels and self.attribute_labels['attribut3']:
                    row_data[self.attribute_labels['attribut3']] = attribut3
                
                data.append(row_data)
            
            if data:
                df = pd.DataFrame(data)
                print(f"🧾 DataFrame erstellt: {len(df)} Zeilen fuer HÄufigkeitsanalyse")
                return df
            else:
                print("❌ Keine Daten fuer DataFrame verfügbar")
                return pd.DataFrame()
                
        except Exception as e:
            print(f"⚠️ Fehler beim Erstellen des DataFrames: {str(e)}")
            return pd.DataFrame()
    
    def _extract_confidence_from_coding(self, coding: Dict) -> float:
        """
        HILFSMETHODE: Extrahiert Konfidenzwert aus Kodierung
        """
        confidence = coding.get('confidence', {})
        if isinstance(confidence, dict):
            return confidence.get('total', 0.0)
        elif isinstance(confidence, (int, float)):
            return float(confidence)
        return 0.0
    
    def _export_main_results(self, writer, codings: List[Dict], original_categories: Dict):
        """
        Exportiert Haupt-Kodierungsergebnisse
        FIX: Übernimmt alle fehlenden Spalten und Logik aus _prepare_coding_for_export
        """
        
        # FIX: Berechne korrekte Instanznummern vor der Verarbeitung
        print("   🔧 Berechne Instanznummern für wiederholte Segmente...")
        updated_codings = self._calculate_instance_info(codings)
        
        # DataFrame erstellen
        data = []
        
        for coding in updated_codings:
            # FIX: Definiere segment_id einheitlich für alle Segmente am Anfang
            segment_id = coding.get('segment_id', '')
            
            # FIX: Normalisiere segment_id für konsistente Relevanz-Lookup
            def normalize_segment_id_for_relevance(sid):
                """Normalisiert segment_id für Relevanz-Lookup - uses standardized approach"""
                if not sid:
                    return sid
                # Use the centralized standardization method instead of hardcoding .pdf
                return self._standardize_segment_id(sid)
            
            # FIX: Hole Relevanz-Details einmal für alle Segmente (kodiert und nicht-kodiert)
            relevance_details = None
            if segment_id:
                # FIX: Debug - zeige segment_id und Suchversuche
                # print(f"   🔍 DEBUG: Suche Relevanz-Details für segment_id: '{segment_id}'")
                
                # Teste verschiedene segment_id Formate
                test_ids = [
                    segment_id,  # Original
                    normalize_segment_id_for_relevance(segment_id),  # Normalisiert
                    # Remove hardcoded .pdf replacements - use standardized approach
                    self._standardize_segment_id(segment_id),
                ]
                
                # FIX: Entferne Mehrfachkodierungs-Suffix für Relevanz-Lookup
                base_segment_id = segment_id.split('-')[0] if '-' in segment_id else segment_id
                if base_segment_id != segment_id:
                    test_ids.extend([
                        base_segment_id,
                        normalize_segment_id_for_relevance(base_segment_id),
                        # Remove hardcoded .pdf replacements - use standardized approach
                        self._standardize_segment_id(base_segment_id),
                    ])
                
                # print(f"   🔍 DEBUG: Test-IDs: {test_ids}")
                pass
                
                for test_id in test_ids:
                    if hasattr(self, 'relevance_checker') and self.relevance_checker:
                        relevance_details = self.relevance_checker.get_relevance_details(test_id)
                        if relevance_details:
                            # print(f"   ✅ Relevanz-Details gefunden mit ID: '{test_id}'")
                            break
                    elif hasattr(self, 'analysis_manager') and self.analysis_manager and hasattr(self.analysis_manager, 'relevance_checker'):
                        relevance_details = self.analysis_manager.relevance_checker.get_relevance_details(test_id)
                        if relevance_details:
                            # print(f"   ✅ Relevanz-Details gefunden mit ID: '{test_id}' (via analysis_manager)")
                            break
                
                if not relevance_details:
                    # print(f"   ❌ Keine Relevanz-Details gefunden für segment_id: '{segment_id}'")
                    # FIX: Debug - zeige verfügbare Keys im relevance_checker
                    if hasattr(self, 'relevance_checker') and self.relevance_checker:
                        available_keys = list(self.relevance_checker.relevance_details.keys())[:5]
                        # print(f"   🔍 Verfügbare Keys (erste 5): {available_keys}")
                        # FIX: Suche nach ähnlichen Keys
                        similar_keys = [k for k in self.relevance_checker.relevance_details.keys() if segment_id.replace('_chunk_', '_') in k or k.replace('_chunk_', '_') in segment_id]
                        if similar_keys:
                            # print(f"   🔍 Ähnliche Keys gefunden: {similar_keys[:3]}")
                            pass
                    elif hasattr(self, 'analysis_manager') and self.analysis_manager and hasattr(self.analysis_manager, 'relevance_checker'):
                        available_keys = list(self.analysis_manager.relevance_checker.relevance_details.keys())[:5]
                        # print(f"   🔍 Verfügbare Keys (erste 5): {available_keys}")
                        # FIX: Suche nach ähnlichen Keys
                        similar_keys = [k for k in self.analysis_manager.relevance_checker.relevance_details.keys() if segment_id.replace('_chunk_', '_') in k or k.replace('_chunk_', '_') in segment_id]
                        if similar_keys:
                            # print(f"   🔍 Ähnliche Keys gefunden: {similar_keys[:3]}")
                            pass
                else:
                    # print(f"   📋 Relevanz-Details gefunden: reasoning='{relevance_details.get('reasoning', 'N/A')[:50]}...'")
                    print(f"   ")
            
            # FIX: Übernehme Dokumentname-Extraktion aus _prepare_coding_for_export
            doc_name = coding.get('document', '')
            if not doc_name:
                # Fallback: Extrahiere aus segment_id
                if segment_id and '_chunk_' in segment_id:
                    doc_name = segment_id.split('_chunk_')[0]
                else:
                    doc_name = 'Unbekanntes_Dokument'
            
            # FIX: Übernehme _extract_metadata Logik fuer drei Attribute
            if hasattr(self, '_extract_metadata'):
                attribut1, attribut2, attribut3 = self._extract_metadata(doc_name)
            else:
                # Fallback: Extrahiere aus Dokumentname
                from pathlib import Path
                tokens = Path(doc_name).stem.split("_")
                attribut1 = tokens[0] if len(tokens) >= 1 else ""
                attribut2 = tokens[1] if len(tokens) >= 2 else ""
                attribut3 = tokens[2] if len(tokens) >= 3 else ""

            # FIX: Erstelle eindeutigen PrÄfix fuer Chunk-Nr mit bis zu 6 Buchstaben pro Attribut
            chunk_prefix = ""
            if attribut1 and attribut2:
                # FIX: Extrahiere bis zu 6 Buchstaben pro Attribut (oder alle verfügbaren)
                import re
                attr1_letters = re.sub(r'[^a-zA-Z0-9]', '', attribut1)[:6]
                attr2_letters = re.sub(r'[^a-zA-Z0-9]', '', attribut2)[:6]
                attr3_letters = re.sub(r'[^a-zA-Z0-9]', '', attribut3)[:6]
                chunk_prefix = (attr1_letters + "_" + attr2_letters + "_" + attr3_letters).upper()
            else:
                chunk_prefix = doc_name[:5].upper()
            
            # FIX: Korrekte chunk_id Extraktion
            chunk_id = 0
            segment_id = coding.get('segment_id', '')
            
            # Versuche chunk_id direkt zu bekommen
            if 'chunk_id' in coding and isinstance(coding['chunk_id'], int):
                chunk_id = coding['chunk_id']
            elif segment_id and '_chunk_' in segment_id:
                # Extrahiere aus segment_id
                try:
                    chunk_part = segment_id.split('_chunk_')[1]
                    # Falls Mehrfachkodierung: "123-1" -> nimm nur "123"
                    if '-' in chunk_part:
                        chunk_id = int(chunk_part.split('-')[0])
                    else:
                        chunk_id = int(chunk_part)
                except (ValueError, IndexError):
                    chunk_id = 0
            
            # FIX: Robuste Text-Extraktion für alle Kodierungstypen
            text = coding.get('text', '')
            if not text and 'result' in coding:
                text = coding['result'].get('text', '')
            
            # FIX: Use standardized segment ID for consistent text retrieval
            if not text:
                segment_id = coding.get('segment_id', '')
                print(f"   🔍 DEBUG: Looking for text for segment {segment_id}...")
                
                # FIX: Check if segment_id looks like a display ID (this should NOT happen)
                if segment_id and not '_chunk_' in segment_id and '-' in segment_id:
                    print(f"   ❌ ERROR: segment_id appears to be a display ID: {segment_id}")
                    print(f"   ❌ This indicates a bug where display_id overwrote segment_id!")
                    print(f"   ❌ The root cause needs to be fixed - display_id should never replace segment_id!")
                
                # Standardize the segment ID for lookup
                standardized_id = self._standardize_segment_id(segment_id)
                
                # Try standardized ID first
                if hasattr(self, '_segment_text_cache'):
                    text = self._segment_text_cache.get(standardized_id, '')
                    if text:
                        print(f"   ✅ Text found via standardized ID: {standardized_id}")
                
                # Try base ID (without multiple coding suffix)
                if not text and '-' in standardized_id:
                    base_id = standardized_id.rsplit('-', 1)[0]
                    text = self._segment_text_cache.get(base_id, '')
                    if text:
                        print(f"   ✅ Text found via base ID: {base_id}")
                
                # Try original segment_id as fallback
                if not text and segment_id != standardized_id:
                    text = self._segment_text_cache.get(segment_id, '')
                    if text:
                        print(f"   ✅ Text found via original segment_id: {segment_id}")
                
                # Try original_segment_id from coding
                if not text:
                    original_id = coding.get('original_segment_id', '')
                    if original_id:
                        standardized_original = self._standardize_segment_id(original_id)
                        text = self._segment_text_cache.get(standardized_original, '')
                        if text:
                            print(f"   ✅ Text found via standardized original ID: {standardized_original}")
                        elif original_id != standardized_original:
                            text = self._segment_text_cache.get(original_id, '')
                            if text:
                                print(f"   ✅ Text found via original_segment_id: {original_id}")
                
                # Try result.original_segment_id as final fallback
                if not text and 'result' in coding:
                    result_original_id = coding['result'].get('original_segment_id', '')
                    if result_original_id:
                        standardized_result = self._standardize_segment_id(result_original_id)
                        text = self._segment_text_cache.get(standardized_result, '')
                        if text:
                            print(f"   ✅ Text found via standardized result original ID: {standardized_result}")
                        elif result_original_id != standardized_result:
                            text = self._segment_text_cache.get(result_original_id, '')
                            if text:
                                print(f"   ✅ Text found via result.original_segment_id: {result_original_id}")
                
                # Warning if still no text found
                if not text:
                    is_multiple = coding.get('is_multiple_coding', False)
                    original_id = coding.get('original_segment_id', 'N/A')
                    cache_size = len(self._segment_text_cache) if hasattr(self, '_segment_text_cache') else 0
                    print(f"⚠️ WARNUNG: Kein Original-Text für Segment {segment_id} gefunden")
                    print(f"   - Standardized ID: {standardized_id}")
                    print(f"   - Mehrfachkodierung: {is_multiple}")
                    print(f"   - Original ID: {original_id}")
                    print(f"   - Cache Größe: {cache_size}")
                    if hasattr(self, '_segment_text_cache') and cache_size > 0:
                        sample_keys = list(self._segment_text_cache.keys())[:5]
                        print(f"   - Sample Cache-Keys: {sample_keys}")
                    text = "[Original-Text nicht verfügbar]"
            
            # FIX: Warnung bei fehlendem Text mit mehr Details
            if not text:
                segment_id = coding.get('segment_id', 'UNKNOWN')
                is_multiple = coding.get('is_multiple_coding', False)
                original_id = coding.get('original_segment_id', 'N/A')
                cache_size = len(self._segment_text_cache) if hasattr(self, '_segment_text_cache') else 0
                print(f"⚠️ WARNUNG: Kein Original-Text für Segment {segment_id} gefunden")
                print(f"   - Mehrfachkodierung: {is_multiple}")
                print(f"   - Original ID: {original_id}")
                print(f"   - Cache Größe: {cache_size}")
                if hasattr(self, '_segment_text_cache') and cache_size > 0:
                    # Zeige ähnliche Keys im Cache
                    if '-' in segment_id:
                        base_id = segment_id.rsplit('-', 1)[0]
                        similar_keys = [k for k in self._segment_text_cache.keys() if base_id in k]
                        print(f"   - Ähnliche Cache-Keys: {similar_keys[:3]}")
                    else:
                        sample_keys = list(self._segment_text_cache.keys())[:3]
                        print(f"   - Cache Beispiel-Keys: {sample_keys}")
                text = "[Original-Text nicht verfügbar]"
            
            paraphrase = coding.get('paraphrase', '')
            if not paraphrase and 'result' in coding:
                paraphrase = coding['result'].get('paraphrase', '')
            
            # FIX: Hole Kategorie und behandle leere Kategorien explizit
            category = coding.get('category', '')
            
            # FIX: Zentrale Behandlung von leeren/fehlenden Kategorien
            if not category or str(category).strip() == "":
                display_category = "Nicht kodiert"
                is_coded = 'Nein'
                category_type = 'unkodiert'
            elif category == "Kein Kodierkonsens":
                display_category = "Kein Kodierkonsens"
                is_coded = 'Nein'
                category_type = 'unkodiert'
            else:
                display_category = category
                is_coded = 'Ja'
                # Kategorietyp bestimmen fuer gültige Kategorien
                if hasattr(self, '_determine_category_type'):
                    category_type = self._determine_category_type(category, original_categories or {})
                else:
                    # Fallback-Logik
                    if original_categories and category in original_categories:
                        category_type = 'deduktiv'
                    else:
                        category_type = 'induktiv'
            # FIX: Ende
            
            # FIX: Subkategorien verarbeiten wie in _prepare_coding_for_export
            subcategories = coding.get('subcategories', [])
            subcats_text = ""
            if subcategories:
                if isinstance(subcategories, str):
                    subcats_text = subcategories.strip()
                elif isinstance(subcategories, (list, tuple)):
                    clean_subcats = []
                    for subcat in subcategories:
                        if subcat and str(subcat).strip():
                            clean_text = str(subcat).strip()
                            clean_text = clean_text.replace('[', '').replace(']', '').replace("'", "").replace('"', '')
                            if clean_text:
                                clean_subcats.append(clean_text)
                    subcats_text = ', '.join(clean_subcats)
                elif isinstance(subcategories, dict):
                    clean_subcats = []
                    for key in subcategories.keys():
                        clean_key = str(key).strip()
                        if clean_key:
                            clean_subcats.append(clean_key)
                    subcats_text = ', '.join(clean_subcats)
                else:
                    subcats_text = str(subcategories).strip()
                    subcats_text = subcats_text.replace('[', '').replace(']', '').replace("'", "").replace('"', '')
            
            # Zusätzliche Bereinigung
            subcats_text = subcats_text.replace('[', '').replace(']', '').replace("'", "")
            
            # FIX: Keywords verarbeiten wie in _prepare_coding_for_export
            raw_keywords = coding.get('keywords', '')
            if isinstance(raw_keywords, list):
                formatted_keywords = [kw.strip() for kw in raw_keywords]
            else:
                formatted_keywords = raw_keywords.replace("[", "").replace("]", "").replace("'", "").split(",")
                formatted_keywords = [kw.strip() for kw in formatted_keywords if kw.strip()]
            keywords_text = ', '.join(formatted_keywords)
            
            # FIX: VERBESSERTE BegründungSVERARBEITUNG - KERNFIX
            justification = ""
            
            # ✅ SPEZIELLE BEHANDLUNG für nicht-kodierte Segmente
            if category in ["Nicht kodiert", ""] or display_category == "Nicht kodiert":
                # Priorität 1: Verwende Relevanz-Details Begründung (wichtigste Quelle für nicht-kodierte Segmente)
                if relevance_details and relevance_details.get('reasoning') and relevance_details['reasoning'] != 'Keine Begründung verfügbar':
                    justification = relevance_details['reasoning']
                    print(f"   ✅ Spezifische LLM-Begründung verwendet für {segment_id}: {justification[:50]}...")
                # Priorität 2: Fallback zu coding reasoning
                elif coding.get('reasoning') and coding.get('reasoning').strip() and coding.get('reasoning') != 'NICHT VORHANDEN':
                    justification = coding.get('reasoning')
                    print(f"   ✅ Coding-Reasoning verwendet für {segment_id}: {justification[:50]}...")
                else:
                    # Priorität 3: Intelligente Fallback-Begründungen basierend auf Textanalyse
                    text_content = text.lower() if text else ""
                    text_length = len(text_content.strip())
                    
                    if text_length < 20:
                        justification = "Segment zu kurz fuer sinnvolle Kodierung"
                    elif any(pattern in text_content for pattern in ['seite ', 'page ', 'copyright', '©', 'inhaltsverzeichnis', 'table of contents']):
                        justification = "Segment als Metadaten (z.B. Seitenzahl, Copyright) identifiziert"
                    elif any(pattern in text_content for pattern in ['abstract', 'zusammenfassung', 'einleitung']):
                        justification = "Segment auẞerhalb des Analysebereichs der Forschungsfrage"
                    elif text_length < 100:
                        justification = "Segment enthÄlt zu wenig Substanz fuer thematische Kodierung"
                    else:
                        justification = "Segment nicht relevant fuer die definierten Analysekategorien"
                    print(f"   ⚠️ Fallback-Begründung verwendet für {segment_id}: {justification}")
            else:
                # ✅ NORMALE BEHANDLUNG für kodierte Segmente
                # 1. Zuerst normale justification (von DeductiveCoder bei Kodierung)
                if coding.get('justification') and coding.get('justification').strip():
                    justification = coding.get('justification')
                # 2. Fallback: RelevanceChecker reasoning
                elif coding.get('reasoning') and coding.get('reasoning').strip() and coding.get('reasoning') != 'NICHT VORHANDEN':
                    justification = coding.get('reasoning')
                # 3. Fallback: original_justification
                elif coding.get('original_justification') and coding.get('original_justification').strip() and coding.get('original_justification') != 'NICHT VORHANDEN':
                    justification = coding.get('original_justification')
                else:
                    # Fallback für kodierte Segmente ohne Begründung
                    justification = "Kodierung ohne spezifische Begründung dokumentiert"
            
            # FIX: Konfidenz korrekt extrahieren
            confidence = coding.get('confidence', {})
            if isinstance(confidence, dict):
                confidence_value = confidence.get('total', 0.0)
            elif isinstance(confidence, (int, float)):
                confidence_value = float(confidence)
            else:
                confidence_value = 0.0
            
            # FIX: Mehrfachkodierungs-Info wie in _prepare_coding_for_export
            consensus_info = coding.get('consensus_info', {})
            # FIX: Verwende Normalisierungsmethode für konsistente Original_Chunk_ID
            original_segment_id = consensus_info.get('original_segment_id', coding.get('segment_id', ''))
            original_chunk_id = self._normalize_segment_id(original_segment_id)
            is_multiple = consensus_info.get('is_multiple_coding_instance', False)
            instance_info = consensus_info.get('instance_info', {})
            
            # FIX: Erstelle eindeutige Chunk-ID mit Mehrfachkodierungs-Suffix
            if coding.get('total_coding_instances', 1) > 1:
                unique_chunk_id = f"{chunk_prefix}-{chunk_id}-{coding.get('multiple_coding_instance', 1)}"
                mehrfachkodierung_status = 'Ja'
            else:
                unique_chunk_id = f"{chunk_prefix}-{chunk_id}"
                mehrfachkodierung_status = 'Nein'
            
            # FIX: Zeile erstellen mit ALLEN Spalten aus _prepare_coding_for_export
            row_data = {
                'Dokument': sanitize_text_for_excel(doc_name),  # FIX: Korrekte ParameterÜbergabe
                self.attribute_labels.get('attribut1', 'Attribut1'): sanitize_text_for_excel(attribut1),
                self.attribute_labels.get('attribut2', 'Attribut2'): sanitize_text_for_excel(attribut2),
            }
            
            # FIX: Füge attribut3 hinzu, wenn es definiert ist
            if 'attribut3' in self.attribute_labels and self.attribute_labels['attribut3']:
                row_data[self.attribute_labels['attribut3']] = sanitize_text_for_excel(attribut3)
            
            # FIX: Review_Typ Logik konsistent mit Standard-Analyse
            if coding.get('manual_review', False):
                review_typ = 'manual'
            elif consensus_info.get('selection_type') == 'consensus':
                review_typ = 'consensus'
            elif consensus_info.get('selection_type') == 'majority':
                review_typ = 'consensus'  # Majority is also a form of consensus
            elif is_multiple:
                review_typ = 'consensus'  # Multiple instances require consensus
            else:
                review_typ = 'single'
            
            # FIX: Alle weiteren Spalten hinzuFügen
            additional_fields = {
                'Chunk_Nr': unique_chunk_id,
                'Text': sanitize_text_for_excel(text),  # FIX: Korrekte Funktionsaufrufe
                'Paraphrase': sanitize_text_for_excel(paraphrase),
                'Kodiert': is_coded,  # FIX: Verwende korrekten is_coded Wert
                'Hauptkategorie': sanitize_text_for_excel(display_category),  # FIX: Korrigiere Leerzeichen-Fehlern
                'Kategorietyp': category_type,  # FIX: Verwende korrekten category_type
                'Subkategorien': sanitize_text_for_excel(subcats_text),
                'Schlüsselwörter': sanitize_text_for_excel(keywords_text),
                'Begründung': sanitize_text_for_excel(justification),
                'Konfidenz': f"{confidence_value:.2f}",  # FIX: Korrekte Konfidenz-Formatierung
                'Mehrfachkodierung': mehrfachkodierung_status,  # FIX: Korrekte Mehrfachkodierung-Logik
                'Instanz_Nr': instance_info.get('instance_number', 1) if is_multiple else 1,  # FIX: HinzugefÜgt
                'Gesamt_Instanzen': instance_info.get('total_instances', 1) if is_multiple else 1,  # FIX: HinzugefÜgt
                'Review_Typ': review_typ,
                'Kodierer': coding.get('coder_id', 'Unbekannt'),  # FIX: HinzugefÜgt               
                'Kontext_verwendet': 'Ja' if coding.get('context_paraphrases_used', False) else 'Nein',  # NEU: Spalte für Kontext-Nutzung
                'Original_Chunk_ID': unique_chunk_id
            }
            
            # NEU: Relevanz-Details hinzufügen (bereits am Anfang geladen)
            if relevance_details:
                additional_fields.update({
                    'Relevanz_Stärke': f"{relevance_details.get('relevance_strength', 0.0):.2f}",
                    'Klassifikations_Konfidenz': f"{relevance_details.get('classification_confidence', 0.0):.2f}",
                    'Forschungsaspekte_gefunden': sanitize_text_for_excel(', '.join(relevance_details.get('aspects_found', []))),
                    'Relevanz_Aspekte': sanitize_text_for_excel(', '.join(relevance_details.get('key_aspects', [])))
                })
            else:
                # Fallback-Werte wenn keine Relevanz-Details verfügbar
                additional_fields.update({
                    'Relevanz_Stärke': 'N/A',
                    'Klassifikations_Konfidenz': 'N/A', 
                    'Forschungsaspekte_gefunden': 'N/A',
                    'Relevanz_Aspekte': 'N/A'
                })
            
            row_data.update(additional_fields)
            
            # FIX: Kontext-bezogene Felder hinzuFügen, wenn vorhanden
            if 'context_summary' in coding and coding['context_summary']:
                row_data['Progressive_Context'] = sanitize_text_for_excel(coding.get('context_summary', ''))
            
            if 'context_influence' in coding and coding['context_influence']:
                row_data['Context_Influence'] = sanitize_text_for_excel(coding.get('context_influence', ''))
            
            data.append(row_data)
        
        # Als DataFrame exportieren
        if data:
            df = pd.DataFrame(data)
            
            # FIX: Initialisiere Kategorie-Farben BEVOR das Sheet exportiert wird
            if not self.category_colors:
                print("   🎨 Initialisiere Kategorie-Farben für konsistente Farbgebung...")
                self._initialize_category_colors(df)
            
            df.to_excel(writer, sheet_name='Kodierungsergebnisse', index=False)
            
            # FIX: Aktiviere TabellenfunktionalitÄt explizit
            worksheet = writer.sheets['Kodierungsergebnisse']
            # print(f"FIX: Formatiere Worksheet 'Kodierungsergebnisse' mit {len(df)} Zeilen und {len(df.columns)} Spalten")
            # print(f"FIX: Spalten: {list(df.columns)}")
            self._format_worksheet(worksheet, as_table=False)  # FIX: Changed to False for robustness - use AutoFilter only
            
        else:
            print("❌ Keine Hauptergebnisse zum Exportieren")
    
    def _calculate_instance_info(self, all_codings: List[Dict]) -> List[Dict]:
        """
        Berechnet korrekte Instanznummern für wiederholte Segmente.
        
        Args:
            all_codings: Liste aller Kodierungen
            
        Returns:
            Liste der Kodierungen mit korrekten instance_info
        """
        # Gruppiere Kodierungen nach Original-Segment-ID
        segment_groups = {}
        for coding in all_codings:
            # Extrahiere Original-Segment-ID
            consensus_info = coding.get('consensus_info', {})
            original_id = consensus_info.get('original_segment_id')
            
            if not original_id:
                # Fallback: Verwende segment_id
                original_id = coding.get('segment_id', '')
                # Entferne mögliche Suffixe wie "-1", "-2" etc.
                if '-' in original_id and original_id.split('-')[-1].isdigit():
                    original_id = '-'.join(original_id.split('-')[:-1])
            
            if original_id not in segment_groups:
                segment_groups[original_id] = []
            segment_groups[original_id].append(coding)
        
        # Berechne Instanznummern für jede Gruppe
        updated_codings = []
        for original_id, codings_group in segment_groups.items():
            total_instances = len(codings_group)
            
            # Sortiere nach Kodierer-ID für konsistente Reihenfolge
            codings_group.sort(key=lambda x: x.get('coder_id', ''))
            
            for instance_number, coding in enumerate(codings_group, 1):
                # Update consensus_info mit korrekten Instanznummern
                if 'consensus_info' not in coding:
                    coding['consensus_info'] = {}
                
                coding['consensus_info']['instance_info'] = {
                    'instance_number': instance_number,
                    'total_instances': total_instances,
                    'original_segment_id': original_id,
                    'all_categories': [c.get('category', '') for c in codings_group]
                }
                
                # Markiere als Mehrfachkodierung wenn mehr als eine Instanz
                coding['consensus_info']['is_multiple_coding_instance'] = total_instances > 1
                
                # Setze auch die alten Felder für Kompatibilität
                coding['total_coding_instances'] = total_instances
                coding['multiple_coding_instance'] = instance_number
                
                updated_codings.append(coding)
        
        return updated_codings
    
    def _normalize_segment_id(self, segment_id: str) -> str:
        """
        Normalisiert Segment-IDs zu einem einheitlichen Format.
        Uses SegmentIDManager for consistent normalization.
        
        Args:
            segment_id: Original Segment-ID
            
        Returns:
            Normalisierte Segment-ID im Format: document_name.ext_chunk_0 (preserves file extension)
        """
        from ..core.segment_id_manager import SegmentIDManager
        
        if not segment_id:
            return 'unknown_segment'
        
        # Use SegmentIDManager for consistent normalization
        return SegmentIDManager.extract_base_segment_id(segment_id)
    
    def _standardize_segment_id(self, segment_id: str, multiple_coding_suffix: str = None) -> str:
        """
        CENTRAL METHOD: Standardizes segment IDs across the entire system.
        Uses SegmentIDManager for consistent ID operations.
        
        Args:
            segment_id: Original segment ID in any format
            multiple_coding_suffix: Optional suffix for multiple coding (e.g., "1", "2")
            
        Returns:
            Standardized segment ID: document.ext_chunk_0 or document.ext_chunk_0-1 (preserves file extension)
        """
        from ..core.segment_id_manager import SegmentIDManager
        
        if not segment_id:
            return 'unknown_segment'
        
        # Use SegmentIDManager for standardization
        base_id = SegmentIDManager.extract_base_segment_id(segment_id)
        standardized_id = SegmentIDManager.standardize_segment_id(base_id)
        
        # Add multiple coding suffix if provided
        if multiple_coding_suffix:
            return SegmentIDManager.add_multiple_coding_suffix(standardized_id, int(multiple_coding_suffix))
        
        return standardized_id

    def _extract_document_from_segment_id(self, segment_id: str) -> str:
        """
        Extrahiert Dokumentnamen aus segment_id falls document-Feld fehlt.
        Uses SegmentIDManager for consistent extraction.
        """
        from ..core.segment_id_manager import SegmentIDManager
        
        if not segment_id:
            return 'Unbekanntes_Dokument'
        
        return SegmentIDManager.extract_document_name(segment_id)
    
    def _extract_three_attributes_from_document(self, doc_name: str) -> tuple:
        """
        Dynamically extracts document attributes without hardcoded extensions.
        Uses SegmentIDManager for consistent attribute extraction.
        """
        # FIX: Use SegmentIDManager for dynamic extension handling
        from ..core.segment_id_manager import SegmentIDManager
        return SegmentIDManager.extract_document_attributes(doc_name)
    
    def _generate_correct_chunk_id(self, coding: Dict, attribut1: str, attribut2: str, attribut3: str) -> str:
        """
        PUNKT 4: Generiert korrekte Chunk-ID im Format AABBCC-01-01
        Uses SegmentIDManager for consistent ID operations.
        
        Format: [Erste 2 Buchstaben Attr1][Erste 2 Buchstaben Attr2][Erste 2 Buchstaben Attr3]-[Segment-Nr]-[Mehrfachkodierungs-Nr]
        """
        from ..core.segment_id_manager import SegmentIDManager
        
        # Use SegmentIDManager to create display ID
        segment_id = coding.get('segment_id', '')
        if not segment_id:
            # Fallback to manual construction
            original_segment_id = coding.get('consensus_info', {}).get('original_segment_id', '')
            segment_id = original_segment_id or 'unknown_segment_0'
        
        return SegmentIDManager.create_display_id(segment_id)
    
    def _extract_segment_number(self, segment_id: str) -> int:
        """
        Extrahiert Segment-Nummer aus segment_id
        """
        if not segment_id:
            return 1
        
        # Suche nach Zahlen in der segment_id
        numbers = re.findall(r'\d+', segment_id)
        
        if numbers:
            # Nimm die erste gefundene Zahl
            return int(numbers[0])
        
        return 1
    
    def _extract_first_letters(self, text: str, count: int) -> str:
        """
        Extrahiert die ersten N Buchstaben aus einem Text
        """
        if not text:
            return 'XX'[:count]
        
        # Nur Buchstaben extrahieren
        letters = re.sub(r'[^a-zA-Z]', '', text)
        if not letters:
            return 'XX'[:count]
        
        # Erste N Buchstaben
        result = letters[:count].upper()
        
        # AuffÜllen falls zu kurz
        while len(result) < count:
            result += 'X'
        
        return result
    
    def _determine_category_type(self, category: str, original_categories: dict) -> str:
        """
        BESTEHENDE METHODE: Bestimmt Kategorietyp (deduktiv/induktiv)
        """
        if not category or category in ['Nicht kodiert', 'Kein Kodierkonsens']:
            return ''
        
        if category in original_categories:
            return 'Deduktiv'
        else:
            return 'Induktiv'
    
    def _format_worksheet(self, worksheet, as_table: bool = False) -> None:
        """
        Formatiert das Detail-Worksheet mit flexibler Farbkodierung und adaptiven Spaltenbreiten
        FIX: Korrigiert AutoFilter/Tabellen-Konflikt
        """
        try:
            if worksheet.max_row <= 1:
                print(f"❌ Worksheet '{worksheet.title}' enthÄlt keine Daten")
                return

            # FIX: Hole DataFrame fuer Farbinitialisierung
            df_data = []
            headers = [cell.value for cell in worksheet[1]]
            
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    df_data.append(row)
            
            if df_data and 'Hauptkategorie' in headers:
                # Erstelle DataFrame fuer Farbinitialisierung
                df = pd.DataFrame(df_data, columns=headers)
                
                # FIX: Initialisiere Kategorie-Farben wenn noch nicht vorhanden
                if not hasattr(self, 'category_colors') or not self.category_colors:
                    self._initialize_category_colors(df)
                
                print(f"verfügbare Kategorie-Farben: {list(self.category_colors.keys())}")

            # Bestimme Spaltenbreiten adaptiv
            column_widths = []
            for col in range(1, worksheet.max_column + 1):
                max_length = 0
                column = get_column_letter(col)
                
                for row in range(1, min(worksheet.max_row + 1, 101)):
                    cell = worksheet[f"{column}{row}"]
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Spaltenbreite setzen (min 8, max 50)
                width = min(max(max_length + 2, 8), 50)
                column_widths.append(width)
                worksheet.column_dimensions[column].width = width

            # Header und Datenformatierung
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # FIX: Finde Hauptkategorie-Spalten-Index korrekt
            hauptkategorie_idx = None
            headers = [cell.value for cell in worksheet[1]]
            for i, header in enumerate(headers, 1):
                if header == 'Hauptkategorie':
                    hauptkategorie_idx = i
                    break
            
            print(f"Hauptkategorie-Spalte gefunden bei Index: {hauptkategorie_idx}")
            
            # Header formatieren
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.border = thin_border

            # FIX: Verbesserte Datenformatierung mit Farbkodierung
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=False, vertical='top')
                    cell.border = thin_border

                    # FIX: Farbkodierung fuer Hauptkategorien-Spalte
                    if (hauptkategorie_idx and 
                        cell.column == hauptkategorie_idx and 
                        cell.value and 
                        hasattr(self, 'category_colors') and 
                        cell.value in self.category_colors):
                        
                        color = self.category_colors[cell.value]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        # print(f"Farbe angewendet fuer '{cell.value}': {color} in Zeile {cell.row}")

            # FIX: Excel-Tabelle oder AutoFilter erstellen - NICHT BEIDES gleichzeitig
            if as_table:
                try:
                    # Entferne vorhandene Tabellen sicher
                    table_names = list(worksheet.tables.keys()).copy()
                    for table_name in table_names:
                        del worksheet.tables[table_name]
                    
                    # FIX: Entferne auch eventuelle AutoFilter vor Tabellenerstellung
                    worksheet.auto_filter.ref = None
                    
                    # Sichere Bestimmung der letzten Spalte und Zeile
                    last_col_index = worksheet.max_column
                    last_col_letter = get_column_letter(last_col_index)
                    last_row = worksheet.max_row
                    
                    # DEBUG: Gebe Dimensionen aus
                    print(f"DEBUG _format_worksheet: Dimensionen = {last_col_index}x{last_row}, max_column={worksheet.max_column}, max_row={worksheet.max_row}")
                    
                    # Robuste Validierung der Dimensionen
                    # WICHTIG: Minimum 2 Zeilen (Header + mindestens 1 Datenzeile)
                    if last_col_index < 1 or last_row < 2:
                        print(f"Warnung: Worksheet hat ungültige Dimensionen ({last_col_index}x{last_row}) - Fallback zu AutoFilter")
                        if as_table:
                            # Fallback zu AutoFilter
                            worksheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
                        return
                    
                    # Generiere eindeutigen Tabellennamen - ROBUST und kurz
                    # Max 31 Zeichen für Excel-Tabellennamen
                    sheet_title = worksheet.title.replace(' ', '_').replace('-', '_')
                    # Entferne Sonderzeichen
                    sheet_title = ''.join(c for c in sheet_title if c.isalnum() or c == '_')[:20]
                    safe_table_name = f"Tbl_{sheet_title}"
                    
                    # Stelle sicher, dass der Tabellenname eindeutig ist
                    counter = 1
                    original_name = safe_table_name
                    while safe_table_name in worksheet.tables:
                        safe_table_name = f"{original_name}_{counter}"
                        counter += 1
                    
                    # Tabellenverweis generieren - ROBUST
                    # WICHTIG: Stelle sicher, dass last_row mindestens 2 ist (Header + 1 Zeile)
                    if last_row < 2:
                        last_row = 2  # Fallback minimum
                    
                    # DEBUG: Überprüfe tatsächlich geschriebene Zeilen
                    actual_last_row = 1
                    for row in worksheet.iter_rows(min_row=1, max_row=last_row):
                        has_data = any(cell.value is not None for cell in row)
                        if has_data:
                            actual_last_row = row[0].row
                    
                    # DEBUG: Validiere Header-Zellenwerte (dürfen keine ungültigen Zeichen haben)
                    header_row = worksheet[1]
                    for cell in header_row:
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            # Prüfe auf problematische Zeichen
                            if any(c in cell_value for c in ['[', ']', '{', '}', '<', '>']):
                                print(f"WARNUNG: Header-Zelle '{cell.coordinate}' enthält problematische Zeichen: '{cell_value}'")
                                # Sanitize
                                sanitized = cell_value.replace('[', '').replace(']', '').replace('{', '').replace('}', '').replace('<', '').replace('>', '')
                                cell.value = sanitized
                    
                    print(f"DEBUG: last_row (max_row)={last_row}, actual_last_row={actual_last_row}")
                    
                    # Verwende die tatsächlich gefüllte Zeile
                    if actual_last_row < 2:
                        print(f"Warnung: Keine ausreichenden Datenzeilen ({actual_last_row}) - verwende AutoFilter")
                        worksheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
                        return
                    
                    table_ref = f"A1:{last_col_letter}{actual_last_row}"
                    
                    # Validiere Tabellenverweis
                    try:
                        # Detaillierte Validierung
                        print(f"DEBUG: Versuche Tabelle '{safe_table_name}' mit Bereich '{table_ref}' zu erstellen")
                        
                        # Prüfe ob Bereich gültig ist
                        if actual_last_row < 2 or last_col_index < 1:
                            raise ValueError(f"Ungültige Tabellen-Dimensionen: Zeilen={actual_last_row}, Spalten={last_col_index}, Bereich={table_ref}")
                        
                        # FIX: Erstelle NUR Excel-Tabelle (nicht AutoFilter)
                        tab = Table(displayName=safe_table_name, ref=table_ref)
                        style = TableStyleInfo(
                            name="TableStyleMedium9", 
                            showFirstColumn=False,
                            showLastColumn=False, 
                            showRowStripes=True, 
                            showColumnStripes=False
                        )
                        tab.tableStyleInfo = style
                        worksheet.add_table(tab)
                        
                        print(f"✓ Excel-Tabelle '{safe_table_name}' erfolgreich erstellt ({table_ref})")
                    except (ValueError, AttributeError, Exception) as ref_error:
                        print(f"Fehler bei Tabellenerstellung: {str(ref_error)}")
                        print(f"DEBUG: Fallback zu AutoFilter (last_row={last_row}, last_col_index={last_col_index})")
                        # Direktes Fallback zu AutoFilter
                        try:
                            filter_range = f"A1:{last_col_letter}{actual_last_row}"
                            worksheet.auto_filter.ref = filter_range
                            print(f"AutoFilter als Fallback erstellt: {filter_range}")
                        except Exception as filter_error:
                            print(f"Auch AutoFilter-Fallback fehlgeschlagen: {str(filter_error)}")
                    
                except Exception as table_error:
                    print(f"Unerwarteter Fehler bei Tabellenerstellung: {str(table_error)}")
            else:
                # AutoFilter verwenden statt Tabelle
                try:
                    last_col_index = worksheet.max_column
                    last_col_letter = get_column_letter(last_col_index)
                    last_row = worksheet.max_row
                    filter_range = f"A1:{last_col_letter}{last_row}"
                    worksheet.auto_filter.ref = filter_range
                    print(f"AutoFilter erstellt: {filter_range}")
                except Exception as filter_error:
                    print(f"Warnung: AutoFilter konnte nicht erstellt werden: {str(filter_error)}")

            print(f"Worksheet '{worksheet.title}' erfolgreich formatiert" + 
                (f" mit Farbkodierung fuer Hauptkategorien (Spalte {hauptkategorie_idx})" if hauptkategorie_idx else ""))
            
        except Exception as e:
            print(f"Fehler bei der Formatierung von {worksheet.title}: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _apply_category_color_coding(self, worksheet, df, category_colors):
        """
        Wendet Farbkodierung auf Hauptkategorien-Spalte an
        """
        try:
            # Finde Spalten-Index fuer Hauptkategorie
            hauptkat_col = None
            for i, col_name in enumerate(df.columns, 1):
                if col_name == 'Hauptkategorie':
                    hauptkat_col = i
                    break
            
            if not hauptkat_col:
                print("❌ Hauptkategorie-Spalte nicht gefunden fuer Farbkodierung")
                return
            
            # Wende Farben auf Zellen an (ab Zeile 2, da Zeile 1 Header ist)
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=hauptkat_col)
                category = cell.value
                
                if category and category in category_colors:
                    color = category_colors[category]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            print(f"✅ Farbkodierung angewendet fuer {len(category_colors)} Kategorien")
            
        except Exception as e:
            print(f"❌ Fehler bei Farbkodierung: {str(e)}")

    def _format_main_worksheet(self, worksheet, num_columns: int, num_rows: int):
        """
        BESTEHENDE METHODE: Formatiert das Haupt-Arbeitsblatt
        """
        # Spaltenbreite anpassen
        column_widths = {
            1: 20,   # Dokument
            2: 15,   # Attribut1
            3: 15,   # Attribut2  
            4: 15,   # Chunk_Nr
            5: 50,   # Text - WICHTIG: Breit fuer Text
            6: 30,   # Paraphrase - WICHTIG: Breit fuer Paraphrase
            7: 10,   # Kodiert
            8: 20,   # Hauptkategorie - WICHTIG
            9: 15,   # Kategorietyp
            10: 25,  # Subkategorien - WICHTIG
            11: 20,  # Schlüsselwörter - WICHTIG
            12: 40,  # Begründung - WICHTIG: Breit fuer Begründung
            13: 10,  # Konfidenz - WICHTIG
            14: 12,  # Mehrfachkodierung
            15: 10,  # Instanz_Nr
            16: 12,  # Gesamt_Instanzen
            17: 12,  # Review_Typ
            18: 15   # Kodierer
        }
        
        for col, width in column_widths.items():
            if col <= num_columns:
                worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # Text-Wrapping fuer wichtige Spalten
        for row in range(2, num_rows + 2):
            # Text-Spalte (5)
            worksheet.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')
            # Paraphrase-Spalte (6) 
            worksheet.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='top')
            # Begründung-Spalte (12)
            worksheet.cell(row=row, column=12).alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_reliability_sheet(self, workbook, reliability: float, export_mode: str):
        """
        ERWEITERTE METHODE: Erstellt Reliabilitäts-Sheet mit Mehrfachkodierungs-Hinweisen
        """
        ws = workbook.create_sheet("Reliabilität")
        
        # Header
        ws.cell(row=1, column=1, value="Intercoder-Reliabilität").font = Font(bold=True, size=14)
        
        # Krippendorff's Alpha
        ws.cell(row=3, column=1, value="Krippendorff's Alpha:")
        ws.cell(row=3, column=2, value=f"{reliability:.3f}")
        ws.cell(row=3, column=2).font = Font(bold=True, size=12)
        
        # Farbkodierung
        if reliability > 0.8:
            fill_color = '90EE90'  # GrÜn
        elif reliability > 0.667:
            fill_color = 'FFFF90'  # Gelb
        else:
            fill_color = 'FFB6C1'  # Rot
        
        ws.cell(row=3, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        # Bewertung
        ws.cell(row=4, column=1, value="Bewertung:")
        rating = ("Exzellent" if reliability > 0.8 else 
                 "Akzeptabel" if reliability > 0.667 else 
                 "Schwach" if reliability > 0.5 else "Unzureichend")
        ws.cell(row=4, column=2, value=rating)
        
        # NEUE INFORMATION: Mehrfachkodierungs-Behandlung
        ws.cell(row=6, column=1, value="Mehrfachkodierungs-Behandlung:").font = Font(bold=True)
        ws.cell(row=7, column=1, value="- Reliabilität basiert auf ursprünglichen Segment-IDs")
        ws.cell(row=8, column=1, value="- Mehrfachkodierungen werden als Set-Variable behandelt")
        ws.cell(row=9, column=1, value="- Kategorie-spezifische Segmentierung fuer Review")
        ws.cell(row=10, column=1, value=f"- Review-Modus: {export_mode}")
    
    def _create_categories_sheet(self, workbook, categories: Dict, original_categories: Dict):
        """
        BESTEHENDE METHODE: Erstellt Kategorien-Übersicht (bleibt unverÄndert)
        """
        ws = workbook.create_sheet("Kategorien")
        
        # Header
        headers = ['Hauptkategorie', 'Typ', 'Definition', 'Anzahl Subkategorien', 'Subkategorien']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Daten
        row = 2
        for cat_name, cat_obj in categories.items():
            cat_type = self._determine_category_type(cat_name, original_categories)
            definition = getattr(cat_obj, 'definition', '')
            subcats = getattr(cat_obj, 'subcategories', [])
            
            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=2, value=cat_type)
            ws.cell(row=row, column=3, value=definition)
            ws.cell(row=row, column=4, value=len(subcats))
            ws.cell(row=row, column=5, value=', '.join(subcats))
            
            row += 1
    
    def _create_multiple_coding_analysis_sheet(self, workbook, codings: List[Dict]):
        """
        NEUE METHODE: Erstellt detaillierte Mehrfachkodierungs-Analyse
        """
        ws = workbook.create_sheet("Mehrfachkodierungs_Analyse")
        
        # Analysiere Mehrfachkodierungen
        original_segments = defaultdict(list)
        for coding in codings:
            consensus_info = coding.get('consensus_info', {})
            original_id = consensus_info.get('original_segment_id', coding.get('segment_id', ''))
            original_segments[original_id].append(coding)
        
        # Übersicht
        ws.cell(row=1, column=1, value="Mehrfachkodierungs-Analyse").font = Font(bold=True, size=14)
        
        # Statistiken
        total_original = len(original_segments)
        multiple_coded = sum(1 for segs in original_segments.values() if len(segs) > 1)
        single_coded = total_original - multiple_coded
        expansion_factor = len(codings) / total_original if total_original > 0 else 1
        
        ws.cell(row=3, column=1, value="Gesamtstatistik:").font = Font(bold=True)
        ws.cell(row=4, column=1, value="- UrsprÜngliche Segmente:")
        ws.cell(row=4, column=2, value=total_original)
        ws.cell(row=5, column=1, value="- Einzelkodierungen:")
        ws.cell(row=5, column=2, value=single_coded)
        ws.cell(row=6, column=1, value="- Mehrfachkodierungen:")
        ws.cell(row=6, column=2, value=multiple_coded)
        ws.cell(row=7, column=1, value="- Finale Segmente:")
        ws.cell(row=7, column=2, value=len(codings))
        ws.cell(row=8, column=1, value="- Expansionsfaktor:")
        ws.cell(row=8, column=2, value=f"{expansion_factor:.2f}")
        
        # Detaillierte Liste der Mehrfachkodierungen
        ws.cell(row=10, column=1, value="Mehrfachkodierungs-Details:").font = Font(bold=True)
        
        headers = ['Original_Segment', 'Anzahl_Kategorien', 'Kategorien', 'Neue_Segment_IDs']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 12
        for original_id, segment_codings in original_segments.items():
            if len(segment_codings) > 1:
                categories = [c.get('category', '') for c in segment_codings]
                unique_categories = list(set(categories))
                segment_ids = [c.get('segment_id', '') for c in segment_codings]
                
                ws.cell(row=row, column=1, value=original_id)
                ws.cell(row=row, column=2, value=len(unique_categories))
                ws.cell(row=row, column=3, value=', '.join(unique_categories))
                ws.cell(row=row, column=4, value=', '.join(segment_ids))
                
                row += 1
    
    async def _export_json_results(self, codings: List[Dict], reliability: float, 
                                 categories: Dict, export_mode: str):
        """
        BESTEHENDE METHODE: JSON-Export (bleibt unverÄndert)
        """
        import json
        from pathlib import Path
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"qca_ergebnisse_{export_mode}_{timestamp}.json"
        output_path = Path(self.output_dir) / filename
        
        export_data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'export_mode': export_mode,
                'total_codings': len(codings),
                'reliability': reliability
            },
            'codings': codings,
            'categories': {name: {
                'definition': getattr(cat, 'definition', ''),
                'subcategories': getattr(cat, 'subcategories', [])
            } for name, cat in categories.items()}
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        print(f"ℹ️ JSON-Export erstellt: {output_path}")
       
    def _extract_doc_and_chunk_id(self, segment_id: str) -> tuple:
        """
        Extrahiert Dokumentnamen und Chunk-ID aus einer Segment-ID.
        
        Args:
            segment_id: Segment-ID im Format "dokument_chunk_X"
                
        Returns:
            tuple: (doc_name, chunk_id)
        """
        parts = segment_id.split('_chunk_')
        if len(parts) == 2:
            return parts[0], parts[1]
        else:
            # Fallback fuer ungültige Segment-IDs
            return segment_id, "unknown"

    def _get_base_segment_id(self, coding: Dict) -> str:
        """
        FIX: Extrahiert Basis-Segment-ID ohne Mehrfachkodierungs-Suffix
        FÜr ResultsExporter Klasse
        """
        segment_id = coding.get('segment_id', '')
        
        # Entferne Mehrfachkodierungs-Suffixe (Format: "doc_chunk_5-1")
        if '-' in segment_id:
            parts = segment_id.rsplit('-', 1)
            if len(parts) == 2 and parts[1].isdigit():
                return parts[0]
        
        return segment_id
    
    def _export_reliability_sheet(self, writer, reliability: float, export_mode: str):
        """
        PUNKT 7: Formatiertes Reliabilitäts-Sheet
        """
        # Erstelle formatierte Reliabilitätsdaten
        reliability_data = [
            ['Metrik', 'Wert', 'Bewertung'],
            ['Krippendorff\'s Alpha', f'{reliability:.3f}', self._get_reliability_rating(reliability)],
            ['Export-Modus', export_mode, ''],
            ['Berechnung', 'Set-basiert fuer Mehrfachkodierungen', ''],
            ['Basis', 'UrsprÜngliche Segment-IDs', '']
        ]
        
        # Als DataFrame exportieren
        df_rel = pd.DataFrame(reliability_data[1:], columns=reliability_data[0])
        df_rel.to_excel(writer, sheet_name='Reliabilität', index=False)
        
        # Formatierung anwenden
        worksheet = writer.sheets['Reliabilität']
        self._apply_professional_formatting(worksheet, df_rel)
        
        # Spezielle Farbkodierung fuer Alpha-Wert
        alpha_cell = worksheet.cell(row=2, column=2)  # Alpha-Wert
        if reliability > 0.8:
            alpha_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        elif reliability > 0.667:
            alpha_cell.fill = PatternFill(start_color='FFFF90', end_color='FFFF90', fill_type='solid')
        else:
            alpha_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    def _get_reliability_rating(self, reliability: float) -> str:
        """Bestimmt Reliabilitäts-Bewertung"""
        if reliability > 0.8:
            return "Exzellent"
        elif reliability > 0.667:
            return "Akzeptabel"
        elif reliability > 0.5:
            return "Schwach"
        else:
            return "Unzureichend"
    
    def _export_categories_sheet_formatted(self, writer, categories: Dict, original_categories: Dict):
        """
        Formatiertes Kategorien-Sheet
        """
        cat_data = []
        
        for cat_name, cat_obj in categories.items():
            cat_type = self._determine_category_type(cat_name, original_categories)
            definition = getattr(cat_obj, 'definition', '')
            subcats = getattr(cat_obj, 'subcategories', [])
            
            cat_data.append({
                'Hauptkategorie': cat_name,
                'Typ': cat_type,
                'Definition': definition,
                'Anzahl_Subkategorien': len(subcats),
                'Subkategorien': ', '.join(subcats)
            })
        
        if cat_data:
            df_cats = pd.DataFrame(cat_data)
            df_cats.to_excel(writer, sheet_name='Kategorien', index=False)
            
            # Formatierung
            worksheet = writer.sheets['Kategorien']
            self._apply_professional_formatting(worksheet, df_cats)
            
            # Spaltenbreiten anpassen
            worksheet.column_dimensions['A'].width = 25  # Hauptkategorie
            worksheet.column_dimensions['B'].width = 12  # Typ
            worksheet.column_dimensions['C'].width = 50  # Definition
            worksheet.column_dimensions['D'].width = 15  # Anzahl
            worksheet.column_dimensions['E'].width = 40  # Subkategorien
    
    def _export_frequency_analysis(self, writer, df_coded: pd.DataFrame, attribut1_label: str, attribut2_label: str) -> None:
        try:
            # Hole alle DatensÄtze, auch "Nicht kodiert"
            df_all = df_coded.copy()
            
            # FIX: Normalize category strings in DataFrame to prevent matching issues
            import unicodedata
            if 'Hauptkategorie' in df_all.columns:
                df_all['Hauptkategorie'] = df_all['Hauptkategorie'].apply(
                    lambda x: unicodedata.normalize('NFKC', str(x).strip()) if x is not None else x
                )
            
            # DEBUG: Check for None values in categories
            none_count = df_all['Hauptkategorie'].isnull().sum()
            if none_count > 0:
                print(f"⚠️ DEBUG: Found {none_count} None/NaN values in Hauptkategorie column")
                print(f"   Sample rows with None categories:")
                none_rows = df_all[df_all['Hauptkategorie'].isnull()].head(3)
                for idx, row in none_rows.iterrows():
                    print(f"   - Row {idx}: Segment={row.get('Segment_ID', 'N/A')}, Coder={row.get('Coder_ID', 'N/A')}")
                
                # Filter out rows with None categories to prevent pivot table issues
                print(f"   🔧 Filtering out {none_count} rows with None categories")
                df_all = df_all[df_all['Hauptkategorie'].notna()]
            
            # Hole eindeutige Hauptkategorien, inkl. "Nicht kodiert" (filter out None values)
            all_categories = df_all['Hauptkategorie'].unique()
            
            # FIX: Normalize category strings to prevent matching issues
            import unicodedata
            main_categories = []
            for cat in all_categories:
                if cat is not None:
                    # Normalize unicode and strip whitespace
                    normalized_cat = unicodedata.normalize('NFKC', str(cat).strip())
                    main_categories.append(normalized_cat)
            
            print(f"   📊 Categories found: {len(all_categories)} total, {len(main_categories)} valid (filtered {len(all_categories) - len(main_categories)} None values)")
            
            # FIX: Stelle sicher, dass category_colors bereits initialisiert ist
            # Falls nicht, initialisiere sie mit derselben Logik wie in _export_main_results
            if not self.category_colors:
                print("⚠️ Kategorie-Farben noch nicht initialisiert, initialisiere jetzt...")
                self._initialize_category_colors(df_all)
                print(f"   ✅ {len(self.category_colors)} Kategorie-Farben initialisiert")
            else:
                print(f"   ✅ Verwende bereits initialisierte {len(self.category_colors)} Kategorie-Farben")
                
                # DEBUG: Show existing category colors
                print(f"   🔍 DEBUG: Existing category colors:")
                for cat, color in self.category_colors.items():
                    print(f"      '{cat}': {color}")
                
                # print(f"   🔍 DEBUG: Categories found in frequency data:")
                for cat in main_categories:
                    is_in_colors = cat in self.category_colors
                    # print(f"      '{cat}' (in colors: {is_in_colors})")
                    
                    # DEBUG: Detailed string analysis for problematic categories
                    if not is_in_colors and cat is not None:
                        # print(f"         🔍 String analysis for '{cat}':")
                        # print(f"            Length: {len(cat)}")
                        # print(f"            Repr: {repr(cat)}")
                        # print(f"            Encoded: {cat.encode('utf-8')}")
                        
                        # Check for similar strings in category_colors
                        for existing_cat in self.category_colors.keys():
                            if existing_cat.strip().lower() == cat.strip().lower():
                                # print(f"            ⚠️ Case/whitespace mismatch with: '{existing_cat}'")
                                # print(f"               Existing repr: {repr(existing_cat)}")
                                # print(f"               Existing encoded: {existing_cat.encode('utf-8')}")
                                pass
                            elif existing_cat.replace(' ', '') == cat.replace(' ', ''):
                                # print(f"            ⚠️ Space difference with: '{existing_cat}'")
                                pass
                        
                        # Check for exact matches with different encoding
                        for existing_cat in self.category_colors.keys():
                            try:
                                if existing_cat.encode('utf-8') == cat.encode('utf-8'):
                                    # print(f"            ⚠️ Encoding match but string comparison failed: '{existing_cat}'")
                                    pass
                            except:
                                pass
                
                # FIX: Check if any categories from frequency data are missing colors
                # Use robust string matching to handle encoding/whitespace issues
                missing_categories = []
                for cat in main_categories:
                    if cat is None:
                        continue
                    
                    # First try exact match
                    if cat in self.category_colors:
                        continue
                    
                    # Try normalized matching (strip whitespace, normalize unicode, normalize case)
                    import unicodedata
                    cat_normalized = unicodedata.normalize('NFKC', str(cat).strip())
                    found_match = False
                    
                    for existing_cat in self.category_colors.keys():
                        existing_normalized = unicodedata.normalize('NFKC', str(existing_cat).strip())
                        
                        # Try exact normalized match
                        if existing_normalized == cat_normalized:
                            print(f"   🔧 Unicode/whitespace mismatch fixed: '{cat}' -> '{existing_cat}'")
                            self.category_colors[cat] = self.category_colors[existing_cat]
                            found_match = True
                            break
                        
                        # Try case-insensitive normalized match
                        elif existing_normalized.lower() == cat_normalized.lower():
                            print(f"   🔧 Case/unicode mismatch fixed: '{cat}' -> '{existing_cat}'")
                            self.category_colors[cat] = self.category_colors[existing_cat]
                            found_match = True
                            break
                    
                    if not found_match:
                        missing_categories.append(cat)
                
                if missing_categories:
                    print(f"\n⚠️ Fehlende Kategorien in Farbzuordnung gefunden: {missing_categories}")
                    
                    # CRITICAL FIX: Don't generate new colors, this breaks consistency!
                    # Instead, log the issue and use a default color
                    for cat in missing_categories:
                        # Use a consistent default color for missing categories
                        self.category_colors[cat] = 'FFE6E6'  # Light pink as fallback
                        print(f"  + {cat}: FFE6E6 (Fallback-Farbe für fehlende Kategorie)")
                        print(f"    ⚠️ WARNUNG: Diese Kategorie sollte bereits in den Hauptergebnissen definiert sein!")
                
                # Stelle sicher dass "Nicht kodiert" grau ist
                if 'Nicht kodiert' in main_categories and 'Nicht kodiert' not in self.category_colors:
                    self.category_colors['Nicht kodiert'] = 'CCCCCC'
                    print(f"  + Nicht kodiert: CCCCCC (Standard-Grau)")
            
            # FIX: Debug-Ausgabe der verwendeten Farben (filter out None values)
            print(f"\n🎨 Verwende Kategorie-Farben für Häufigkeitsanalyse:")
            valid_categories = [cat for cat in main_categories if cat is not None]
            for cat in sorted(valid_categories):
                color = self.category_colors.get(cat, 'FFFFFF')
                print(f"  - {cat}: {color}")
            
            # FIX: Vergleiche mit Hauptergebnisse-Farben
            print(f"\n🔍 Farbkonsistenz-Check:")
            print(f"  - Hauptergebnisse haben {len(self.category_colors)} Farben definiert")
            print(f"  - Häufigkeitsanalyse benötigt {len(main_categories)} Farben")
            missing_in_main = [cat for cat in main_categories if cat not in self.category_colors]
            if missing_in_main:
                print(f"  ⚠️ Fehlende Farben: {missing_in_main}")
            else:
                print(f"  ✅ Alle Kategorien haben definierte Farben")

            if 'Häufigkeitsanalysen' not in writer.sheets:
                writer.book.create_sheet('Häufigkeitsanalysen')
            
            worksheet = writer.sheets['Häufigkeitsanalysen']
            worksheet.delete_rows(1, worksheet.max_row)  # Bestehende Daten löschen

            current_row = 1
            
            from openpyxl.styles import Font
            title_font = Font(bold=True, size=12)

            # 1. Hauptkategorien nach Dokumenten
            cell = worksheet.cell(row=current_row, column=1, value="1. Verteilung der Hauptkategorien")
            cell.font = title_font
            current_row += 2

            # FIX: Verwende dieselbe einfache Methode wie fuer Subkategorien
            pivot_main = pd.pivot_table(
                df_all,
                index=['Hauptkategorie'],
                columns='Dokument',
                values='Chunk_Nr',
                aggfunc='count',
                margins=True,
                margins_name='Gesamt',
                fill_value=0
            )

            # FIX: Konvertiere zu DataFrame fuer einfache Ausgabe
            temp_df_main = pivot_main.copy().reset_index()
            
            # Formatierte Spaltenbezeichnungen
            formatted_columns = []
            for col in pivot_main.columns:
                if isinstance(col, tuple):
                    col_parts = [str(part) for part in col if part and part != '']
                    formatted_columns.append(' - '.join(col_parts))
                else:
                    formatted_columns.append(str(col))
            
            # FIX: Erstelle Header-Zeile mit korrekten Spaltennamen wie bei Subkategorien
            headers = ['Hauptkategorie'] + formatted_columns
            for col_idx, header in enumerate(headers):
                worksheet.cell(row=current_row, column=col_idx+1, value=header)
            
            # FIX: Exportiere Daten-Zeilen (beginne bei current_row + 1)
            for row_idx, (index, row) in enumerate(temp_df_main.iterrows()):
                for col_idx, value in enumerate(row):
                    worksheet.cell(row=current_row + 1 + row_idx, column=col_idx+1, value=value)
            
            # Formatiere den Bereich (Header + Daten)
            self._apply_professional_formatting_to_range(worksheet, current_row, 1, len(temp_df_main) + 1, len(headers))
            
            # FIX: Verbesserte Farbkodierung fuer Hauptkategorien - verwende exakt dieselben Farben
            for row_idx in range(1, len(temp_df_main) + 1):
                kategorie = temp_df_main.iloc[row_idx-1]['Hauptkategorie']
                if kategorie != 'Gesamt' and kategorie in self.category_colors:
                    color = self.category_colors[kategorie]
                    from openpyxl.styles import PatternFill
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    worksheet.cell(row=current_row + row_idx, column=1).fill = fill
                    print(f"  🎨 Farbe angewendet: {kategorie} -> {color}")
            
            current_row += len(temp_df_main) + 3
            # FIX: Ende - einfache Methode wie bei Subkategorien

            # 2. Subkategorien-Hierarchie (nur fuer kodierte Segmente)
            cell = worksheet.cell(row=current_row, column=1, value="2. Subkategorien nach Hauptkategorien")
            cell.font = title_font
            current_row += 2

            # Filtere "Nicht kodiert" fuer Subkategorien-Analyse aus
            df_sub = df_all[df_all['Hauptkategorie'] != "Nicht kodiert"].copy()
            df_sub['Subkategorie'] = df_sub['Subkategorien'].str.split(', ')
            df_sub = df_sub.explode('Subkategorie')
            
            # Erstelle Pivot-Tabelle
            pivot_sub = pd.pivot_table(
                df_sub,
                index=['Hauptkategorie', 'Subkategorie'],
                columns='Dokument',
                values='Chunk_Nr',
                aggfunc='count',
                margins=True,
                margins_name='Gesamt',
                fill_value=0
            )

            # DataFrame fuer Subkategorien mit korrekten Spaltennamen
            temp_df_sub = pivot_sub.copy().reset_index()
            
            # Formatierte Spaltenbezeichnungen
            formatted_columns = []
            for col in pivot_sub.columns:
                if isinstance(col, tuple):
                    col_parts = [str(part) for part in col if part and part != '']
                    formatted_columns.append(' - '.join(col_parts))
                else:
                    formatted_columns.append(str(col))
            
            # Erstelle Header-Zeile mit korrekten Spaltennamen
            headers = ['Hauptkategorie', 'Subkategorie'] + formatted_columns
            for col_idx, header in enumerate(headers):
                worksheet.cell(row=current_row, column=col_idx+1, value=header)
            
            # Exportiere Daten-Zeilen (beginne bei current_row + 1)
            for row_idx, (index, row) in enumerate(temp_df_sub.iterrows()):
                for col_idx, value in enumerate(row):
                    worksheet.cell(row=current_row + 1 + row_idx, column=col_idx+1, value=value)
            
            # Formatiere den Bereich (Header + Daten)
            self._apply_professional_formatting_to_range(worksheet, current_row, 1, len(temp_df_sub) + 1, len(headers))
            
            # Zusätzliche Farbkodierung fuer Hauptkategorien in Subkategorien-Tabelle
            for row_idx in range(1, len(temp_df_sub) + 1):
                kategorie = temp_df_sub.iloc[row_idx-1]['Hauptkategorie']
                if kategorie != 'Gesamt' and kategorie in self.category_colors:
                    color = self.category_colors[kategorie]
                    from openpyxl.styles import PatternFill
                    worksheet.cell(row=current_row + row_idx, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            current_row += len(temp_df_sub) + 2

            # 3. Attribut-Analysen
            cell = worksheet.cell(row=current_row, column=1, value="3. Verteilung nach Attributen")
            cell.font = title_font
            current_row += 2

            # 3.1 Attribut 1
            cell = worksheet.cell(row=current_row, column=1, value=f"3.1 Verteilung nach {attribut1_label}")
            cell.font = title_font
            current_row += 1

            attr1_counts = df_coded[attribut1_label].value_counts()
            attr1_counts['Gesamt'] = attr1_counts.sum()
            
            attr1_data = [[attribut1_label, 'Anzahl']] + [[idx, value] for idx, value in attr1_counts.items()]
            
            for row_idx, row_data in enumerate(attr1_data):
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
            
            self._apply_professional_formatting_to_range(worksheet, current_row, 1, len(attr1_data), 2)
            current_row += len(attr1_data) + 3

            # 3.2 Attribut 2
            cell = worksheet.cell(row=current_row, column=1, value=f"3.2 Verteilung nach {attribut2_label}")
            cell.font = title_font
            current_row += 1

            attr2_counts = df_coded[attribut2_label].value_counts()
            attr2_counts['Gesamt'] = attr2_counts.sum()
            
            attr2_data = [[attribut2_label, 'Anzahl']] + [[idx, value] for idx, value in attr2_counts.items()]
            
            for row_idx, row_data in enumerate(attr2_data):
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
            
            self._apply_professional_formatting_to_range(worksheet, current_row, 1, len(attr2_data), 2)
            current_row += len(attr2_data) + 3

            # 3.3 Attribut 3 (nur wenn definiert)
            attribut3_label = self.attribute_labels.get('attribut3', '')
            if attribut3_label and attribut3_label in df_coded.columns:
                cell = worksheet.cell(row=current_row, column=1, value=f"3.3 Verteilung nach {attribut3_label}")
                cell.font = title_font
                current_row += 1

                attr3_counts = df_coded[attribut3_label].value_counts()
                attr3_counts['Gesamt'] = attr3_counts.sum()
                
                attr3_data = [[attribut3_label, 'Anzahl']] + [[idx, value] for idx, value in attr3_counts.items()]
                
                for row_idx, row_data in enumerate(attr3_data):
                    for col_idx, value in enumerate(row_data):
                        worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
                
                self._apply_professional_formatting_to_range(worksheet, current_row, 1, len(attr3_data), 2)
                current_row += len(attr3_data) + 3

            # FIX: 3.4 Kreuztabellen der Attribute wieder hinzugefÜgt
            cell = worksheet.cell(row=current_row, column=1, value="3.4 Kreuztabelle der Attribute")
            cell.font = title_font
            current_row += 1

            # Kreuztabelle 1-2
            cross_tab = pd.crosstab(
                df_coded[attribut1_label], 
                df_coded[attribut2_label],
                margins=True,
                margins_name='Gesamt'
            )
            
            cross_tab_df = cross_tab.copy().reset_index()
            cross_tab_df.columns.name = None
            
            # FIX: Spaltenüberschriften korrekt exportieren
            # Erste Zeile: Spaltenüberschriften
            col_headers = [attribut1_label] + list(cross_tab_df.columns[1:])
            for col_idx, header in enumerate(col_headers):
                worksheet.cell(row=current_row, column=col_idx+1, value=header)
            current_row += 1
            
            # Datenzeilen
            for row_idx, (index, row) in enumerate(cross_tab_df.iterrows()):
                for col_idx, value in enumerate(row):
                    worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
            
            self._apply_professional_formatting_to_range(worksheet, current_row - 1, 1, len(cross_tab_df) + 1, len(cross_tab_df.columns))
            current_row += len(cross_tab_df) + 3

            # Weitere Kreuztabellen fuer Attribut 3, wenn vorhanden
            if attribut3_label and attribut3_label in df_coded.columns:
                # Kreuztabelle 1-3
                cross_tab_1_3 = pd.crosstab(
                    df_coded[attribut1_label], 
                    df_coded[attribut3_label],
                    margins=True,
                    margins_name='Gesamt'
                )
                
                cross_tab_1_3_df = cross_tab_1_3.copy().reset_index()
                cross_tab_1_3_df.columns.name = None
                
                # FIX: Spaltenüberschriften korrekt exportieren
                col_headers_1_3 = [attribut1_label] + list(cross_tab_1_3_df.columns[1:])
                for col_idx, header in enumerate(col_headers_1_3):
                    worksheet.cell(row=current_row, column=col_idx+1, value=header)
                current_row += 1
                
                for row_idx, (index, row) in enumerate(cross_tab_1_3_df.iterrows()):
                    for col_idx, value in enumerate(row):
                        worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
                
                self._apply_professional_formatting_to_range(worksheet, current_row - 1, 1, len(cross_tab_1_3_df) + 1, len(cross_tab_1_3_df.columns))
                current_row += len(cross_tab_1_3_df) + 3
                
                # Kreuztabelle 2-3
                cross_tab_2_3 = pd.crosstab(
                    df_coded[attribut2_label], 
                    df_coded[attribut3_label],
                    margins=True,
                    margins_name='Gesamt'
                )
                
                cross_tab_2_3_df = cross_tab_2_3.copy().reset_index()
                cross_tab_2_3_df.columns.name = None
                
                # FIX: Spaltenüberschriften korrekt exportieren
                col_headers_2_3 = [attribut2_label] + list(cross_tab_2_3_df.columns[1:])
                for col_idx, header in enumerate(col_headers_2_3):
                    worksheet.cell(row=current_row, column=col_idx+1, value=header)
                current_row += 1
                
                for row_idx, (index, row) in enumerate(cross_tab_2_3_df.iterrows()):
                    for col_idx, value in enumerate(row):
                        worksheet.cell(row=current_row + row_idx, column=col_idx+1, value=value)
                
                self._apply_professional_formatting_to_range(worksheet, current_row - 1, 1, len(cross_tab_2_3_df) + 1, len(cross_tab_2_3_df.columns))
                current_row += len(cross_tab_2_3_df) + 3
            
            print("✅ Häufigkeitsanalysen erfolgreich mit standardisierter Formatierung exportiert")
            
        except Exception as e:
            print(f"⚠️ Fehler bei Häufigkeitsanalysen: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _apply_professional_formatting_to_range(self, worksheet, start_row: int, start_col: int, num_rows: int, num_cols: int) -> None:
        """
        FIX: Hilfsmethode fuer die Formatierung eines bestimmten Bereichs ohne StyleProxy-Probleme
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header formatieren (erste Zeile des Bereichs)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center')
        
        for col in range(start_col, start_col + num_cols):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Rahmen fuer alle Zellen
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, start_row + num_rows):
            for col in range(start_col, start_col + num_cols):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        
        # Abwechselnde ZeilenfÄrbung (ab zweiter Zeile)
        alternate_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for row_num in range(start_row + 1, start_row + num_rows):
            if (row_num - start_row) % 2 == 0:
                for col in range(start_col, start_col + num_cols):
                    cell = worksheet.cell(row=row_num, column=col)
                    if not cell.fill or cell.fill.start_color.rgb == 'FFFFFF':
                        cell.fill = alternate_fill
    
    def _create_formatted_attribute_analysis(self, writer, df_coded, attribute_label):
        """Erstellt formatierte Attribut-Analyse"""
        try:
            # Kreuztabelle erstellen
            crosstab = pd.crosstab(df_coded[attribute_label], df_coded['Hauptkategorie'], margins=True)
            
            if not crosstab.empty:
                sheet_name = f'Analyse_{attribute_label}'[:31]  # Excel Sheet-Name Limit
                crosstab.to_excel(writer, sheet_name=sheet_name)
                
                # Formatierung
                worksheet = writer.sheets[sheet_name]
                self._apply_crosstable_formatting(worksheet, crosstab)
                
        except Exception as e:
            print(f"⚠️ Fehler bei {attribute_label}-Analyse: {str(e)}")
    
    def _create_formatted_crosstable(self, writer, df_coded, attr1_label, attr2_label):
        """Erstellt formatierte Kreuztabelle"""
        try:
            crosstab = pd.crosstab(df_coded[attr1_label], df_coded[attr2_label], margins=True)
            
            if not crosstab.empty:
                crosstab.to_excel(writer, sheet_name='Kreuztabelle_Attribute')
                
                # Formatierung
                worksheet = writer.sheets['Kreuztabelle_Attribute']
                self._apply_crosstable_formatting(worksheet, crosstab)
                
        except Exception as e:
            print(f"⚠️ Fehler bei Kreuztabelle: {str(e)}")
    
    def _export_progressive_summaries(self, writer, document_summaries):
        """
        PUNKT 7: Formatierte Progressive Summaries
        """
        try:
            print("🔀 Erstelle formatierte Progressive Summaries...")
            
            if not document_summaries:
                print("ℹ️ Keine Document-Summaries verfügbar")
                return
            
            summary_data = []
            for doc_name, summary in document_summaries.items():
                # FIX: Verwende 'summary' statt 'self.summary'
                clean_summary = summary.replace('\n', ' ').replace('\r', ' ').strip()
                
                
                summary_data.append({
                    'Dokument': doc_name,
                    'Finales_Summary': clean_summary,
                    'Wortanzahl': len(summary.split()),
                    'Zeichenanzahl': len(summary),
                    'Durchschnittliche_WortlÄnge': f"{len(summary)/len(summary.split()):.1f}" if summary.split() else "0"
                })
            
            if summary_data:
                df_summaries = pd.DataFrame(summary_data)
                df_summaries.to_excel(writer, sheet_name='Progressive_Summaries', index=False)
                
                # Formatierung
                worksheet = writer.sheets['Progressive_Summaries']
                self._apply_professional_formatting(worksheet, df_summaries)
                
                # Spaltenbreiten anpassen
                worksheet.column_dimensions['A'].width = 25  # Dokument
                worksheet.column_dimensions['B'].width = 80  # Summary (breit aber ohne wrapping)
                worksheet.column_dimensions['C'].width = 12  # Wortanzahl
                worksheet.column_dimensions['D'].width = 15  # Zeichenanzahl
                worksheet.column_dimensions['E'].width = 20  # Durchschnittliche WortlÄnge
                
                print(f"✅ {len(summary_data)} formatierte Document-Summaries exportiert")
            
        except Exception as e:
            print(f"⚠️ Fehler bei formatierten Progressive Summaries: {str(e)}")
    
    def _calculate_review_statistics(self, codings: List[Dict], export_mode: str, original_codings: List[Dict] = None) -> Dict[str, int]:
        """
        # FIX: Erweiterte Berechnung der Review-Statistiken mit Vor-/Nach-Review Vergleich
        Berechnet umfassende Statistiken des Review-Prozesses.
        
        Args:
            codings: Liste der finalen Kodierungen nach Review
            export_mode: Verwendeter Review-Modus
            original_codings: Liste der ursprünglichen Kodierungen vor Review (optional)
            
        Returns:
            Dict[str, int]: Erweiterte Statistiken des Review-Prozesses
        """
        stats = {
            'consensus_found': 0,
            'majority_found': 0,
            'manual_priority': 0,
            'no_consensus': 0,
            'single_coding': 0,
            'multiple_coding_consolidated': 0,
            # FIX: Neue Statistiken fuer bessere Übersicht
            'segments_before_review': 0,
            'segments_after_review': 0,
            'segments_with_conflicts': 0,
            'segments_resolved': 0,
            'total_original_codings': 0,
            'categories_involved': 0
        }
        
        # FIX: Analysiere ursprüngliche Kodierungen vor Review
        if original_codings:
            stats['total_original_codings'] = len(original_codings)
            
            # Gruppiere ursprüngliche Kodierungen nach Segmenten
            from collections import defaultdict
            original_segments = defaultdict(list)
            all_categories = set()
            
            for coding in original_codings:
                segment_id = coding.get('segment_id', '')
                category = coding.get('category', '')
                if segment_id and category not in ['Nicht kodiert', 'Kein Kodierkonsens']:
                    original_segments[segment_id].append(coding)
                    all_categories.add(category)
            
            stats['segments_before_review'] = len(original_segments)
            stats['categories_involved'] = len(all_categories)
            
            # ZÄhle Segmente mit Konflikten (mehrere verschiedene Kategorien)
            for segment_id, segment_codings in original_segments.items():
                categories = [c.get('category', '') for c in segment_codings]
                unique_categories = set(categories)
                if len(unique_categories) > 1:
                    stats['segments_with_conflicts'] += 1
        # FIX: Ende
        
        stats['segments_after_review'] = len(codings)
        
        for coding in codings:
            # Bestimme den Typ der Kodierung basierend auf verfügbaren Informationen
            if coding.get('manual_review', False):
                stats['manual_priority'] += 1
            elif coding.get('consolidated_from_multiple', False):
                stats['multiple_coding_consolidated'] += 1
            elif coding.get('consensus_info', {}).get('selection_type') == 'consensus':
                stats['consensus_found'] += 1
                stats['segments_resolved'] += 1  # FIX: ZÄhle als aufgelöst
            elif coding.get('consensus_info', {}).get('selection_type') == 'majority':
                stats['majority_found'] += 1
                stats['segments_resolved'] += 1  # FIX: ZÄhle als aufgelöst
            elif coding.get('consensus_info', {}).get('selection_type') == 'no_consensus':
                stats['no_consensus'] += 1
            elif coding.get('category') == 'Kein Kodierkonsens':
                stats['no_consensus'] += 1
            else:
                stats['single_coding'] += 1
        
        return stats

    def _export_review_statistics(self, writer, review_stats: Dict, export_mode: str):
        """
        # FIX: Erweiterte Exportfunktion fuer Review-Statistiken mit detaillierteren Informationen
        Exportiert umfassende Statistiken des Review-Prozesses in ein separates Excel-Sheet.
        
        Args:
            writer: Excel Writer Objekt
            review_stats: Erweiterte Statistiken des Review-Prozesses
            export_mode: Verwendeter Review-Modus
        """
        try:
            if 'Review_Statistiken' not in writer.sheets:
                writer.book.create_sheet('Review_Statistiken')
                
            worksheet = writer.sheets['Review_Statistiken']
            current_row = 1
            
            # Titel
            worksheet.cell(row=current_row, column=1, value=f"Review-Statistiken ({export_mode}-Modus)")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            current_row += 2
            
            # FIX: Übersichts-Statistiken vor Detail-AufschlÜsselung
            worksheet.cell(row=current_row, column=1, value="ÜBERSICHT")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            # Grundlegende Zahlen
            basic_stats = [
                ("UrsprÜngliche Kodierungen", review_stats.get('total_original_codings', 0)),
                ("Segmente vor Review", review_stats.get('segments_before_review', 0)),
                ("Segmente nach Review", review_stats.get('segments_after_review', 0)),
                ("Segmente mit Konflikten", review_stats.get('segments_with_conflicts', 0)),
                ("Segmente aufgelöst", review_stats.get('segments_resolved', 0)),
                ("Involvierte Kategorien", review_stats.get('categories_involved', 0))
            ]
            
            for label, value in basic_stats:
                if value > 0:  # Nur anzeigen wenn Werte vorhanden
                    worksheet.cell(row=current_row, column=1, value=label)
                    worksheet.cell(row=current_row, column=2, value=value)
                    current_row += 1
            
            current_row += 1
            # FIX: Ende Übersichts-Sektion
            
            # Header fuer Detail-AufschlÜsselung
            worksheet.cell(row=current_row, column=1, value="REVIEW-TYPEN AUFSCHLÜSSELUNG")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Review-Typ")
            worksheet.cell(row=current_row, column=2, value="Anzahl")
            worksheet.cell(row=current_row, column=3, value="Prozent")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            worksheet.cell(row=current_row, column=2).font = Font(bold=True)
            worksheet.cell(row=current_row, column=3).font = Font(bold=True)
            current_row += 1
            
            # FIX: Berechne Gesamtsumme nur der Review-spezifischen Kategorien
            review_specific_stats = {
                'consensus_found': review_stats.get('consensus_found', 0),
                'majority_found': review_stats.get('majority_found', 0),
                'manual_priority': review_stats.get('manual_priority', 0),
                'no_consensus': review_stats.get('no_consensus', 0),
                'single_coding': review_stats.get('single_coding', 0),
                'multiple_coding_consolidated': review_stats.get('multiple_coding_consolidated', 0)
            }
            
            total_reviewed = sum(review_specific_stats.values())
            
            for stat_name, count in review_specific_stats.items():
                if count > 0:
                    # Übersetze Statistik-Namen
                    german_names = {
                        'consensus_found': 'Konsens gefunden',
                        'majority_found': 'Mehrheit gefunden', 
                        'manual_priority': 'Manuelle PrioritÄt',
                        'no_consensus': 'Kein Konsens',
                        'single_coding': 'Einzelkodierung',
                        'multiple_coding_consolidated': 'Mehrfachkodierung konsolidiert'
                    }
                    
                    display_name = german_names.get(stat_name, stat_name)
                    worksheet.cell(row=current_row, column=1, value=display_name)
                    worksheet.cell(row=current_row, column=2, value=count)
                    
                    # Prozentangabe
                    if total_reviewed > 0:
                        percentage = (count / total_reviewed) * 100
                        worksheet.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
                    
                    current_row += 1
            
            # Gesamtsumme
            worksheet.cell(row=current_row, column=1, value="GESAMT")
            worksheet.cell(row=current_row, column=2, value=total_reviewed)
            worksheet.cell(row=current_row, column=3, value="100.0%")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            worksheet.cell(row=current_row, column=2).font = Font(bold=True)
            worksheet.cell(row=current_row, column=3).font = Font(bold=True)
            
            # FIX: Zusätzliche Analyse-Sektion
            current_row += 2
            worksheet.cell(row=current_row, column=1, value="REVIEW-EFFIZIENZ")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            # Effizienz-Berechnung
            conflicts = review_stats.get('segments_with_conflicts', 0)
            resolved = review_stats.get('segments_resolved', 0)
            
            if conflicts > 0:
                resolution_rate = (resolved / conflicts) * 100
                worksheet.cell(row=current_row, column=1, value="Konflikt-Auflösungsrate")
                worksheet.cell(row=current_row, column=2, value=f"{resolution_rate:.1f}%")
                current_row += 1
            
            # Reduzierungsrate berechnen
            original_count = review_stats.get('total_original_codings', 0)
            final_count = review_stats.get('segments_after_review', 0)
            
            if original_count > 0 and final_count > 0:
                reduction_rate = ((original_count - final_count) / original_count) * 100
                worksheet.cell(row=current_row, column=1, value="Kodierungs-Reduzierung")
                worksheet.cell(row=current_row, column=2, value=f"{reduction_rate:.1f}%")
                current_row += 1
            # FIX: Ende Zusätzliche Analyse
            
            # Spaltenbreiten anpassen
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            
            print("✅ Erweiterte Review-Statistiken erfolgreich exportiert")
            
        except Exception as e:
            print(f"⚠️ Fehler beim Export der Review-Statistiken: {str(e)}")
            import traceback
            traceback.print_exc()

    
    def _export_detailed_intercoder_analysis(self, writer, codings: List[Dict], reliability: float):
        """
        PUNKT 6: Detaillierte Intercoder-Analyse mit Subkategorien-Unstimmigkeiten
        """
        try:
            print("ðŸ‘¥ Erstelle detaillierte Intercoder-Analyse...")
            
            # Gruppiere Kodierungen nach ursprünglicher Segment-ID
            from collections import defaultdict
            
            original_segments = defaultdict(list)
            for coding in codings:
                # Extrahiere ursprüngliche Segment-ID
                consensus_info = coding.get('consensus_info', {})
                original_id = consensus_info.get('original_segment_id', coding.get('segment_id', ''))
                if original_id:
                    original_segments[original_id].append(coding)
            
            # Analysiere Unstimmigkeiten
            disagreement_data = []
            
            for original_id, segment_codings in original_segments.items():
                if len(segment_codings) < 2:
                    continue  # Keine Unstimmigkeit mÖglich bei nur einem Kodierer
                
                # Analysiere Hauptkategorien-Unstimmigkeiten
                main_categories = [c.get('category', '') for c in segment_codings]
                unique_main_cats = set(main_categories)
                
                # Analysiere Subkategorien-Unstimmigkeiten
                for main_cat in unique_main_cats:
                    if main_cat in ['Nicht kodiert', 'Kein Kodierkonsens', '']:
                        continue
                    
                    # Finde alle Kodierungen fuer diese Hauptkategorie
                    cat_codings = [c for c in segment_codings if c.get('category') == main_cat]
                    
                    if len(cat_codings) < 2:
                        continue
                    
                    # Subkategorien-Analyse
                    subcat_sets = []
                    coder_info = []
                    
                    for coding in cat_codings:
                        subcats = set(coding.get('subcategories', []))
                        subcat_sets.append(subcats)
                        coder_info.append({
                            'coder': coding.get('coder_id', 'Unbekannt'),
                            'subcats': list(subcats),
                            'confidence': self._extract_confidence_from_coding(coding)
                        })
                    
                    # Prüfe auf Subkategorien-Unstimmigkeiten
                    all_subcats_identical = all(s == subcat_sets[0] for s in subcat_sets)
                    
                    if not all_subcats_identical or len(unique_main_cats) > 1:
                        # Unstimmigkeit gefunden
                        disagreement_data.append({
                            'Segment_ID': original_id,
                            'Hauptkategorie': main_cat,
                            'Anzahl_Kodierer': len(cat_codings),
                            'Hauptkat_Konsens': 'Ja' if len(unique_main_cats) == 1 else 'Nein',
                            'Subkat_Konsens': 'Ja' if all_subcats_identical else 'Nein',
                            'Kodierer_1': coder_info[0]['coder'] if len(coder_info) > 0 else '',
                            'Subkats_1': ', '.join(coder_info[0]['subcats']) if len(coder_info) > 0 else '',
                            'Konfidenz_1': f"{coder_info[0]['confidence']:.2f}" if len(coder_info) > 0 else '',
                            'Kodierer_2': coder_info[1]['coder'] if len(coder_info) > 1 else '',
                            'Subkats_2': ', '.join(coder_info[1]['subcats']) if len(coder_info) > 1 else '',
                            'Konfidenz_2': f"{coder_info[1]['confidence']:.2f}" if len(coder_info) > 1 else '',
                            'Kodierer_3': coder_info[2]['coder'] if len(coder_info) > 2 else '',
                            'Subkats_3': ', '.join(coder_info[2]['subcats']) if len(coder_info) > 2 else '',
                            'Konfidenz_3': f"{coder_info[2]['confidence']:.2f}" if len(coder_info) > 2 else '',
                            'Unstimmigkeits_Typ': self._classify_disagreement_type(unique_main_cats, all_subcats_identical)
                        })
            
            # Exportiere Unstimmigkeits-Analyse
            if disagreement_data:
                df_disagreements = pd.DataFrame(disagreement_data)
                df_disagreements.to_excel(writer, sheet_name='Intercoder_Unstimmigkeiten', index=False)
                
                # Formatierung anwenden
                worksheet = writer.sheets['Intercoder_Unstimmigkeiten']
                self._format_intercoder_sheet(worksheet, df_disagreements)
                
                print(f"✅ {len(disagreement_data)} Intercoder-Unstimmigkeiten analysiert")
            else:
                # Leeres Sheet mit Info erstellen
                empty_data = [{'Info': 'Keine Intercoder-Unstimmigkeiten gefunden'}]
                df_empty = pd.DataFrame(empty_data)
                df_empty.to_excel(writer, sheet_name='Intercoder_Unstimmigkeiten', index=False)
                print("ℹ️ Keine Intercoder-Unstimmigkeiten gefunden")
            
            # Zusätzlich: Übersichts-Statistiken
            self._create_intercoder_summary(writer, original_segments, reliability)
            
        except Exception as e:
            print(f"⚠️ Fehler bei Intercoder-Analyse: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _classify_disagreement_type(self, unique_main_cats, all_subcats_identical):
        """
        Klassifiziert den Typ der Unstimmigkeit
        """
        if len(unique_main_cats) > 1:
            return 'Hauptkategorie-Konflikt'
        elif not all_subcats_identical:
            return 'Subkategorie-Konflikt'
        else:
            return 'Andere Unstimmigkeit'
    
    def _format_intercoder_sheet(self, worksheet, df):
        """
        PUNKT 7: Formatierung des Intercoder-Sheets
        """
        # Header formatieren
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Rahmen fuer alle Zellen
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Spaltenbreiten anpassen
        column_widths = {
            'A': 15,  # Segment_ID
            'B': 20,  # Hauptkategorie
            'C': 12,  # Anzahl_Kodierer
            'D': 15,  # Hauptkat_Konsens
            'E': 15,  # Subkat_Konsens
            'F': 12,  # Kodierer_1
            'G': 25,  # Subkats_1
            'H': 10,  # Konfidenz_1
            # ... weitere Spalten
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    
    def _export_intercoder_bericht(self, writer, original_codings: List[Dict], reliability: float):
        """
        FIX: Intercoder-Bericht mit bereits berechneter Reliabilität
        FÜr ResultsExporter Klasse
        """
        try:
            print("🧾 Erstelle IntercoderBericht mit ursprünglichen Daten...")
            
            worksheet = writer.book.create_sheet("IntercoderBericht")
            current_row = 1
            
            # Titel
            title_cell = worksheet.cell(row=current_row, column=1, value="Intercoder-Reliabilitäts-Bericht")
            title_cell.font = Font(bold=True, size=14)
            current_row += 2
            
            # FIX: Verwende bereits berechnete Reliabilität (aus main())
            print(f"🧾 Verwende bereits berechnete Reliabilität: {reliability:.3f}")
            
            # FIX: Verwende bereits berechneten comprehensive report statt Neuberechnung
            if hasattr(self, 'comprehensive_reliability_report') and self.comprehensive_reliability_report:
                print("🧾 Verwende bereits berechneten comprehensive reliability report")
                comprehensive_report = self.comprehensive_reliability_report
                statistics = comprehensive_report['statistics']
                agreement_analysis = comprehensive_report['agreement_analysis']
            else:
                print("🧾 Fallback: Berechne reliability report (comprehensive report nicht verfügbar)")
                # Fallback: Berechne nur wenn nicht bereits verfügbar
                reliability_calc = ReliabilityCalculator()
                comprehensive_report = reliability_calc.calculate_comprehensive_reliability(original_codings)
                statistics = reliability_calc._calculate_basic_statistics(original_codings)
                agreement_analysis = reliability_calc._calculate_detailed_agreement_analysis(original_codings)
            
            # 1. Reliabilitäts-Übersicht
            worksheet.cell(row=current_row, column=1, value="1. Reliabilitäts-Übersicht")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # FIX: Overall Alpha fett gedruckt und prominent
            worksheet.cell(row=current_row, column=1, value="Overall Alpha (Jaccard-basiert):")
            alpha_cell = worksheet.cell(row=current_row, column=2, value=f"{comprehensive_report['overall_alpha']:.3f}")
            alpha_cell.font = Font(bold=True, size=12)
            current_row += 1
            
            # FIX: Bewertung hinzuFügen
            worksheet.cell(row=current_row, column=1, value="Bewertung:")
            overall_alpha = comprehensive_report['overall_alpha']
            rating = "Exzellent" if overall_alpha > 0.8 else "Akzeptabel" if overall_alpha > 0.667 else "Unzureichend"
            rating_cell = worksheet.cell(row=current_row, column=2, value=rating)
            rating_cell.font = Font(bold=True)
            
            # FIX: Farbkodierung fuer die Bewertung
            if overall_alpha > 0.8:
                rating_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif overall_alpha > 0.667:
                rating_cell.fill = PatternFill(start_color='FFFF90', end_color='FFFF90', fill_type='solid')
            else:
                rating_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            current_row += 2
            
            # FIX: Zusätzliche Alpha-Werte aus dem comprehensive report
            worksheet.cell(row=current_row, column=1, value="Hauptkategorien Alpha (Jaccard):")
            worksheet.cell(row=current_row, column=2, value=f"{comprehensive_report['main_categories_alpha']:.3f}")
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Subkategorien Alpha (Jaccard):")
            worksheet.cell(row=current_row, column=2, value=f"{comprehensive_report['subcategories_alpha']:.3f}")
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Vergleichbare Segmente:")
            worksheet.cell(row=current_row, column=2, value=comprehensive_report['statistics']['vergleichbare_segmente'])
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Anzahl Kodierer:")
            worksheet.cell(row=current_row, column=2, value=comprehensive_report['statistics']['anzahl_kodierer'])
            current_row += 2
            
            # FIX: Methodik-Informationen hinzuFügen
            worksheet.cell(row=current_row, column=1, value="2. Methodik")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Alle Alpha-Werte verwenden:")
            worksheet.cell(row=current_row, column=2, value="Jaccard-Ähnlichkeit")
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Berechnung:")
            worksheet.cell(row=current_row, column=2, value="Set-basierte Berechnung")
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Konsistenz:")
            worksheet.cell(row=current_row, column=2, value="Overall zwischen Haupt- und Sub-Alpha")
            current_row += 2
            
            # 3. Detaillierte Set-Analyse für Nachprüfung
            worksheet.cell(row=current_row, column=1, value="3. Detaillierte Set-Analyse (Nachprüfung)")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            worksheet.cell(row=current_row, column=1, value="Diese Tabelle zeigt die Sets je Kodierer und Segment für die manuelle Nachprüfung der Jaccard-Berechnung.")
            current_row += 2
            
            # Erstelle detaillierte Set-Analyse
            current_row = self._add_detailed_set_analysis(worksheet, original_codings, current_row)
            
            # Spaltenbreiten anpassen
            worksheet.column_dimensions['A'].width = 25  # Segment ID
            worksheet.column_dimensions['B'].width = 15  # Kodierer 1
            worksheet.column_dimensions['C'].width = 15  # Kodierer 2
            worksheet.column_dimensions['D'].width = 40  # Set 1
            worksheet.column_dimensions['E'].width = 40  # Set 2
            worksheet.column_dimensions['F'].width = 30  # Schnittmenge
            worksheet.column_dimensions['G'].width = 30  # Vereinigung
            worksheet.column_dimensions['H'].width = 18  # Jaccard-Ähnlichkeit
            
            print("✅ IntercoderBericht mit ursprünglichen Daten und detaillierter Set-Analyse erstellt")
            
        except Exception as e:
            print(f"⚠️ Fehler beim IntercoderBericht: {str(e)}")
            import traceback
            traceback.print_exc()

    def _add_detailed_set_analysis(self, worksheet, original_codings: List[Dict], start_row: int) -> int:
        """
        Fügt detaillierte Set-Analyse für Nachprüfung der Jaccard-Berechnung hinzu.
        
        Args:
            worksheet: Excel worksheet
            original_codings: Liste der ursprünglichen Kodierungen
            start_row: Startzeile für die Tabelle
            
        Returns:
            Nächste verfügbare Zeile nach der Tabelle
        """
        try:
            from ..quality.reliability import ReliabilityCalculator
            
            # Initialisiere ReliabilityCalculator für Hilfsmethoden
            reliability_calc = ReliabilityCalculator()
            
            # Gruppiere Kodierungen nach Segmenten
            segment_data = reliability_calc._group_by_original_segments(original_codings)
            
            current_row = start_row
            
            # Header für die Set-Analyse Tabelle
            headers = [
                "Segment ID", "Kodierer 1", "Kodierer 2", 
                "Set 1 (Kategorien + Subkategorien)", "Set 2 (Kategorien + Subkategorien)",
                "Schnittmenge", "Vereinigung", "Jaccard-Ähnlichkeit"
            ]
            
            for col, header in enumerate(headers, 1):
                header_cell = worksheet.cell(row=current_row, column=col, value=header)
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            current_row += 1
            
            # FIX: Korrekte Behandlung von Mehrfachkodierungen
            # Sammle alle Vergleiche für die Tabelle (inklusive Mehrfachkodierungen)
            comparisons = []
            
            # NEUE LOGIK: Gruppiere nach tatsächlichen Segment-IDs (mit Mehrfachkodierungs-Suffixen)
            actual_segment_data = {}
            for coding in original_codings:
                actual_segment_id = coding.get('segment_id', '')
                coder_id = coding.get('coder_id', 'unknown')
                
                if actual_segment_id not in actual_segment_data:
                    actual_segment_data[actual_segment_id] = {}
                
                if coder_id not in actual_segment_data[actual_segment_id]:
                    actual_segment_data[actual_segment_id][coder_id] = []
                
                actual_segment_data[actual_segment_id][coder_id].append(coding)
            
            # Sammle alle Basis-Segment-IDs
            base_segments = {}
            for actual_segment_id in actual_segment_data.keys():
                base_id = reliability_calc._extract_base_segment_id({'segment_id': actual_segment_id})
                if base_id not in base_segments:
                    base_segments[base_id] = []
                base_segments[base_id].append(actual_segment_id)
            
            # Erstelle Vergleiche für jede Kombination von Mehrfachkodierungen
            for base_segment_id, actual_segment_ids in base_segments.items():
                # Sammle alle Kodierer für dieses Basis-Segment
                all_coders_for_base = set()
                for actual_id in actual_segment_ids:
                    all_coders_for_base.update(actual_segment_data[actual_id].keys())
                
                all_coders_list = list(all_coders_for_base)
                
                if len(all_coders_list) < 2:
                    continue
                
                # Paarweise Vergleiche zwischen allen Kodierern
                for i in range(len(all_coders_list)):
                    for j in range(i + 1, len(all_coders_list)):
                        coder1, coder2 = all_coders_list[i], all_coders_list[j]
                        
                        # Finde alle Kodierungen von beiden Kodierern für dieses Basis-Segment
                        coder1_codings = []
                        coder2_codings = []
                        
                        for actual_id in actual_segment_ids:
                            if coder1 in actual_segment_data[actual_id]:
                                coder1_codings.extend(actual_segment_data[actual_id][coder1])
                            if coder2 in actual_segment_data[actual_id]:
                                coder2_codings.extend(actual_segment_data[actual_id][coder2])
                        
                        if not coder1_codings or not coder2_codings:
                            continue
                        
                        # Erstelle Vergleiche für alle Kombinationen von Mehrfachkodierungen
                        for coding1 in coder1_codings:
                            for coding2 in coder2_codings:
                                # Sammle Sets für beide Kodierungen
                                set1 = set()
                                set2 = set()
                                
                                # Set 1 (Kodierung 1)
                                main_cat1 = coding1.get('category', '')
                                if main_cat1 and main_cat1 not in ['Nicht kodiert', 'Kein Kodierkonsens']:
                                    set1.add(f"MAIN:{main_cat1}")
                                
                                subcats1 = coding1.get('subcategories', [])
                                if isinstance(subcats1, (list, tuple)):
                                    for subcat in subcats1:
                                        if subcat:
                                            set1.add(f"SUB:{subcat}")
                                
                                # Set 2 (Kodierung 2)
                                main_cat2 = coding2.get('category', '')
                                if main_cat2 and main_cat2 not in ['Nicht kodiert', 'Kein Kodierkonsens']:
                                    set2.add(f"MAIN:{main_cat2}")
                                
                                subcats2 = coding2.get('subcategories', [])
                                if isinstance(subcats2, (list, tuple)):
                                    for subcat in subcats2:
                                        if subcat:
                                            set2.add(f"SUB:{subcat}")
                                
                                # Berechne Jaccard-Ähnlichkeit
                                if len(set1) == 0 and len(set2) == 0:
                                    jaccard_similarity = 1.0
                                    intersection = set()
                                    union = set()
                                elif len(set1) == 0 or len(set2) == 0:
                                    jaccard_similarity = 0.0
                                    intersection = set()
                                    union = set1.union(set2)
                                else:
                                    intersection = set1.intersection(set2)
                                    union = set1.union(set2)
                                    jaccard_similarity = len(intersection) / len(union) if len(union) > 0 else 0.0
                                
                                # Formatiere Sets für Anzeige
                                set1_display = ", ".join(sorted(set1)) if set1 else "(leer)"
                                set2_display = ", ".join(sorted(set2)) if set2 else "(leer)"
                                intersection_display = ", ".join(sorted(intersection)) if intersection else "(leer)"
                                union_display = ", ".join(sorted(union)) if union else "(leer)"
                                
                                # Verwende tatsächliche Segment-IDs für Anzeige
                                display_segment_id = f"{coding1.get('segment_id', '')} vs {coding2.get('segment_id', '')}"
                                
                                comparisons.append({
                                    'segment_id': display_segment_id,
                                    'coder1': coder1,
                                    'coder2': coder2,
                                    'set1': set1_display,
                                    'set2': set2_display,
                                    'intersection': intersection_display,
                                    'union': union_display,
                                    'jaccard': jaccard_similarity
                                })
            
            # Sortiere Vergleiche nach Segment ID
            comparisons.sort(key=lambda x: x['segment_id'])
            
            # Füge Vergleiche zur Tabelle hinzu
            for comparison in comparisons:
                worksheet.cell(row=current_row, column=1, value=comparison['segment_id'])
                worksheet.cell(row=current_row, column=2, value=comparison['coder1'])
                worksheet.cell(row=current_row, column=3, value=comparison['coder2'])
                worksheet.cell(row=current_row, column=4, value=comparison['set1'])
                worksheet.cell(row=current_row, column=5, value=comparison['set2'])
                worksheet.cell(row=current_row, column=6, value=comparison['intersection'])
                worksheet.cell(row=current_row, column=7, value=comparison['union'])
                
                # Jaccard-Ähnlichkeit mit Farbkodierung
                jaccard_cell = worksheet.cell(row=current_row, column=8, value=f"{comparison['jaccard']:.3f}")
                
                # Farbkodierung basierend auf Ähnlichkeit
                if comparison['jaccard'] >= 0.8:
                    jaccard_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Grün
                elif comparison['jaccard'] >= 0.5:
                    jaccard_cell.fill = PatternFill(start_color='FFFF90', end_color='FFFF90', fill_type='solid')  # Gelb
                else:
                    jaccard_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Rosa
                
                current_row += 1
            
            # Füge Zusammenfassung hinzu
            current_row += 1
            worksheet.cell(row=current_row, column=1, value="Zusammenfassung:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            if comparisons:
                avg_jaccard = sum(c['jaccard'] for c in comparisons) / len(comparisons)
                worksheet.cell(row=current_row, column=1, value=f"Durchschnittliche Jaccard-Ähnlichkeit:")
                worksheet.cell(row=current_row, column=2, value=f"{avg_jaccard:.3f}")
                current_row += 1
                
                worksheet.cell(row=current_row, column=1, value=f"Anzahl Vergleiche:")
                worksheet.cell(row=current_row, column=2, value=len(comparisons))
                current_row += 1
                
                high_agreement = sum(1 for c in comparisons if c['jaccard'] >= 0.8)
                worksheet.cell(row=current_row, column=1, value=f"Hohe Übereinstimmung (≥0.8):")
                worksheet.cell(row=current_row, column=2, value=f"{high_agreement}/{len(comparisons)} ({high_agreement/len(comparisons)*100:.1f}%)")
                current_row += 1
            else:
                worksheet.cell(row=current_row, column=1, value="Keine Vergleiche verfügbar")
                current_row += 1
            
            current_row += 1
            
            # Erklärung der Berechnung
            worksheet.cell(row=current_row, column=1, value="4. Erklärung der Jaccard-Berechnung")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            explanations = [
                "• Jaccard-Ähnlichkeit = |Schnittmenge| / |Vereinigungsmenge|",
                "• MAIN: kennzeichnet Hauptkategorien",
                "• SUB: kennzeichnet Subkategorien", 
                "• Beide Sets werden kombiniert (Haupt- + Subkategorien)",
                "• Farbkodierung: Grün (≥0.8), Gelb (≥0.5), Rosa (<0.5)",
                "• Leere Sets werden als perfekte Übereinstimmung (1.0) gewertet"
            ]
            
            for explanation in explanations:
                worksheet.cell(row=current_row, column=1, value=explanation)
                current_row += 1
            
            print(f"✅ Detaillierte Set-Analyse hinzugefügt: {len(comparisons)} Vergleiche")
            
            return current_row + 1
            
        except Exception as e:
            print(f"⚠️ Fehler bei detaillierter Set-Analyse: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Fallback: Fehlermeldung in Sheet
            worksheet.cell(row=start_row, column=1, value=f"Fehler bei Set-Analyse: {str(e)}")
            return start_row + 2

    def _create_empty_intercoder_sheet(self, writer):
        """
        FIX: Erstellt Info-Sheet wenn keine Reliabilitätsdaten verfügbar
        FÜr ResultsExporter Klasse
        """
        worksheet = writer.book.create_sheet("IntercoderBericht")
        
        worksheet.cell(row=1, column=1, value="Intercoder-Reliabilitäts-Bericht").font = Font(bold=True, size=14)
        worksheet.cell(row=3, column=1, value="❌ Keine ursprünglichen Kodierungen fuer Reliabilitätsberechnung verfügbar")
        worksheet.cell(row=4, column=1, value="Reliabilität muss vor dem Review-Prozess berechnet werden")
        
        worksheet.column_dimensions['A'].width = 60


    def _create_intercoder_summary(self, writer, original_segments, reliability):
        """
        FIX: Erstellt erweiterte Intercoder-Übersicht mit detaillierten Alpha-Informationen
        FÜr ResultsExporter Klasse
        """
        try:
            print("🧾 Erstelle erweiterte Intercoder-Übersicht...")
            
            # FIX: Verwende bereits berechneten comprehensive report statt Neuberechnung
            if hasattr(self, 'comprehensive_reliability_report') and self.comprehensive_reliability_report:
                print("🧾 Verwende bereits berechneten comprehensive reliability report")
                comprehensive_report = self.comprehensive_reliability_report
                overall_alpha = comprehensive_report['overall_alpha']
                main_categories_alpha = comprehensive_report['main_categories_alpha']
                subcategories_alpha = comprehensive_report['subcategories_alpha']
                agreement_analysis = comprehensive_report['agreement_analysis']
                statistics = comprehensive_report['statistics']
            else:
                print("🧾 Fallback: Berechne reliability report (comprehensive report nicht verfügbar)")
                # Extrahiere ursprüngliche Kodierungen aus den Segment-Daten
                original_codings = []
                for segment_codings in original_segments.values():
                    original_codings.extend(segment_codings)
                
                # FIX: Berechne umfassende Reliabilität mit Details nur als Fallback
                if original_codings:
                    reliability_calc = ReliabilityCalculator()
                    comprehensive_report = reliability_calc.calculate_comprehensive_reliability(original_codings)
                    overall_alpha = comprehensive_report['overall_alpha']
                    main_categories_alpha = comprehensive_report['main_categories_alpha']
                    subcategories_alpha = comprehensive_report['subcategories_alpha']
                    agreement_analysis = comprehensive_report['agreement_analysis']
                    statistics = comprehensive_report['statistics']
                else:
                    # Fallback wenn keine Daten
                    overall_alpha = reliability
                    main_categories_alpha = 0.0
                    subcategories_alpha = 0.0
                    agreement_analysis = {'Vollständige Übereinstimmung': 0, 'Hauptkategorie gleich, Subkat. unterschiedlich': 0, 'Hauptkategorie unterschiedlich': 0, 'Gesamt': 0}
                    statistics = {'vergleichbare_segmente': 0, 'anzahl_kodierer': 0}
            
            # Berechne Übersichtsstatistiken
            total_segments = len(original_segments)
            segments_with_multiple_coders = sum(1 for segs in original_segments.values() if len(segs) > 1)
            segments_single_coder = total_segments - segments_with_multiple_coders
            
            # Kodierer-Statistiken
            all_coders = set()
            for segment_codings in original_segments.values():
                for coding in segment_codings:
                    coder_id = coding.get('coder_id', 'Unbekannt')
                    all_coders.add(coder_id)
            
            # FIX: Erweiterte Daten mit allen Alpha-Details
            summary_data = [
                ['Metrik', 'Wert'],
                ['Gesamte Segmente', total_segments],
                ['Segmente mit mehreren Kodierern', segments_with_multiple_coders],
                ['Segmente mit einem Kodierer', segments_single_coder],
                ['Anteil Mehrfachkodierung', f"{(segments_with_multiple_coders/total_segments)*100:.1f}%" if total_segments > 0 else "0%"],
                ['Anzahl Kodierer', len(all_coders)],
                ['Kodierer', ', '.join(sorted(all_coders))],
                ['', ''],  # Leerzeile als Trenner
                ['--- KRIPPENDORFF\'S ALPHA DETAILS ---', ''],
                ['Overall Alpha (Jaccard-basiert)', f"{overall_alpha:.3f}"],
                ['Hauptkategorien Alpha (Jaccard)', f"{main_categories_alpha:.3f}"],
                ['Subkategorien Alpha (Jaccard)', f"{subcategories_alpha:.3f}"],
                ['Vergleichbare Segmente', statistics.get('vergleichbare_segmente', 0)],
                ['', ''],  # Leerzeile als Trenner
                ['--- ÜBEREINSTIMMUNGSANALYSE ---', ''],
                ['Vollständige Übereinstimmung', agreement_analysis.get('Vollständige Übereinstimmung', 0)],
                ['Hauptkategorie gleich, Subkat. unterschiedlich', agreement_analysis.get('Hauptkategorie gleich, Subkat. unterschiedlich', 0)],
                ['Hauptkategorie unterschiedlich', agreement_analysis.get('Hauptkategorie unterschiedlich', 0)],
                ['Gesamt analysiert', agreement_analysis.get('Gesamt', 0)],
                ['', ''],  # Leerzeile als Trenner
                ['--- BEWERTUNG ---', ''],
                ['Overall Alpha Bewertung', self._get_reliability_rating(overall_alpha)],
                ['Hauptkategorien Bewertung', self._get_reliability_rating(main_categories_alpha)],
                ['Subkategorien Bewertung', self._get_reliability_rating(subcategories_alpha)]
            ]
            
            df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            df_summary.to_excel(writer, sheet_name='Intercoder_Übersicht', index=False)
            
            # Formatierung
            worksheet = writer.sheets['Intercoder_Übersicht']
            self._apply_professional_formatting(worksheet, df_summary)
            
            # FIX: Spezielle Formatierung fuer Trennzeilen
            for row_idx in range(1, len(summary_data)):
                cell_value = summary_data[row_idx][0]
                if cell_value.startswith('---') and cell_value.endswith('---'):
                    # Fette Formatierung fuer Sektions-Header
                    worksheet.cell(row=row_idx + 1, column=1).font = Font(bold=True, size=11)
                elif cell_value == '':
                    # Leerzeilen
                    worksheet.cell(row=row_idx + 1, column=1).value = ''
            
            # Spaltenbreiten
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 25
            
            print("✅ Erweiterte Intercoder-Übersicht mit Alpha-Details erstellt")
            
        except Exception as e:
            print(f"⚠️ Fehler bei erweiterter Intercoder-Übersicht: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _apply_professional_formatting(self, worksheet, df):
        """
        Wendet professionelle Formatierung an
        """
        # Header formatieren
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Rahmen fuer alle Zellen
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
                # PUNKT 1: Keine Text-Wrapping
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        
        # Abwechselnde ZeilenfÄrbung
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col)
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    def _apply_crosstable_formatting(self, worksheet, crosstab):
        """
        PUNKT 7: Spezielle Formatierung fuer Kreuztabellen
        """
        # Header-Zeile und -Spalte formatieren
        for cell in worksheet[1]:  # Erste Zeile
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
        
        for row in range(1, len(crosstab) + 2):  # Erste Spalte
            cell = worksheet.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
        
        # Rahmen
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(crosstab)+1, min_col=1, max_col=len(crosstab.columns)+1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Marginal-Zeilen/-Spalten hervorheben (falls vorhanden)
        if 'All' in crosstab.index:
            last_row = len(crosstab) + 1
            for col in range(1, len(crosstab.columns) + 2):
                cell = worksheet.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        
        if 'All' in crosstab.columns:
            last_col = len(crosstab.columns) + 1
            for row in range(1, len(crosstab) + 2):
                cell = worksheet.cell(row=row, column=last_col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
    
    def _extract_confidence_from_coding(self, coding: Dict) -> float:
        """
        Extrahiert Konfidenzwert aus Kodierung
        """
        confidence = coding.get('confidence', {})
        if isinstance(confidence, dict):
            return confidence.get('total', 0.0)
        elif isinstance(confidence, (int, float)):
            return float(confidence)
        return 0.0
    
    def _determine_category_type(self, category: str, original_categories: dict) -> str:
        """
        # FIX: Erweiterte Bestimmung des Kategorietyps (deduktiv/induktiv/grounded)
        Bestimmt Kategorietyp (deduktiv/induktiv/grounded) basierend auf Analysemodus und Kategoriendefinition
        
        Args:
            category: Name der Kategorie
            original_categories: Dictionary der ursprünglichen Kategorien mit CategoryDefinition-Objekten
            
        Returns:
            str: 'Deduktiv', 'Induktiv', 'Grounded' oder '' (fuer nicht kodiert)
        """
        if not category or category in ['Nicht kodiert', 'Kein Kodierkonsens']:
            return ''
        
        # FIX: Prüfe zuerst ob es eine deduktive Kategorie ist
        if category in original_categories:
            return 'Deduktiv'
        
        # FIX: Prüfe ob die Analyse im grounded mode durchgefÜhrt wurde
        analysis_mode = CONFIG.get('ANALYSIS_MODE', 'deductive')
        if analysis_mode == 'grounded':
            return 'Grounded'
        
        # FIX: Alternative: Prüfe ob in der CategoryDefinition ein development_type gespeichert ist
        if original_categories and category in original_categories:
            category_obj = original_categories[category]
            if hasattr(category_obj, 'development_type'):
                dev_type = getattr(category_obj, 'development_type', '')
                if dev_type == 'grounded':
                    return 'Grounded'
                elif dev_type == 'deductive':
                    return 'Deduktiv'
                elif dev_type == 'abductive':
                    return 'Abduktiv'
                elif dev_type == 'inductive':
                    return 'Induktiv'
        
        # FIX: Fallback: Alle anderen Kategorien sind induktiv
        return 'Induktiv'
    
    def _extract_attributes_from_document(self, doc_name: str) -> tuple:
        """
        Extrahiert alle 3 Attribute aus Dokumentname
        """
        return self._extract_three_attributes_from_document(doc_name)
    
    def _prepare_dataframe_for_frequency_analysis(self, codings: List[Dict]) -> pd.DataFrame:
        """
        Bereitet DataFrame fuer HÄufigkeitsanalyse vor
        """
        try:
            data = []
            
            for coding in codings:
                doc_name = coding.get('document', '')
                if not doc_name:
                    doc_name = self._extract_document_from_segment_id(coding.get('segment_id', ''))
                
                attribut1, attribut2, attribut3 = self._extract_three_attributes_from_document(doc_name)
                
                row_data = {
                    'Dokument': doc_name,
                    self.attribute_labels.get('attribut1', 'Attribut1'): attribut1,
                    self.attribute_labels.get('attribut2', 'Attribut2'): attribut2,
                    self.attribute_labels.get('attribut3', 'Attribut3'): attribut3,
                    'Chunk_Nr': coding.get('chunk_id', coding.get('segment_id', '')),
                    'Hauptkategorie': coding.get('category', ''),
                    'Kodiert': 'Ja' if coding.get('category') and coding.get('category') not in ['Nicht kodiert', 'Kein Kodierkonsens'] else 'Nein',
                    'Subkategorien': ', '.join(coding.get('subcategories', [])),
                    'Konfidenz': self._extract_confidence_from_coding(coding)
                }
                
                data.append(row_data)
            
            return pd.DataFrame(data) if data else pd.DataFrame()
            
        except Exception as e:
            print(f"⚠️ Fehler beim DataFrame erstellen: {str(e)}")
            return pd.DataFrame()
    
    def export_annotated_pdfs(self, 
                             codings: List[Dict], 
                             chunks: Dict[str, List[str]], 
                             data_dir: str) -> List[str]:
        """
        FIX: Neue Methode fuer ResultsExporter Klasse
        Exportiert annotierte PDFs fuer alle gefundenen PDF-Eingabedateien
        
        Args:
            codings: Liste der finalen Kodierungen
            chunks: Dictionary mit chunk_id -> text mapping
            data_dir: Input-Verzeichnis mit Original-PDF-Dateien (aus CONFIG['DATA_DIR'])
            
        Returns:
            List[str]: Liste der Pfade zu erstellten annotierten PDFs
        """
        print(f"\n💾 Beginne PDF-Annotations-Export...")
        
        try:
            # FIX: Importiere PDF-Annotator (nur wenn benÖtigt)
            from ..utils.export.pdf_annotator import PDFAnnotator
        except ImportError:
            print("   ⚠️ PyMuPDF nicht verfügbar - PDF-Annotation Übersprungen")
            print("   ℹ️ Installieren Sie mit: pip install PyMuPDF")
            return []
        
        # FIX: Initialisiere PDF-Annotator
        pdf_annotator = PDFAnnotator(self)
        
        # FIX: Nutze os.path (wie im Rest von QCA-AID) statt Path
        pdf_files = []
        
        # FIX: Finde PDF-Dateien mit der gleichen Logik wie DocumentReader
        try:
            if not os.path.exists(data_dir):
                print(f"   ❌ Verzeichnis {data_dir} existiert nicht")
                return []

            for file in os.listdir(data_dir):
                file_path = os.path.join(data_dir, file)
                file_ext = os.path.splitext(file)[1].lower()
                
                if os.path.isfile(file_path) and file_ext == '.pdf':
                    pdf_files.append((file, file_path))
                    
        except Exception as e:
            print(f"   ⚠️ Fehler beim Durchsuchen des Verzeichnisses: {e}")
            return []
        
        if not pdf_files:
            print("   ℹ️ Keine PDF-Dateien im Input-Verzeichnis gefunden")
            return []
        
        print(f"   ℹ️ {len(pdf_files)} PDF-Dateien gefunden")
        annotated_files = []
        
        # FIX: Annotiere jede PDF-Datei
        for filename, file_path in pdf_files:
            print(f"\n   ℹ️ Verarbeite: {filename}")
            
            # FIX: Filtere nur konsolidierte/Review-Kodierungen fuer diese Datei
            file_stem = os.path.splitext(filename)[0]
            file_codings = []
            
            for coding in codings:
                # FIX: Nur Kodierungen nach Review/Consensus nehmen
                is_review_coding = (
                    coding.get('consensus_info') is not None or          # Hat Consensus-Info
                    coding.get('review_decision') is not None or         # Hat Review-Entscheidung  
                    coding.get('selection_type') in ['consensus', 'majority', 'manual_priority'] or  # Ist Review-Ergebnis
                    len([c for c in codings if c.get('segment_id') == coding.get('segment_id')]) == 1  # Einzige Kodierung fuer Segment
                )
                
                # FIX: Prüfe ob Kodierung zu dieser Datei gehÖrt
                matches_file = (
                    file_stem in coding.get('document', '') or 
                    file_stem in coding.get('segment_id', '')
                )
                
                if is_review_coding and matches_file:
                    file_codings.append(coding)
                    
            print(f"      🔀 {len(file_codings)} Review-Kodierungen fuer {filename} gefunden")
            
            if not file_codings:
                print(f"      ❌ Keine Kodierungen fuer {filename} gefunden")
                continue
            
            # FIX: Erstelle Ausgabepfad mit os.path - in "Annotated" Unterordner
            annotated_dir = os.path.join(self.output_dir, "Annotated")
            os.makedirs(annotated_dir, exist_ok=True)
            
            output_filename = f"{file_stem}_QCA_annotiert.pdf"
            output_file = os.path.join(annotated_dir, output_filename)
            
            # FIX: Annotiere PDF
            try:
                result_path = pdf_annotator.annotate_pdf_with_codings(
                    file_path,
                    file_codings,
                    chunks,
                    output_file
                )
                
                if result_path:
                    annotated_files.append(result_path)
                    print(f"      ✅ Erstellt: {os.path.basename(result_path)}")
                
            except Exception as e:
                print(f"      ⚠️ Fehler bei {filename}: {e}")
                continue
        
        print(f"\n✅ PDF-Annotation abgeschlossen: {len(annotated_files)} Dateien erstellt")
        return annotated_files
    
    def export_annotated_pdfs_all_formats(self, 
                                         codings: List[Dict], 
                                         chunks: Dict[str, List[str]], 
                                         data_dir: str) -> List[str]:
        """
        FIX: Neue Methode fuer ResultsExporter Klasse - Erweiterte PDF-Annotation fuer alle Formate
        
        Args:
            codings: Liste der finalen Kodierungen
            chunks: Dictionary mit chunk_id -> text mapping
            data_dir: Input-Verzeichnis mit Original-Dateien
            
        Returns:
            List[str]: Liste der Pfade zu erstellten annotierten PDFs
        """
        print(f"\n💾 Beginne erweiterte PDF-Annotations-Export fuer alle Formate...")
        
        try:
            # FIX: Importiere benÖtigte Klassen
            from ..utils.export.pdf_annotator import PDFAnnotator
            from ..utils.export.converters import DocumentToPDFConverter
        except ImportError:
            print("   ⚠️ BenÖtigte Bibliotheken nicht verfügbar")
            print("   ℹ️ Installieren Sie mit: pip install PyMuPDF reportlab")
            return []
        
        # FIX: Initialisiere Konverter und Annotator
        try:
            pdf_converter = DocumentToPDFConverter(self.output_dir)
            pdf_annotator = PDFAnnotator(self)
        except Exception as e:
            print(f"   ⚠️ Fehler beim Initialisieren: {e}")
            return []
        
        # FIX: Finde alle unterstÜtzten Dateien
        supported_extensions = ['.pdf', '.txt', '.docx', '.doc']
        input_files = []
        
        try:
            if not os.path.exists(data_dir):
                print(f"   ❌ Verzeichnis {data_dir} existiert nicht")
                return []

            for file in os.listdir(data_dir):
                file_path = os.path.join(data_dir, file)
                file_ext = os.path.splitext(file)[1].lower()
                
                if os.path.isfile(file_path) and file_ext in supported_extensions:
                    input_files.append((file, file_path, file_ext))
                    
        except Exception as e:
            print(f"   ⚠️ Fehler beim Durchsuchen des Verzeichnisses: {e}")
            return []
        
        if not input_files:
            print("   ℹ️ Keine unterstÜtzten Dateien im Input-Verzeichnis gefunden")
            return []
        
        print(f"   🔀 {len(input_files)} Dateien gefunden:")
        for filename, _, ext in input_files:
            print(f"      - {filename} ({ext})")
        
        annotated_files = []
        
        # FIX: Verarbeite jede Datei
        for filename, file_path, file_ext in input_files:
            print(f"\n   ℹ️ Verarbeite: {filename}")
            
            # FIX: Filtere Review-Kodierungen fuer diese Datei
            file_stem = os.path.splitext(filename)[0]
            file_codings = []
            
            for coding in codings:
                # FIX: Nur Kodierungen nach Review/Consensus nehmen
                is_review_coding = (
                    coding.get('consensus_info') is not None or          
                    coding.get('review_decision') is not None or         
                    coding.get('selection_type') in ['consensus', 'majority', 'manual_priority'] or  
                    len([c for c in codings if c.get('segment_id') == coding.get('segment_id')]) == 1  
                )
                
                # FIX: Prüfe ob Kodierung zu dieser Datei gehÖrt
                matches_file = (
                    file_stem in coding.get('document', '') or 
                    file_stem in coding.get('segment_id', '')
                )
                
                if is_review_coding and matches_file:
                    file_codings.append(coding)
            
            if not file_codings:
                print(f"      ❌ Keine Review-Kodierungen fuer {filename} gefunden")
                continue
            
            print(f"      🔀 {len(file_codings)} Review-Kodierungen gefunden")
            
            # FIX: Konvertiere zu PDF falls nÖtig
            if file_ext == '.pdf':
                pdf_path = file_path
                print(f"      ✅ Bereits PDF")
            else:
                print(f"      ℹ️ Konvertiere {file_ext.upper()} zu PDF...")
                pdf_path = pdf_converter.convert(file_path)
                
                if not pdf_path:
                    print(f"      ⚠️ Konvertierung fehlgeschlagen")
                    continue
                
                print(f"      ✅ PDF erstellt: {os.path.basename(pdf_path)}")
            
            # FIX: Annotiere PDF
            try:
                output_filename = f"{file_stem}_QCA_annotiert.pdf"
                output_file = os.path.join(self.output_dir, output_filename)
                
                result_path = pdf_annotator.annotate_pdf_with_codings(
                    pdf_path,
                    file_codings,
                    chunks,
                    output_file
                )
                
                if result_path:
                    annotated_files.append(result_path)
                    print(f"      ✅ Annotiert: {os.path.basename(result_path)}")
                else:
                    print(f"      ⚠️ Annotation fehlgeschlagen")
                
            except Exception as e:
                print(f"      ⚠️ Fehler bei Annotation: {e}")
                continue
        
        # FIX: Bereinige temporÄre Dateien
        try:
            pdf_converter.cleanup_temp_pdfs()
        except Exception as e:
            print(f"   ❌ Fehler bei Bereinigung: {e}")
        
        print(f"\n✅ Erweiterte PDF-Annotation abgeschlossen: {len(annotated_files)} Dateien erstellt")
        return annotated_files
    
    def _export_configuration(self, writer, export_mode: str):
        """
        # FIX: Fehlende Methode _export_configuration hinzugefÜgt
        Exportiert die Konfiguration in ein separates Excel-Sheet.
        FÜr ResultsExporter Klasse
        
        Args:
            writer: Excel Writer Objekt
            export_mode: Verwendeter Export-Modus
        """
        try:
            from openpyxl.styles import Font, PatternFill
            
            if 'Konfiguration' not in writer.sheets:
                writer.book.create_sheet('Konfiguration')
                
            worksheet = writer.sheets['Konfiguration']
            current_row = 1
            
            # Titel
            worksheet.cell(row=current_row, column=1, value="QCA-AID Konfiguration")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            current_row += 2
            
            # Header fuer die Konfigurationstabelle
            worksheet.cell(row=current_row, column=1, value="Parameter")
            worksheet.cell(row=current_row, column=2, value="Wert")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            worksheet.cell(row=current_row, column=2).font = Font(bold=True)
            worksheet.cell(row=current_row, column=1).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            worksheet.cell(row=current_row, column=2).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            current_row += 1
            
            # Import version information
            try:
                from ..__version__ import __version__, __version_date__
            except ImportError:
                __version__ = "0.12.3"
                __version_date__ = "2025-12-01"
            
            # FIX: Alle wichtigen Konfigurationsparameter exportieren
            config_params = [
                ('QCA-AID_VERSION', __version__),
                ('VERSION_DATE', __version_date__),
                ('MODEL_PROVIDER', CONFIG.get('MODEL_PROVIDER', 'OpenAI')),
                ('MODEL_NAME', CONFIG.get('MODEL_NAME', 'gpt-4o-mini')),
                ('CHUNK_SIZE', CONFIG.get('CHUNK_SIZE', 2000)),
                ('CHUNK_OVERLAP', CONFIG.get('CHUNK_OVERLAP', 200)),
                ('BATCH_SIZE', CONFIG.get('BATCH_SIZE', 5)),
                ('CODE_WITH_CONTEXT', CONFIG.get('CODE_WITH_CONTEXT', False)),
                ('ANALYSIS_MODE', CONFIG.get('ANALYSIS_MODE', 'deductive')),
                ('REVIEW_MODE', export_mode),
                ('EXPORT_ANNOTATED_PDFS', CONFIG.get('EXPORT_ANNOTATED_PDFS', True)),
                ('ATTRIBUT1_LABEL', CONFIG.get('ATTRIBUTE_LABELS', {}).get('attribut1', 'Attribut1')),
                ('ATTRIBUT2_LABEL', CONFIG.get('ATTRIBUTE_LABELS', {}).get('attribut2', 'Attribut2')),
                ('TIMESTAMP', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            
            for param_name, param_value in config_params:
                worksheet.cell(row=current_row, column=1, value=param_name)
                worksheet.cell(row=current_row, column=2, value=str(param_value))
                current_row += 1
            
            # Spaltenbreiten anpassen
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 40
            
            print("✅ Konfiguration erfolgreich exportiert")
            
        except Exception as e:
            print(f"⚠️ Fehler beim Export der Konfiguration: {str(e)}")
            import traceback
            traceback.print_exc()
                