"""
Configuration Converter

Converts between Excel (XLSX) and JSON configuration formats for QCA-AID Explorer.
Handles bidirectional conversion while preserving data structure.
"""

import json
import os
from typing import Dict, Any
import pandas as pd
from openpyxl import Workbook, load_workbook


class ConfigConverter:
    """Konvertiert zwischen XLSX und JSON Konfigurationsformaten"""
    
    @staticmethod
    def xlsx_to_json(xlsx_path: str) -> Dict[str, Any]:
        """
        Konvertiert Excel-Konfigurationsdatei zu JSON-Struktur
        
        Args:
            xlsx_path: Pfad zur Excel-Datei
            
        Returns:
            Dictionary mit 'base_config' und 'analysis_configs'
            
        Raises:
            FileNotFoundError: Wenn Excel-Datei nicht existiert
            ValueError: Wenn Excel-Struktur ungültig ist
        """
        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {xlsx_path}")
        
        try:
            # Lese Basis-Sheet
            # keep_default_na=False verhindert dass Strings wie "null", "NA" als NaN interpretiert werden
            base_df = pd.read_excel(xlsx_path, sheet_name='Basis', keep_default_na=False)
            base_config = {}
            
            # Definiere erwartete Typen für base_config Parameter
            base_config_types = {
                'clean_keywords': bool,
                'temperature': float,
                'similarity_threshold': float
            }
            
            for _, row in base_df.iterrows():
                param_name = str(row['Parameter'])
                param_value = row['Wert']
                
                # Mit keep_default_na=False sind leere Zellen bereits leere Strings
                # Aber wir müssen noch NaN-Werte behandeln die durch andere Wege entstehen können
                if pd.isna(param_value):
                    param_value = ''
                
                # Konvertiere zu erwartetem Typ falls bekannt
                if param_name in base_config_types and param_value != '':
                    expected_type = base_config_types[param_name]
                    if expected_type == bool:
                        # Konvertiere 0/1 zurück zu bool
                        if isinstance(param_value, (int, float)):
                            param_value = bool(param_value)
                    elif expected_type == float:
                        # Stelle sicher dass Floats als Float bleiben
                        if isinstance(param_value, (int, float)):
                            param_value = float(param_value)
                    
                base_config[param_name] = param_value
            
            # Lese alle anderen Sheets für Analyse-Konfigurationen
            excel = pd.ExcelFile(xlsx_path)
            sheet_names = excel.sheet_names
            analysis_configs = []
            
            for sheet_name in sheet_names:
                if sheet_name.lower() != 'basis':
                    # keep_default_na=False verhindert dass Strings wie "null", "NA" als NaN interpretiert werden
                    analysis_df = pd.read_excel(xlsx_path, sheet_name=sheet_name, keep_default_na=False)
                    
                    analysis_config = {'name': sheet_name}
                    filter_params = {}
                    other_params = {}
                    
                    for _, row in analysis_df.iterrows():
                        param_name = str(row['Parameter'])
                        param_value = row['Wert']
                        
                        # Spezielle Behandlung für active/enabled Parameter
                        if param_name.lower() == 'active' or param_name.lower() == 'enabled':
                            if pd.isna(param_value):
                                param_value = True
                            elif isinstance(param_value, str):
                                param_value = param_value.lower() in ('true', 'ja', 'yes', '1')
                            else:
                                param_value = bool(param_value)
                            other_params['active'] = param_value
                            continue
                        
                        # Nur NaN/NaT als None behandeln, nicht leere Strings
                        if pd.isna(param_value):
                            param_value = None
                        elif isinstance(param_value, str) and param_value == '':
                            # Leere Strings bleiben leere Strings
                            param_value = ''
                        
                        # Filter-Parameter vs. andere Parameter
                        if param_name.startswith('filter_'):
                            filter_name = param_name[7:]
                            filter_params[filter_name] = param_value
                        else:
                            other_params[param_name] = param_value
                    
                    # Stelle sicher, dass 'active' existiert
                    if 'active' not in other_params:
                        other_params['active'] = True
                    
                    analysis_config['filters'] = filter_params
                    analysis_config['params'] = other_params
                    analysis_configs.append(analysis_config)
            
            return {
                'base_config': base_config,
                'analysis_configs': analysis_configs
            }
            
        except Exception as e:
            raise ValueError(f"Fehler beim Lesen der Excel-Datei: {str(e)}")
    
    @staticmethod
    def json_to_xlsx(json_data: Dict[str, Any], xlsx_path: str) -> None:
        """
        Schreibt JSON-Struktur in Excel-Datei
        
        Args:
            json_data: Dictionary mit 'base_config' und 'analysis_configs'
            xlsx_path: Pfad zur Ziel-Excel-Datei
            
        Raises:
            ValueError: Wenn JSON-Struktur ungültig ist
        """
        if 'base_config' not in json_data or 'analysis_configs' not in json_data:
            raise ValueError("JSON muss 'base_config' und 'analysis_configs' enthalten")
        
        try:
            wb = Workbook()
            
            # Erstelle Basis-Sheet
            ws_basis = wb.active
            ws_basis.title = 'Basis'
            ws_basis.append(['Parameter', 'Wert'])
            
            for param_name, param_value in json_data['base_config'].items():
                # Konvertiere None zu leerem String für Excel
                if param_value is None:
                    param_value = ''
                ws_basis.append([param_name, param_value])
            
            # Erstelle Sheets für Analyse-Konfigurationen
            for analysis_config in json_data['analysis_configs']:
                sheet_name = analysis_config['name']
                ws = wb.create_sheet(title=sheet_name)
                ws.append(['Parameter', 'Wert'])
                
                # Schreibe Filter-Parameter
                filters = analysis_config.get('filters', {})
                for filter_name, filter_value in filters.items():
                    # Konvertiere None zu leerem String für Excel
                    if filter_value is None:
                        filter_value = ''
                    ws.append([f'filter_{filter_name}', filter_value])
                
                # Schreibe andere Parameter
                params = analysis_config.get('params', {})
                for param_name, param_value in params.items():
                    # Konvertiere None zu leerem String für Excel
                    if param_value is None:
                        param_value = ''
                    ws.append([param_name, param_value])
            
            # Speichere Workbook
            wb.save(xlsx_path)
            wb.close()
            
        except Exception as e:
            raise ValueError(f"Fehler beim Schreiben der Excel-Datei: {str(e)}")
    
    @staticmethod
    def save_json(json_data: Dict[str, Any], json_path: str) -> None:
        """
        Speichert JSON-Daten mit UTF-8 Encoding und Formatierung
        
        Args:
            json_data: Dictionary zum Speichern
            json_path: Pfad zur Ziel-JSON-Datei
            
        Raises:
            IOError: Wenn Datei nicht geschrieben werden kann
        """
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            raise IOError(f"Fehler beim Schreiben der JSON-Datei: {str(e)}")
    
    @staticmethod
    def load_json(json_path: str) -> Dict[str, Any]:
        """
        Lädt JSON-Datei mit Fehlerbehandlung
        
        Args:
            json_path: Pfad zur JSON-Datei
            
        Returns:
            Dictionary mit JSON-Daten
            
        Raises:
            FileNotFoundError: Wenn JSON-Datei nicht existiert
            ValueError: Wenn JSON ungültig ist
        """
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON-Datei nicht gefunden: {json_path}")
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            raise ValueError(f"Ungültiges JSON-Format: {str(e)}")
        except Exception as e:
            raise IOError(f"Fehler beim Lesen der JSON-Datei: {str(e)}")
    
    @staticmethod
    def qca_aid_xlsx_to_json(xlsx_path: str) -> Dict[str, Any]:
        """
        Konvertiert QCA-AID Excel-Codebook zu JSON-Struktur
        
        Args:
            xlsx_path: Pfad zur Excel-Datei
            
        Returns:
            Dictionary mit 'forschungsfrage', 'kodierregeln', 'deduktive_kategorien', 'config'
            
        Raises:
            FileNotFoundError: Wenn Excel-Datei nicht existiert
            ValueError: Wenn Excel-Struktur ungültig ist
        """
        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {xlsx_path}")
        
        try:
            result = {}
            
            # 1. Lese FORSCHUNGSFRAGE Sheet - Wert aus Zelle B1
            forschungsfrage_df = pd.read_excel(xlsx_path, sheet_name='FORSCHUNGSFRAGE', header=None)
            result['forschungsfrage'] = str(forschungsfrage_df.iloc[0, 1]) if len(forschungsfrage_df) > 0 else ""
            
            # 2. Lese KODIERREGELN Sheet - gruppiere nach general/format/exclusion
            kodierregeln_df = pd.read_excel(xlsx_path, sheet_name='KODIERREGELN', keep_default_na=False)
            kodierregeln = {
                'general': [],
                'format': [],
                'exclusion': []
            }
            
            # Die Spalten sind: "Allgemeine Kodierregeln", "Formatregeln", "Ausschlussregeln"
            for col_name, key in [
                ('Allgemeine Kodierregeln', 'general'),
                ('Formatregeln', 'format'),
                ('Ausschlussregeln', 'exclusion')
            ]:
                if col_name in kodierregeln_df.columns:
                    for value in kodierregeln_df[col_name]:
                        if pd.notna(value) and str(value).strip() != '':
                            kodierregeln[key].append(str(value))
            
            result['kodierregeln'] = kodierregeln
            
            # 3. Lese DEDUKTIVE_KATEGORIEN Sheet - Key-Sub-Key-Sub-Sub-Key Hierarchie
            kategorien_df = pd.read_excel(xlsx_path, sheet_name='DEDUKTIVE_KATEGORIEN', keep_default_na=False)
            deduktive_kategorien = {}
            
            for _, row in kategorien_df.iterrows():
                key = str(row['Key'])
                sub_key = row['Sub-Key']
                sub_sub_key = row['Sub-Sub-Key']
                value = row['Value']
                
                # Initialisiere Kategorie falls noch nicht vorhanden
                if key not in deduktive_kategorien:
                    deduktive_kategorien[key] = {
                        'definition': '',
                        'rules': [],
                        'examples': [],
                        'subcategories': {}
                    }
                
                # Verarbeite basierend auf Sub-Key
                if pd.notna(sub_key):
                    sub_key_str = str(sub_key)
                    
                    if sub_key_str == 'definition':
                        deduktive_kategorien[key]['definition'] = str(value) if pd.notna(value) else ''
                    
                    elif sub_key_str == 'rules':
                        if pd.notna(value) and str(value).strip() != '':
                            deduktive_kategorien[key]['rules'].append(str(value))
                    
                    elif sub_key_str == 'examples':
                        if pd.notna(value) and str(value).strip() != '':
                            deduktive_kategorien[key]['examples'].append(str(value))
                    
                    elif sub_key_str == 'subcategories':
                        if pd.notna(sub_sub_key) and pd.notna(value):
                            deduktive_kategorien[key]['subcategories'][str(sub_sub_key)] = str(value)
            
            result['deduktive_kategorien'] = deduktive_kategorien
            
            # 4. Lese CONFIG Sheet - verschachtelte Strukturen
            config_df = pd.read_excel(xlsx_path, sheet_name='CONFIG', keep_default_na=False)
            config = {}
            
            for _, row in config_df.iterrows():
                key = str(row['Key'])
                sub_key = row['Sub-Key']
                sub_sub_key = row['Sub-Sub-Key']
                value = row['Value']
                
                # Einfache Key-Value Paare (keine Sub-Keys)
                if pd.isna(sub_key) or str(sub_key).strip() == '':
                    # Konvertiere Datentypen
                    if pd.notna(value):
                        value_str = str(value).strip().lower()
                        
                        # Boolean-Konvertierung
                        if value_str in ('true', '1', 'yes', 'ja'):
                            config[key] = True
                        elif value_str in ('false', '0', 'no', 'nein'):
                            config[key] = False
                        else:
                            # Versuche numerische Konvertierung
                            try:
                                if '.' in str(value):
                                    config[key] = float(value)
                                else:
                                    config[key] = int(value)
                            except (ValueError, TypeError):
                                config[key] = str(value)
                    else:
                        config[key] = None
                
                # Verschachtelte Strukturen
                else:
                    sub_key_str = str(sub_key)
                    
                    # Listen-Strukturen wie CODER_SETTINGS[0], CODER_SETTINGS[1]
                    if sub_key_str.startswith('[') and sub_key_str.endswith(']'):
                        index = int(sub_key_str[1:-1])
                        
                        if key not in config:
                            config[key] = []
                        
                        # Erweitere Liste falls nötig
                        while len(config[key]) <= index:
                            config[key].append({})
                        
                        # Füge Sub-Sub-Key hinzu
                        if pd.notna(sub_sub_key) and pd.notna(value):
                            sub_sub_key_str = str(sub_sub_key)
                            
                            # Konvertiere Wert
                            value_str = str(value).strip()
                            try:
                                if '.' in value_str:
                                    config[key][index][sub_sub_key_str] = float(value)
                                else:
                                    config[key][index][sub_sub_key_str] = int(value)
                            except (ValueError, TypeError):
                                config[key][index][sub_sub_key_str] = value_str
                    
                    # Dictionary-Strukturen wie ATTRIBUTE_LABELS, VALIDATION
                    else:
                        if key not in config:
                            config[key] = {}
                        
                        # Verschachtelte Dictionary-Strukturen (z.B. VALIDATION.ENGLISH_WORDS)
                        if pd.notna(sub_sub_key) and str(sub_sub_key).strip() != '':
                            if sub_key_str not in config[key]:
                                config[key][sub_key_str] = {}
                            
                            # Konvertiere Wert
                            if pd.notna(value):
                                value_str = str(value).strip().lower()
                                if value_str in ('true', '1', 'yes', 'ja'):
                                    config[key][sub_key_str][str(sub_sub_key)] = True
                                elif value_str in ('false', '0', 'no', 'nein'):
                                    config[key][sub_key_str][str(sub_sub_key)] = False
                                else:
                                    try:
                                        if '.' in str(value):
                                            config[key][sub_key_str][str(sub_sub_key)] = float(value)
                                        else:
                                            config[key][sub_key_str][str(sub_sub_key)] = int(value)
                                    except (ValueError, TypeError):
                                        config[key][sub_key_str][str(sub_sub_key)] = str(value)
                            else:
                                config[key][sub_key_str][str(sub_sub_key)] = None
                        
                        # Einfache Dictionary-Strukturen
                        else:
                            if pd.notna(value):
                                value_str = str(value).strip()
                                # Konvertiere None-String zu None
                                if value_str.lower() == 'none' or value_str == '':
                                    config[key][sub_key_str] = None
                                else:
                                    config[key][sub_key_str] = value_str
                            else:
                                config[key][sub_key_str] = None
            
            result['config'] = config
            
            return result
            
        except Exception as e:
            raise ValueError(f"Fehler beim Lesen der Excel-Datei: {str(e)}")
    
    @staticmethod
    def qca_aid_json_to_xlsx(json_data: Dict[str, Any], xlsx_path: str) -> None:
        """
        Schreibt QCA-AID JSON-Struktur in Excel-Datei
        
        Args:
            json_data: Dictionary mit 'forschungsfrage', 'kodierregeln', 'deduktive_kategorien', 'config'
            xlsx_path: Pfad zur Ziel-Excel-Datei
            
        Raises:
            ValueError: Wenn JSON-Struktur ungültig ist
        """
        required_keys = ['forschungsfrage', 'kodierregeln', 'deduktive_kategorien', 'config']
        for key in required_keys:
            if key not in json_data:
                raise ValueError(f"JSON muss '{key}' enthalten")
        
        try:
            wb = Workbook()
            
            # 1. Erstelle FORSCHUNGSFRAGE Sheet
            ws_forschungsfrage = wb.active
            ws_forschungsfrage.title = 'FORSCHUNGSFRAGE'
            ws_forschungsfrage.append(['Forschungsfrage:', json_data['forschungsfrage']])
            
            # 2. Erstelle KODIERREGELN Sheet
            ws_kodierregeln = wb.create_sheet(title='KODIERREGELN')
            ws_kodierregeln.append(['Allgemeine Kodierregeln', 'Formatregeln', 'Ausschlussregeln'])
            
            kodierregeln = json_data['kodierregeln']
            general = kodierregeln.get('general', [])
            format_rules = kodierregeln.get('format', [])
            exclusion = kodierregeln.get('exclusion', [])
            
            # Finde maximale Länge
            max_len = max(len(general), len(format_rules), len(exclusion))
            
            for i in range(max_len):
                row = [
                    general[i] if i < len(general) else '',
                    format_rules[i] if i < len(format_rules) else '',
                    exclusion[i] if i < len(exclusion) else ''
                ]
                ws_kodierregeln.append(row)
            
            # 3. Erstelle DEDUKTIVE_KATEGORIEN Sheet
            ws_kategorien = wb.create_sheet(title='DEDUKTIVE_KATEGORIEN')
            ws_kategorien.append(['Key', 'Sub-Key', 'Sub-Sub-Key', 'Value'])
            
            deduktive_kategorien = json_data['deduktive_kategorien']
            
            for key, kategorie in deduktive_kategorien.items():
                # Definition
                if 'definition' in kategorie and kategorie['definition']:
                    ws_kategorien.append([key, 'definition', '', kategorie['definition']])
                
                # Rules
                if 'rules' in kategorie:
                    for rule in kategorie['rules']:
                        ws_kategorien.append([key, 'rules', '', rule])
                
                # Subcategories
                if 'subcategories' in kategorie:
                    for sub_key, sub_value in kategorie['subcategories'].items():
                        ws_kategorien.append([key, 'subcategories', sub_key, sub_value])
                
                # Examples
                if 'examples' in kategorie:
                    for idx, example in enumerate(kategorie['examples']):
                        ws_kategorien.append([key, 'examples', f'[{idx}]', example])
            
            # 4. Erstelle CONFIG Sheet
            ws_config = wb.create_sheet(title='CONFIG')
            ws_config.append(['Key', 'Sub-Key', 'Sub-Sub-Key', 'Value', 'Erläuterung'])
            
            config = json_data['config']
            
            for key, value in config.items():
                # Einfache Werte
                if not isinstance(value, (dict, list)):
                    # Boolean-Konvertierung: JSON true/false → Excel 1/0
                    if isinstance(value, bool):
                        excel_value = 1 if value else 0
                    elif value is None:
                        excel_value = ''
                    else:
                        excel_value = value
                    
                    ws_config.append([key, '', '', excel_value, ''])
                
                # Listen (z.B. CODER_SETTINGS)
                elif isinstance(value, list):
                    for idx, item in enumerate(value):
                        if isinstance(item, dict):
                            for sub_key, sub_value in item.items():
                                ws_config.append([key, f'[{idx}]', sub_key, sub_value, ''])
                        else:
                            ws_config.append([key, f'[{idx}]', '', item, ''])
                
                # Dictionaries (z.B. ATTRIBUTE_LABELS, VALIDATION)
                elif isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        # Verschachtelte Dictionaries
                        if isinstance(sub_value, dict):
                            for sub_sub_key, sub_sub_value in sub_value.items():
                                # Boolean-Konvertierung
                                if isinstance(sub_sub_value, bool):
                                    excel_value = 1 if sub_sub_value else 0
                                elif sub_sub_value is None:
                                    excel_value = ''
                                else:
                                    excel_value = sub_sub_value
                                
                                ws_config.append([key, sub_key, sub_sub_key, excel_value, ''])
                        else:
                            # Einfache Dictionary-Werte
                            if sub_value is None:
                                excel_value = ''
                            else:
                                excel_value = sub_value
                            
                            ws_config.append([key, sub_key, '', excel_value, ''])
            
            # Speichere Workbook
            wb.save(xlsx_path)
            wb.close()
            
        except Exception as e:
            raise ValueError(f"Fehler beim Schreiben der Excel-Datei: {str(e)}")
