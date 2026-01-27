"""
Configuration Loader

Loads and manages QCA-AID configuration from Excel codebook files.
Handles category definitions, coding rules, settings validation, and sanitization.
"""

# Standard library imports
import os
from typing import Dict, Optional, Any

# Third-party imports
import pandas as pd
from openpyxl import load_workbook

# Local imports
from .converter import ConfigConverter
from .synchronizer import ConfigSynchronizer


class ConfigLoader:
    """
    Lädt und verwaltet die QCA-AID Konfiguration aus Excel-Codebüchern.
    
    Handles:
    - Loading research question from Excel
    - Loading coding rules
    - Loading deductive category definitions
    - Loading general configuration settings
    - Validation and sanitization of all loaded values
    - Boolean parameter handling
    - Directory path management
    
    The loader stores all values in a global_config dictionary which is
    updated throughout the loading process.
    """
    
    def __init__(self, script_dir: str, global_config: Dict) -> None:
        """
        Initialisiert den ConfigLoader.
        
        Args:
            script_dir: Directory where QCA-AID-Codebook.xlsx is located
            global_config: Dictionary to store loaded configuration (passed by reference)
        """
        self.script_dir = os.path.abspath(script_dir)
        self.excel_path = os.path.join(self.script_dir, "QCA-AID-Codebook.xlsx")
        self.json_path = os.path.join(self.script_dir, "QCA-AID-Codebook.json")
        self.config_path = None  # Will be set by _sync_configs()
        self.global_config = global_config  # Reference zum globalen CONFIG-Objekt
    
    def _sync_configs(self) -> None:
        """
        Führt Synchronisation zwischen XLSX und JSON durch.
        
        Bestimmt Pfade für XLSX und JSON basierend auf config_path.
        Prüft Existenz beider Dateien und ruft ConfigSynchronizer.sync() auf.
        Behandelt FileNotFoundError wenn beide fehlen.
        Aktualisiert config_path basierend auf Synchronisationsergebnis.
        
        Raises:
            FileNotFoundError: Wenn weder XLSX noch JSON existiert
        """
        xlsx_exists = os.path.exists(self.excel_path)
        json_exists = os.path.exists(self.json_path)
        
        # Fall 1: Beide Dateien fehlen
        if not xlsx_exists and not json_exists:
            raise FileNotFoundError(
                f"Weder XLSX noch JSON gefunden:\n"
                f"  XLSX: {self.excel_path}\n"
                f"  JSON: {self.json_path}"
            )
        
        # Fall 2: Nur eine Datei existiert - keine Synchronisation nötig
        if xlsx_exists and not json_exists:
            print(f"Nur XLSX gefunden: {self.excel_path}")
            self.config_path = self.excel_path
            return
        
        if json_exists and not xlsx_exists:
            print(f"Nur JSON gefunden: {self.json_path}")
            self.config_path = self.json_path
            return
        
        # Fall 3: Beide existieren - Synchronisation durchführen
        try:
            print("Beide Konfigurationsdateien gefunden. Prüfe Zeitstempel...")
            
            # Vergleiche Änderungszeitpunkte
            xlsx_mtime = os.path.getmtime(self.excel_path)
            json_mtime = os.path.getmtime(self.json_path)
            
            # Wenn XLSX neuer ist, aktualisiere JSON
            if xlsx_mtime > json_mtime:
                print("Excel-Datei ist neuer. Aktualisiere JSON...")
                json_data = ConfigConverter.qca_aid_xlsx_to_json(self.excel_path)
                ConfigConverter.save_json(json_data, self.json_path)
                print("JSON-Datei aktualisiert")
                self.config_path = self.json_path
            
            # Wenn JSON neuer ist, aktualisiere XLSX
            elif json_mtime > xlsx_mtime:
                print("JSON-Datei ist neuer. Aktualisiere Excel...")
                json_data = ConfigConverter.load_json(self.json_path)
                ConfigConverter.qca_aid_json_to_xlsx(json_data, self.excel_path)
                print("Excel-Datei aktualisiert")
                self.config_path = self.json_path
            
            # Wenn gleich alt, verwende JSON
            else:
                print("Beide Dateien haben gleichen Zeitstempel. Verwende JSON.")
                self.config_path = self.json_path
                
        except Exception as e:
            print(f"Warnung: Synchronisation fehlgeschlagen: {str(e)}")
            print("Verwende JSON-Datei als Fallback.")
            self.config_path = self.json_path
    
    def load_codebook(self) -> bool:
        """
        Lädt die komplette Konfiguration aus dem Codebook (JSON oder Excel).
        
        Ruft _sync_configs() vor dem Laden auf.
        Bestimmt JSON-Pfad basierend auf excel_path.
        Prüft ob JSON existiert und bevorzugt JSON.
        Falls JSON fehlt oder fehlschlägt, verwendet XLSX.
        Erstellt automatisch fehlende Datei nach erfolgreichem Laden.
        
        Returns:
            bool: True if loading successful, False otherwise
        """
        # Synchronisiere Konfigurationsdateien
        try:
            self._sync_configs()
        except FileNotFoundError as e:
            print(f"❌ {str(e)}")
            return False
        
        # Bestimme welche Datei geladen werden soll
        json_exists = os.path.exists(self.json_path)
        xlsx_exists = os.path.exists(self.excel_path)
        
        # Versuche JSON zu laden wenn vorhanden
        if json_exists:
            print(f"Versuche Konfiguration zu laden von: {self.json_path}")
            success = self._load_from_json(self.json_path)
            
            if success:
                # Erstelle XLSX wenn es fehlt
                if not xlsx_exists:
                    try:
                        print(f"Erstelle Excel-Datei aus JSON: {self.excel_path}")
                        json_data = ConfigConverter.load_json(self.json_path)
                        ConfigConverter.qca_aid_json_to_xlsx(json_data, self.excel_path)
                        print(f"✅ Excel-Datei erstellt: {self.excel_path}")
                    except Exception as e:
                        print(f"⚠️ Warnung: Konnte Excel-Datei nicht erstellen: {str(e)}")
                
                return True
            else:
                print("⚠️ JSON-Laden fehlgeschlagen, versuche Excel...")
        
        # Fallback zu Excel wenn JSON nicht existiert oder fehlgeschlagen ist
        if not xlsx_exists:
            print(f"Excel-Datei nicht gefunden: {self.excel_path}")
            return False
        
        print(f"Versuche Konfiguration zu laden von: {self.excel_path}")

        try:
            # FIX: Erweiterte Optionen für das Öffnen der Excel-Datei
            print("\nÖffne Excel-Datei...")
            wb = load_workbook(
                self.excel_path,
                read_only=True,
                data_only=True,
                keep_vba=False  # Verhindert Probleme mit VBA-Code
            )
            print(f"Excel-Datei erfolgreich geladen. verfügbare Sheets: {wb.sheetnames}")
            
            # Prüfe DEDUKTIVE_KATEGORIEN Sheet
            if 'DEDUKTIVE_KATEGORIEN' in wb.sheetnames:
                print("\nLese DEDUKTIVE_KATEGORIEN Sheet...")
                sheet = wb['DEDUKTIVE_KATEGORIEN']
            
            # Lade die verschiedenen Komponenten
            self._load_research_question(wb)
            self._load_coding_rules(wb)
            self._load_config(wb)
            
            # Debug-Ausgabe vor dem Laden der Kategorien
            print("\nStarte Laden der deduktiven Kategorien...")
            kategorien = self._load_deduktive_kategorien(wb)
            
            # Prüfe das Ergebnis
            if kategorien:
                # Speichere in Config
                self.global_config['DEDUKTIVE_KATEGORIEN'] = kategorien
                print("\nKategorien erfolgreich in Config gespeichert")
                
                # Erstelle JSON wenn es fehlt
                if not json_exists:
                    try:
                        print(f"Erstelle JSON-Datei aus Excel: {self.json_path}")
                        json_data = ConfigConverter.qca_aid_xlsx_to_json(self.excel_path)
                        ConfigConverter.save_json(json_data, self.json_path)
                        print(f"✅ JSON-Datei erstellt: {self.json_path}")
                    except Exception as e:
                        print(f"⚠️ Warnung: Konnte JSON-Datei nicht erstellen: {str(e)}")
                
                return True
            else:
                print("\nKeine Kategorien geladen!")
                return False

        except PermissionError:
            # Spezifische Behandlung für geöffnete Dateien
            print("\n⚠️ Datei ist bereits geöffnet (z.B. in Excel).")
            print("Versuche alternative Lademethode...")
            try:
                # Versuche mit keep_links=False für bessere Kompatibilität
                wb = load_workbook(
                    self.excel_path,
                    read_only=True,
                    data_only=True,
                    keep_vba=False,
                    keep_links=False  # Entfernt externe Links
                )
                
                # Lade die verschiedenen Komponenten
                self._load_research_question(wb)
                self._load_coding_rules(wb)
                self._load_config(wb)
                kategorien = self._load_deduktive_kategorien(wb)
                
                if kategorien:
                    self.global_config['DEDUKTIVE_KATEGORIEN'] = kategorien
                    print("\n✅ Konfiguration trotz geöffneter Datei erfolgreich geladen!")
                    
                    # Erstelle JSON wenn es fehlt
                    if not json_exists:
                        try:
                            print(f"Erstelle JSON-Datei aus Excel: {self.json_path}")
                            json_data = ConfigConverter.qca_aid_xlsx_to_json(self.excel_path)
                            ConfigConverter.save_json(json_data, self.json_path)
                            print(f"✅ JSON-Datei erstellt: {self.json_path}")
                        except Exception as e:
                            print(f"⚠️ Warnung: Konnte JSON-Datei nicht erstellen: {str(e)}")
                    
                    return True
                else:
                    print("\n‼️ Keine Kategorien geladen, auch bei alternativer Methode")
                    return False
                    
            except Exception as alt_e:
                print(f"\n‼️ Auch alternative Lademethode fehlgeschlagen: {str(alt_e)}")
                print("\nLösung: Schließen Sie die Excel-Datei und versuchen Sie erneut,")
                print("oder deaktivieren Sie das automatische Speichern in OneDrive.")
                return False

        except Exception as e:
            print(f"Fehler beim Lesen der Excel-Datei: {str(e)}")
            print("Details:")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_from_json(self, json_path: str) -> bool:
        """
        Lädt Konfiguration aus JSON-Datei.
        
        Liest JSON-Datei mit ConfigConverter.load_json() und extrahiert:
        - forschungsfrage → global_config
        - kodierregeln → global_config
        - deduktive_kategorien → CategoryDefinition Objekte
        - config-Parameter → global_config
        
        Ruft _sanitize_config() auf für Validierung.
        
        Args:
            json_path: Pfad zur JSON-Datei
            
        Returns:
            bool: True wenn erfolgreich geladen
        """
        print(f"Lade Konfiguration aus JSON: {json_path}")
        
        try:
            # Lade JSON-Datei
            json_data = ConfigConverter.load_json(json_path)
            
            # 1. Extrahiere forschungsfrage
            if 'forschungsfrage' in json_data:
                self.global_config['FORSCHUNGSFRAGE'] = json_data['forschungsfrage']
                print(f"Forschungsfrage geladen: {json_data['forschungsfrage'][:50]}...")
            
            # 2. Extrahiere kodierregeln
            if 'kodierregeln' in json_data:
                self.global_config['KODIERREGELN'] = json_data['kodierregeln']
                print(f"Kodierregeln geladen:")
                print(f"  - Allgemeine Regeln: {len(json_data['kodierregeln'].get('general', []))}")
                print(f"  - Formatregeln: {len(json_data['kodierregeln'].get('format', []))}")
                print(f"  - Ausschlussregeln: {len(json_data['kodierregeln'].get('exclusion', []))}")
            
            # 3. Extrahiere deduktive_kategorien und konvertiere zu CategoryDefinition Objekten
            if 'deduktive_kategorien' in json_data:
                from datetime import datetime
                from QCA_AID_assets.core.data_models import CategoryDefinition
                
                kategorien = {}
                for key, kategorie_data in json_data['deduktive_kategorien'].items():
                    # Erstelle CategoryDefinition Objekt
                    kategorien[key] = CategoryDefinition(
                        name=key,
                        definition=kategorie_data.get('definition', ''),
                        examples=kategorie_data.get('examples', []),
                        rules=kategorie_data.get('rules', []),
                        subcategories=kategorie_data.get('subcategories', {}),
                        added_date=datetime.now().strftime("%Y-%m-%d"),
                        modified_date=datetime.now().strftime("%Y-%m-%d")
                    )
                
                self.global_config['DEDUKTIVE_KATEGORIEN'] = kategorien
                print(f"✅ {len(kategorien)} Kategorien geladen")
            
            # 4. Extrahiere config-Parameter
            if 'config' in json_data:
                config = json_data['config']
                
                # Konvertiere snake_case Keys zu UPPER_CASE für Kompatibilität
                config_normalized = {}
                for key, value in config.items():
                    # Konvertiere z.B. "manual_coding_enabled" zu "MANUAL_CODING_ENABLED"
                    normalized_key = key.upper()
                    config_normalized[normalized_key] = value
                
                # Rufe _sanitize_config() auf für Validierung
                self._sanitize_config(config_normalized)
                
                # Update global_config mit Werten die nicht speziell behandelt wurden
                specially_handled_keys = {
                    'DATA_DIR', 'OUTPUT_DIR', 'INPUT_DIR',
                    'CHUNK_SIZE', 'CHUNK_OVERLAP',
                    'CODER_SETTINGS',
                    'CODE_WITH_CONTEXT',
                    'MULTIPLE_CODINGS',
                    'MULTIPLE_CODING_THRESHOLD',
                    'SIMILARITY_THRESHOLD',
                    'PDF_ANNOTATION_FUZZY_THRESHOLD',
                    'BATCH_SIZE',
                    'MANUAL_CODING_ENABLED'
                }
                
                for key, value in config_normalized.items():
                    if key not in specially_handled_keys:
                        self.global_config[key] = value
                
                # Explizit MANUAL_CODING_ENABLED setzen
                if 'MANUAL_CODING_ENABLED' in config_normalized:
                    self.global_config['MANUAL_CODING_ENABLED'] = bool(config_normalized['MANUAL_CODING_ENABLED'])
                    print(f"[CONFIG] MANUAL_CODING_ENABLED geladen: {self.global_config['MANUAL_CODING_ENABLED']}")
                
                print("Konfigurationsparameter geladen und validiert")
            
            return True
            
        except FileNotFoundError as e:
            print(f"❌ JSON-Datei nicht gefunden: {str(e)}")
            return False
        except ValueError as e:
            print(f"❌ Ungültiges JSON-Format: {str(e)}")
            return False
        except Exception as e:
            print(f"❌ Fehler beim Laden der JSON-Datei: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_research_question(self, wb) -> None:
        """Lädt die Forschungsfrage aus dem Excel-Codebook."""
        if 'FORSCHUNGSFRAGE' in wb.sheetnames:
            sheet = wb['FORSCHUNGSFRAGE']
            value = sheet['B1'].value
            self.global_config['FORSCHUNGSFRAGE'] = value

    def _load_coding_rules(self, wb) -> None:
        """Lädt Kodierregeln aus dem Excel-Codebook."""
        if 'KODIERREGELN' in wb.sheetnames:
            df = pd.read_excel(self.excel_path, sheet_name='KODIERREGELN', header=0)
            
            # Initialisiere Regelkategorien
            rules = {
                'general': [],       # Allgemeine Kodierregeln
                'format': [],        # Formatregeln
                'exclusion': []      # Ausschlussregeln
            }
            
            # Verarbeite jede Spalte
            for column in df.columns:
                rules_list = df[column].dropna().tolist()
                
                if 'Allgemeine' in column:
                    rules['general'].extend(rules_list)
                elif 'Format' in column:
                    rules['format'].extend(rules_list)
                elif 'Ausschluss' in column:
                    rules['exclusion'].extend(rules_list)
            
            print("\nKodierregeln geladen:")
            print(f"- Allgemeine Regeln: {len(rules['general'])}")
            print(f"- Formatregeln: {len(rules['format'])}")
            print(f"- Ausschlussregeln: {len(rules['exclusion'])}")
            
            self.global_config['KODIERREGELN'] = rules

    def _load_deduktive_kategorien(self, wb) -> Dict:
        """
        Lädt deduktive Kategorien mit Hierarchie aus dem Excel-Codebook.
        
        Supports multi-level hierarchy:
        - Key: Hauptkategorie
        - Sub-Key: definition, rules, examples, subcategories
        - Sub-Sub-Key: Unterkategorie-Namen
        
        Returns:
            Dictionary with category structure or empty dict on error
        """
        try:
            if 'DEDUKTIVE_KATEGORIEN' not in wb.sheetnames:
                print("Warnung: Sheet 'DEDUKTIVE_KATEGORIEN' nicht gefunden")
                return {}

            print("\nLade deduktive Kategorien...")
            sheet = wb['DEDUKTIVE_KATEGORIEN']
            
            # Initialisiere Kategorien
            kategorien = {}
            current_category = None
            
            # Hole Header-Zeile
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Indizes für Spalten finden
            key_idx = headers.index('Key') if 'Key' in headers else None
            sub_key_idx = headers.index('Sub-Key') if 'Sub-Key' in headers else None
            sub_sub_key_idx = headers.index('Sub-Sub-Key') if 'Sub-Sub-Key' in headers else None
            value_idx = headers.index('Value') if 'Value' in headers else None
            
            if None in [key_idx, sub_key_idx, value_idx]:
                print("Fehler: Erforderliche Spalten fehlen!")
                return {}
                
            # Verarbeite Zeilen
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                try:
                    key = row[key_idx].value
                    sub_key = row[sub_key_idx].value if row[sub_key_idx].value else None
                    sub_sub_key = row[sub_sub_key_idx].value if sub_sub_key_idx is not None and row[sub_sub_key_idx].value else None
                    value = row[value_idx].value if row[value_idx].value else None
                    
                    # Neue Hauptkategorie
                    if key and isinstance(key, str):
                        key = key.strip()
                        if key not in kategorien:
                            current_category = key
                            kategorien[key] = {
                                'definition': '',
                                'rules': [],
                                'examples': [],
                                'subcategories': {}
                            }
                    
                    # Verarbeite Unterkategorien und Werte
                    if current_category and sub_key:
                        sub_key = sub_key.strip()
                        if isinstance(value, str):
                            value = value.strip()
                            
                        if sub_key == 'definition':
                            kategorien[current_category]['definition'] = value
                            
                        elif sub_key == 'rules':
                            if value:
                                kategorien[current_category]['rules'].append(value)
                                
                        elif sub_key == 'examples':
                            if value:
                                kategorien[current_category]['examples'].append(value)
                                
                        elif sub_key == 'subcategories' and sub_sub_key:
                            # FIX: Auch leere Werte akzeptieren (für Subcategories ohne Beschreibung)
                            if value is None:
                                value = ""  # Leere Beschreibung
                            kategorien[current_category]['subcategories'][sub_sub_key] = value
                                
                except Exception as e:
                    print(f"Fehler in Zeile {row_idx}: {str(e)}")
                    continue

            # Ergebnis
            if kategorien:
                print(f"Erfolgreich {len(kategorien)} Kategorien geladen")
                return kategorien
            else:
                print("Keine Kategorien gefunden!")
                return {}

        except Exception as e:
            print(f"Fehler beim Laden der Kategorien: {str(e)}")
            print("Details:")
            import traceback
            traceback.print_exc()
            return {}
    
    def _load_config(self, wb) -> bool:
        """
        Lädt allgemeine Konfigurationswerte aus dem CONFIG Sheet.
        
        Handles:
        - Nested dictionary structures (Key -> Sub-Key -> Sub-Sub-Key)
        - List structures for CODER_SETTINGS
        - Boolean parameter conversion
        - Analysis mode and review mode validation
        
        Returns:
            bool: True if config loaded, False if CONFIG sheet not found
        """
        if 'CONFIG' not in wb.sheetnames:
            return False
            
        df = pd.read_excel(self.excel_path, sheet_name='CONFIG')
        config = {}
        
        for _, row in df.iterrows():
            key = row['Key']
            sub_key = row['Sub-Key']
            sub_sub_key = row['Sub-Sub-Key']
            value = row['Value']

            if key not in config:
                config[key] = value if pd.isna(sub_key) else {}

            if not pd.isna(sub_key):
                if sub_key.startswith('['):  # Für Listen wie CODER_SETTINGS
                    if not isinstance(config[key], list):
                        config[key] = []
                    index = int(sub_key.strip('[]'))
                    while len(config[key]) <= index:
                        config[key].append({})
                    if pd.isna(sub_sub_key):
                        config[key][index] = value
                    else:
                        config[key][index][sub_sub_key] = value
                else:  # Für verschachtelte Dicts
                    if not isinstance(config[key], dict):
                        config[key] = {}
                    if pd.isna(sub_sub_key):
                        # Extended Boolean handling für alle relevanten Parameter
                        if sub_key in ['EXPORT_ANNOTATED_PDFS', 'CODE_WITH_CONTEXT', 'SAVE_PROGRESS', 'PARALLEL_PROCESSING', 'MULTIPLE_CODINGS', 'MANUAL_CODING_ENABLED']:
                            if isinstance(value, str):
                                config[key][sub_key] = value.lower() in ['true', '1', 'yes', 'ja', 'on', 'wahr']
                            elif isinstance(value, (int, float)):
                                config[key][sub_key] = bool(value) and value != 0
                            else:
                                config[key][sub_key] = bool(value)
                            print(f"[CONFIG] {sub_key} geladen: {config[key][sub_key]} (ursprünglich: '{value}')")
                        elif sub_key in ['BATCH_SIZE', 'CHUNK_SIZE', 'CHUNK_OVERLAP', 'MAX_RETRIES'] or key == 'BATCH_SIZE':
                            try:
                                config[key][sub_key] = int(value)
                                print(f"[CONFIG] {sub_key} geladen: {config[key][sub_key]}")
                            except (ValueError, TypeError):
                                default_values = {'BATCH_SIZE': 5, 'CHUNK_SIZE': 2000, 'CHUNK_OVERLAP': 200, 'MAX_RETRIES': 3}
                                config[key][sub_key] = default_values.get(sub_key, 5)
                                print(f"[CONFIG] Ungültiger {sub_key} Wert '{value}', verwende Standard: {config[key][sub_key]}")
                        elif sub_key in ['MULTIPLE_CODING_THRESHOLD', 'SIMILARITY_THRESHOLD', 'PDF_ANNOTATION_FUZZY_THRESHOLD']:
                            try:
                                config[key][sub_key] = float(value)
                                print(f"[CONFIG] {sub_key} geladen: {config[key][sub_key]}")
                            except (ValueError, TypeError):
                                default_values = {'MULTIPLE_CODING_THRESHOLD': 0.85, 'SIMILARITY_THRESHOLD': 0.7, 'PDF_ANNOTATION_FUZZY_THRESHOLD': 0.85}
                                config[key][sub_key] = default_values.get(sub_key, 0.85)
                                print(f"[CONFIG] Ungültiger {sub_key} Wert '{value}', verwende Standard: {config[key][sub_key]}")
                        else:
                            config[key][sub_key] = value
                    else:
                        if sub_key not in config[key]:
                            config[key][sub_key] = {}
                        config[key][sub_key][sub_sub_key] = value

            # Top-level Boolean parameters
            if key == 'EXPORT_ANNOTATED_PDFS' and pd.isna(sub_key):
                if isinstance(value, str):
                    config['EXPORT_ANNOTATED_PDFS'] = value.lower() in ['true', '1', 'yes', 'ja', 'on', 'wahr']
                elif isinstance(value, (int, float)):
                    config['EXPORT_ANNOTATED_PDFS'] = bool(value) and value != 0
                else:
                    config['EXPORT_ANNOTATED_PDFS'] = bool(value)
                print(f"[CONFIG] EXPORT_ANNOTATED_PDFS geladen: {config['EXPORT_ANNOTATED_PDFS']}")

        # Validierung: ANALYSIS_MODE
        if 'ANALYSIS_MODE' in config:
            valid_modes = {'abductive', 'deductive', 'inductive', 'grounded'}
            if config['ANALYSIS_MODE'] not in valid_modes:
                print(f"[CONFIG] Warnung: Ungültiger ANALYSIS_MODE '{config['ANALYSIS_MODE']}'. Verwende 'deductive'.")
                config['ANALYSIS_MODE'] = 'deductive'
            else:
                print(f"[CONFIG] ANALYSIS_MODE geladen: {config['ANALYSIS_MODE']}")
        else:
            config['ANALYSIS_MODE'] = 'deductive'
            print(f"[CONFIG] ANALYSIS_MODE nicht gefunden, verwende Standard: deductive")

        # Validierung: REVIEW_MODE
        if 'REVIEW_MODE' in config:
            valid_modes = {'auto', 'manual', 'consensus', 'majority'}
            if config['REVIEW_MODE'] not in valid_modes:
                print(f"[CONFIG] Warnung: Ungültiger REVIEW_MODE '{config['REVIEW_MODE']}'. Verwende 'consensus'.")
                config['REVIEW_MODE'] = 'consensus'
            else:
                print(f"[CONFIG] REVIEW_MODE geladen: {config['REVIEW_MODE']}")
        else:
            config['REVIEW_MODE'] = 'consensus'
            print(f"[CONFIG] REVIEW_MODE nicht gefunden, verwende Standard: consensus")

        # Stelle sicher, dass ATTRIBUTE_LABELS vorhanden ist
        if 'ATTRIBUTE_LABELS' not in config:
            config['ATTRIBUTE_LABELS'] = {'attribut1': 'Attribut1', 'attribut2': 'Attribut2', 'attribut3': 'Attribut3'}
        elif isinstance(config.get('ATTRIBUTE_LABELS'), dict) and 'attribut3' not in config['ATTRIBUTE_LABELS']:
            config['ATTRIBUTE_LABELS']['attribut3'] = 'Attribut3'

        # Extract top-level booleans from nested structure
        for key in ['SETTINGS', 'OPTIONS', 'GENERAL']:
            if key in config and isinstance(config[key], dict):
                for param_name in ['CODE_WITH_CONTEXT', 'MULTIPLE_CODINGS', 'PARALLEL_PROCESSING', 'SAVE_PROGRESS', 'MANUAL_CODING_ENABLED']:
                    if param_name in config[key]:
                        config[param_name] = config[key][param_name]
                        print(f"[CONFIG] Extrahierte {param_name} aus {key}: {config[param_name]}")
                # Also extract threshold values
                for threshold_name in ['MULTIPLE_CODING_THRESHOLD', 'SIMILARITY_THRESHOLD', 'PDF_ANNOTATION_FUZZY_THRESHOLD']:
                    if threshold_name in config[key]:
                        config[threshold_name] = config[key][threshold_name]
                        print(f"[CONFIG] Extrahierte {threshold_name} aus {key}: {config[threshold_name]}")
                # Extract ATTRIBUTE_LABELS if nested
                if 'ATTRIBUTE_LABELS' in config[key]:
                    config['ATTRIBUTE_LABELS'] = config[key]['ATTRIBUTE_LABELS']
                    print(f"[CONFIG] Extrahierte ATTRIBUTE_LABELS aus {key}: {config['ATTRIBUTE_LABELS']}")
                # Extract CODER_SETTINGS if nested
                if 'CODER_SETTINGS' in config[key]:
                    config['CODER_SETTINGS'] = config[key]['CODER_SETTINGS']
                    print(f"[CONFIG] Extrahierte CODER_SETTINGS aus {key}: {len(config['CODER_SETTINGS'])} Coder")

        
        self._sanitize_config(config)
        
        # Update global_config with values that weren't specially handled in _sanitize_config
        # Only add keys that aren't already set by _sanitize_config
        specially_handled_keys = {
            'DATA_DIR', 'OUTPUT_DIR', 'INPUT_DIR',
            'CHUNK_SIZE', 'CHUNK_OVERLAP',
            'CODER_SETTINGS',
            'CODE_WITH_CONTEXT',
            'MULTIPLE_CODINGS',
            'MULTIPLE_CODING_THRESHOLD',
            'SIMILARITY_THRESHOLD',
            'PDF_ANNOTATION_FUZZY_THRESHOLD',
        }
        
        for key, value in config.items():
            if key not in specially_handled_keys and key not in self.global_config:
                self.global_config[key] = value
        
        return True

    def _sanitize_config(self, loaded_config: Dict) -> None:
        """
        Bereinigt und validiert Konfigurationswerte.
        
        Handles:
        - Directory path resolution and creation (Requirements 7.1, 7.2, 7.4, 7.5)
        - Numeric value conversion and validation (Requirements 8.1, 8.2, 8.3, 8.4, 8.5)
        - Enum parameter validation (Requirements 9.1, 9.2, 9.3, 9.4)
        - CODER_SETTINGS type conversion
        - Parameter constraint checking
        - Chunk size validation
        
        Args:
            loaded_config: Dictionary with raw configuration values from codebook
        """
        try:
            # Use the directory containing the configuration files as project root
            project_root = self.script_dir
            
            # Default values for validation
            DEFAULT_VALUES = {
                'CHUNK_SIZE': 1200,
                'CHUNK_OVERLAP': 50,
                'BATCH_SIZE': 8,
                'MULTIPLE_CODING_THRESHOLD': 0.85,
                'SIMILARITY_THRESHOLD': 0.7,
                'PDF_ANNOTATION_FUZZY_THRESHOLD': 0.85,
                'ANALYSIS_MODE': 'deductive',
                'REVIEW_MODE': 'consensus'
            }
            
            # ===== SUBTASK 3.3: Pfadverwaltung =====
            # Unterscheide zwischen relativen und absoluten Pfaden
            # Löse relative Pfade relativ zum Projektverzeichnis auf
            # Verwende absolute Pfade direkt
            # Erstelle nicht-existierende Verzeichnisse
            for key in ['DATA_DIR', 'OUTPUT_DIR', 'INPUT_DIR']:
                if key in loaded_config:
                    path_value = str(loaded_config[key]).strip()
                    
                    if os.path.isabs(path_value):
                        # Absolute path - use directly
                        resolved_path = path_value
                    else:
                        # Relative path - relative to selected project root
                        resolved_path = os.path.join(project_root, path_value.lstrip('/\\'))
                    
                    # Erstelle Verzeichnis wenn nicht vorhanden
                    os.makedirs(resolved_path, exist_ok=True)
                    self.global_config[key] = resolved_path
                    print(f"🩹 Verzeichnis {key}: {resolved_path}")
            
            # ===== SUBTASK 3.1: Numerische Validierung =====
            
            # CHUNK_SIZE: Prüfe >= 1, verwende Standard bei ungültig
            if 'CHUNK_SIZE' in loaded_config:
                try:
                    chunk_size = int(loaded_config['CHUNK_SIZE'])
                    if chunk_size < 1:
                        print(f"⚠️ Warnung: CHUNK_SIZE muss >= 1 sein. Wert '{chunk_size}' ist ungültig.")
                        print(f"   Verwende Standardwert: {DEFAULT_VALUES['CHUNK_SIZE']}")
                        self.global_config['CHUNK_SIZE'] = DEFAULT_VALUES['CHUNK_SIZE']
                    else:
                        self.global_config['CHUNK_SIZE'] = chunk_size
                        print(f"🩹 CHUNK_SIZE = {chunk_size}")
                except (ValueError, TypeError) as e:
                    print(f"⚠️ Warnung: Ungültiger CHUNK_SIZE Wert '{loaded_config['CHUNK_SIZE']}'.")
                    print(f"   Verwende Standardwert: {DEFAULT_VALUES['CHUNK_SIZE']}")
                    self.global_config['CHUNK_SIZE'] = DEFAULT_VALUES['CHUNK_SIZE']
            else:
                self.global_config['CHUNK_SIZE'] = DEFAULT_VALUES['CHUNK_SIZE']
            
            # CHUNK_OVERLAP: Prüfe < CHUNK_SIZE, korrigiere bei ungültig
            if 'CHUNK_OVERLAP' in loaded_config:
                try:
                    chunk_overlap = int(loaded_config['CHUNK_OVERLAP'])
                    if chunk_overlap < 0:
                        print(f"⚠️ Warnung: CHUNK_OVERLAP muss >= 0 sein. Wert '{chunk_overlap}' ist ungültig.")
                        print(f"   Verwende Standardwert: {DEFAULT_VALUES['CHUNK_OVERLAP']}")
                        self.global_config['CHUNK_OVERLAP'] = DEFAULT_VALUES['CHUNK_OVERLAP']
                    else:
                        self.global_config['CHUNK_OVERLAP'] = chunk_overlap
                        print(f"🩹 CHUNK_OVERLAP = {chunk_overlap}")
                except (ValueError, TypeError):
                    print(f"⚠️ Warnung: Ungültiger CHUNK_OVERLAP Wert '{loaded_config['CHUNK_OVERLAP']}'.")
                    print(f"   Verwende Standardwert: {DEFAULT_VALUES['CHUNK_OVERLAP']}")
                    self.global_config['CHUNK_OVERLAP'] = DEFAULT_VALUES['CHUNK_OVERLAP']
            else:
                self.global_config['CHUNK_OVERLAP'] = DEFAULT_VALUES['CHUNK_OVERLAP']
            
            # Stelle sicher dass CHUNK_OVERLAP < CHUNK_SIZE
            if self.global_config['CHUNK_OVERLAP'] >= self.global_config['CHUNK_SIZE']:
                print(f"⚠️ Warnung: CHUNK_OVERLAP ({self.global_config['CHUNK_OVERLAP']}) >= CHUNK_SIZE ({self.global_config['CHUNK_SIZE']}).")
                corrected_overlap = max(1, self.global_config['CHUNK_SIZE'] // 10)
                print(f"   Korrigiere CHUNK_OVERLAP auf: {corrected_overlap}")
                self.global_config['CHUNK_OVERLAP'] = corrected_overlap
            
            # BATCH_SIZE: Prüfe zwischen 1-20, verwende Standard bei ungültig
            if 'BATCH_SIZE' in loaded_config:
                try:
                    batch_size = int(loaded_config['BATCH_SIZE'])
                    if batch_size < 1:
                        print(f"⚠️ Warnung: BATCH_SIZE muss >= 1 sein. Wert '{batch_size}' ist ungültig.")
                        print(f"   Verwende Standardwert: {DEFAULT_VALUES['BATCH_SIZE']}")
                        self.global_config['BATCH_SIZE'] = DEFAULT_VALUES['BATCH_SIZE']
                    elif batch_size > 20:
                        print(f"⚠️ Warnung: BATCH_SIZE > 20 könnte Performance-Probleme verursachen. Wert '{batch_size}' ist zu hoch.")
                        print(f"   Verwende Standardwert: {DEFAULT_VALUES['BATCH_SIZE']}")
                        self.global_config['BATCH_SIZE'] = DEFAULT_VALUES['BATCH_SIZE']
                    else:
                        self.global_config['BATCH_SIZE'] = batch_size
                        print(f"🩹 BATCH_SIZE = {batch_size}")
                except (ValueError, TypeError):
                    print(f"⚠️ Warnung: Ungültiger BATCH_SIZE Wert '{loaded_config['BATCH_SIZE']}'.")
                    print(f"   Verwende Standardwert: {DEFAULT_VALUES['BATCH_SIZE']}")
                    self.global_config['BATCH_SIZE'] = DEFAULT_VALUES['BATCH_SIZE']
            else:
                self.global_config['BATCH_SIZE'] = DEFAULT_VALUES['BATCH_SIZE']
            
            # Float-Thresholds: Prüfe zwischen 0.0-1.0, verwende Standard bei ungültig
            float_thresholds = {
                'MULTIPLE_CODING_THRESHOLD': DEFAULT_VALUES['MULTIPLE_CODING_THRESHOLD'],
                'SIMILARITY_THRESHOLD': DEFAULT_VALUES['SIMILARITY_THRESHOLD'],
                'PDF_ANNOTATION_FUZZY_THRESHOLD': DEFAULT_VALUES['PDF_ANNOTATION_FUZZY_THRESHOLD']
            }
            
            for threshold_name, default_value in float_thresholds.items():
                if threshold_name in loaded_config:
                    try:
                        threshold = float(loaded_config[threshold_name])
                        if threshold < 0.0 or threshold > 1.0:
                            print(f"⚠️ Warnung: {threshold_name} muss zwischen 0.0 und 1.0 liegen. Wert '{threshold}' ist ungültig.")
                            print(f"   Verwende Standardwert: {default_value}")
                            self.global_config[threshold_name] = default_value
                        else:
                            self.global_config[threshold_name] = threshold
                            print(f"🩹 {threshold_name} = {threshold}")
                    except (ValueError, TypeError):
                        print(f"⚠️ Warnung: Ungültiger {threshold_name} Wert '{loaded_config[threshold_name]}'.")
                        print(f"   Verwende Standardwert: {default_value}")
                        self.global_config[threshold_name] = default_value
                else:
                    self.global_config[threshold_name] = default_value
            
            # ===== SUBTASK 3.2: Enum-Validierung =====
            
            # ANALYSIS_MODE: Prüfe gegen gültige Werte
            valid_analysis_modes = {'abductive', 'deductive', 'inductive', 'grounded'}
            if 'ANALYSIS_MODE' in loaded_config:
                analysis_mode = loaded_config['ANALYSIS_MODE']
                if analysis_mode not in valid_analysis_modes:
                    print(f"⚠️ Warnung: Ungültiger ANALYSIS_MODE '{analysis_mode}'.")
                    print(f"   Gültige Werte: {valid_analysis_modes}")
                    print(f"   Verwende Standardwert: '{DEFAULT_VALUES['ANALYSIS_MODE']}'")
                    self.global_config['ANALYSIS_MODE'] = DEFAULT_VALUES['ANALYSIS_MODE']
                else:
                    self.global_config['ANALYSIS_MODE'] = analysis_mode
                    print(f"🩹 ANALYSIS_MODE = {analysis_mode}")
            else:
                self.global_config['ANALYSIS_MODE'] = DEFAULT_VALUES['ANALYSIS_MODE']
            
            # REVIEW_MODE: Prüfe gegen gültige Werte
            valid_review_modes = {'auto', 'manual', 'consensus', 'majority'}
            if 'REVIEW_MODE' in loaded_config:
                review_mode = loaded_config['REVIEW_MODE']
                if review_mode not in valid_review_modes:
                    print(f"⚠️ Warnung: Ungültiger REVIEW_MODE '{review_mode}'.")
                    print(f"   Gültige Werte: {valid_review_modes}")
                    print(f"   Verwende Standardwert: '{DEFAULT_VALUES['REVIEW_MODE']}'")
                    self.global_config['REVIEW_MODE'] = DEFAULT_VALUES['REVIEW_MODE']
                else:
                    self.global_config['REVIEW_MODE'] = review_mode
                    print(f"🩹 REVIEW_MODE = {review_mode}")
            else:
                self.global_config['REVIEW_MODE'] = DEFAULT_VALUES['REVIEW_MODE']
            
            # ===== Weitere Validierungen =====
            
            # Coder settings with type conversion
            if 'CODER_SETTINGS' in loaded_config:
                self.global_config['CODER_SETTINGS'] = [
                    {
                        'temperature': float(coder['temperature'])
                            if isinstance(coder.get('temperature'), (int, float, str))
                            else 0.3,
                        'coder_id': str(coder.get('coder_id', f'auto_{i}'))
                    }
                    for i, coder in enumerate(loaded_config['CODER_SETTINGS'])
                ]
                print(f"🩹 CODER_SETTINGS geladen: {len(self.global_config['CODER_SETTINGS'])} Coder")
                for i, coder in enumerate(self.global_config['CODER_SETTINGS']):
                    print(f"   Coder {i}: ID='{coder['coder_id']}', Temp={coder['temperature']}")
            
            # CODE_WITH_CONTEXT handling
            if 'CODE_WITH_CONTEXT' in loaded_config:
                value = loaded_config['CODE_WITH_CONTEXT']
                if isinstance(value, str):
                    self.global_config['CODE_WITH_CONTEXT'] = value.lower() in ('true', 'ja', 'yes', '1')
                else:
                    self.global_config['CODE_WITH_CONTEXT'] = bool(value)
                print(f"🩹 CODE_WITH_CONTEXT = {self.global_config['CODE_WITH_CONTEXT']}")
            else:
                self.global_config['CODE_WITH_CONTEXT'] = False

            # MULTIPLE_CODINGS handling
            if 'MULTIPLE_CODINGS' in loaded_config:
                value = loaded_config['MULTIPLE_CODINGS']
                if isinstance(value, str):
                    self.global_config['MULTIPLE_CODINGS'] = value.lower() in ('true', 'ja', 'yes', '1')
                else:
                    self.global_config['MULTIPLE_CODINGS'] = bool(value)
                print(f"🩹 MULTIPLE_CODINGS = {self.global_config['MULTIPLE_CODINGS']}")
            else:
                self.global_config['MULTIPLE_CODINGS'] = True
            
            # MANUAL_CODING_ENABLED handling
            if 'MANUAL_CODING_ENABLED' in loaded_config:
                value = loaded_config['MANUAL_CODING_ENABLED']
                if isinstance(value, str):
                    self.global_config['MANUAL_CODING_ENABLED'] = value.lower() in ('true', 'ja', 'yes', '1')
                else:
                    self.global_config['MANUAL_CODING_ENABLED'] = bool(value)
                print(f"🩹 MANUAL_CODING_ENABLED = {self.global_config['MANUAL_CODING_ENABLED']}")
            else:
                self.global_config['MANUAL_CODING_ENABLED'] = False
            
            # Pass through all other values
            for key, value in loaded_config.items():
                if key not in self.global_config:
                    self.global_config[key] = value
            
            # Debug output for ATTRIBUTE_LABELS
            if 'ATTRIBUTE_LABELS' in self.global_config:
                print(f"🩹 ATTRIBUTE_LABELS geladen: {self.global_config['ATTRIBUTE_LABELS']}")
            else:
                print("⚠️ Warnung: ATTRIBUTE_LABELS nicht gefunden in loaded_config")
                # Set default
                self.global_config['ATTRIBUTE_LABELS'] = {
                    'attribut1': 'Attribut1',
                    'attribut2': 'Attribut2',
                    'attribut3': 'Attribut3'
                }
                print(f"   Verwende Standard-ATTRIBUTE_LABELS: {self.global_config['ATTRIBUTE_LABELS']}")
                    
        except Exception as e:
            print(f"🩹 Fehler: {str(e)}")
            import traceback
            traceback.print_exc()

    def get_config(self) -> Dict:
        """
        Returns the global configuration dictionary.
        
        Returns:
            Dictionary containing all loaded and sanitized configuration
        """
        return self.global_config
